"""
Парсер Orizhka.ru - магазин запчастей для Apple
API: Tilda Store (store.tildaapi.com)
БД: db_orizhka
"""

import httpx
import json
import time
import os
import re
import sys
import argparse
import psycopg2
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# === КОНФИГУРАЦИЯ ===

BASE_URL = "https://orizhka.ru"
API_URL = "https://store.tildaapi.com/api/getproductslist/"

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
PRODUCTS_JSON = os.path.join(DATA_DIR, "products.json")
PRODUCTS_XLSX = os.path.join(DATA_DIR, "products.xlsx")

REQUEST_DELAY = 0.3
REQUEST_TIMEOUT = 30
PER_PAGE = 50

# === КОНФИГУРАЦИЯ БД (Supabase) ===
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db_wrapper import get_db  # Автоматически маппит таблицы на новые имена

# Категории (storepart ID -> название)
CATEGORIES = {
    # iPhone 7/7+
    "346541913282": "iPhone 7",
    "776914721572": "iPhone 7 Plus",
    # iPhone 8/8+
    "553331889302": "iPhone 8",
    "381404029652": "iPhone 8 Plus",
    # iPhone X/XS/XR
    "133757333162": "iPhone X",
    "512547526732": "iPhone XS",
    "894668999552": "iPhone XS Max",
    "647621281332": "iPhone XR",
    # iPhone 11
    "685529849602": "iPhone 11",
    "270513488312": "iPhone 11 Pro",
    "457325399132": "iPhone 11 Pro Max",
    # iPhone 12
    "861463486052": "iPhone 12",
    "601615617652": "iPhone 12 Mini",
    "641694305782": "iPhone 12 Pro",
    "646033160842": "iPhone 12 Pro Max",
    # iPhone 13
    "699239210582": "iPhone 13",
    "542123349722": "iPhone 13 Mini",
    "525005961452": "iPhone 13 Pro",
    "200559278262": "iPhone 13 Pro Max",
    # iPhone 14
    "322696400962": "iPhone 14",
    "546796606802": "iPhone 14 Plus",
    "211493978242": "iPhone 14 Pro",
    "844936907582": "iPhone 14 Pro Max",
    # iPhone 15
    "960390868672": "iPhone 15",
    "550674235482": "iPhone 15 Plus",
    "748555222622": "iPhone 15 Pro",
    "734533983092": "iPhone 15 Pro Max",
    # iPhone 16
    "292850428202": "iPhone 16",
    "568380906832": "iPhone 16 Plus",
    "853707969302": "iPhone 16 Pro",
    "988341629032": "iPhone 16 Pro Max",
    # iPhone 17
    "872337508192": "iPhone 17 Air",
    # iPhone SE
    "756534640782": "iPhone SE 2020",
    # Apple Watch
    "458225102632": "Apple Watch",
    # iPad
    "869303579112": "iPad",
    # MacBook
    "345714019292": "MacBook",
    # Аксессуары
    "175877187102": "Проклейки Apple Watch",
    "449856930902": "Проклейки iPhone",
    "483534211232": "Проклейки разные",
    "610416219912": "Инструменты JCID",
    "875620649822": "Платы SWAP",
}


def init_db():
    """Создать таблицы если не существуют"""
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id SERIAL PRIMARY KEY,
            product_id VARCHAR(50) UNIQUE NOT NULL,
            sku VARCHAR(100),
            name TEXT NOT NULL,
            price NUMERIC(10,2),
            old_price NUMERIC(10,2),
            availability INTEGER DEFAULT 0,
            category VARCHAR(200),
            brand VARCHAR(100) DEFAULT 'Apple',
            url TEXT,
            city_id INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS cities (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            slug VARCHAR(100) UNIQUE,
            region VARCHAR(200)
        );

        CREATE INDEX IF NOT EXISTS idx_products_category ON products(category);
        CREATE INDEX IF NOT EXISTS idx_products_updated ON products(updated_at);
    """)

    # Добавляем город по умолчанию
    cur.execute("""
        INSERT INTO cities (id, name, slug, region)
        VALUES (1, 'Санкт-Петербург', 'spb', 'Ленинградская область')
        ON CONFLICT (slug) DO NOTHING
    """)

    conn.commit()
    cur.close()
    conn.close()
    print("[DB] Таблицы инициализированы")


@dataclass
class Product:
    """Товар"""
    uid: str
    title: str
    sku: str
    price: float
    price_old: float
    quantity: int
    url: str
    category: str
    description: str = ""

    def to_dict(self) -> dict:
        return {
            "uid": self.uid,
            "title": self.title,
            "sku": self.sku,
            "price": self.price,
            "price_old": self.price_old,
            "quantity": self.quantity,
            "url": self.url,
            "category": self.category,
        }


class OrizhkaParser:
    """Парсер каталога Orizhka.ru"""

    def __init__(self, save_to_db: bool = True):
        self.client = httpx.Client(
            timeout=REQUEST_TIMEOUT,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36",
                "Accept": "application/json",
                "Accept-Language": "ru,en;q=0.9",
                "Referer": "https://orizhka.ru/catalog",
            },
            follow_redirects=True,
        )
        self.products: List[Product] = []
        self.seen_uids: set = set()
        self.errors: List[Dict] = []
        self.last_request = 0
        self.save_to_db = save_to_db

        os.makedirs(DATA_DIR, exist_ok=True)

    def _rate_limit(self):
        """Ограничение частоты запросов"""
        elapsed = time.time() - self.last_request
        if elapsed < REQUEST_DELAY:
            time.sleep(REQUEST_DELAY - elapsed)
        self.last_request = time.time()

    def get_products(self, storepart_uid: str, category_name: str) -> List[Product]:
        """Получить товары из категории"""
        products = []
        slice_num = 1

        while True:
            self._rate_limit()

            params = {
                "storepartuid": storepart_uid,
                "getparts": "true",
                "getoptions": "true",
                "slice": str(slice_num),
                "size": str(PER_PAGE),
            }

            try:
                response = self.client.get(API_URL, params=params)
                response.raise_for_status()
                data = response.json()
            except Exception as e:
                self.errors.append({
                    "category": category_name,
                    "storepart": storepart_uid,
                    "slice": slice_num,
                    "error": str(e),
                    "time": datetime.now().isoformat()
                })
                print(f"  [ERROR] {e}")
                break

            items = data.get("products", [])
            total = data.get("total", 0)

            if not items:
                break

            for item in items:
                uid = str(item.get("uid", ""))
                if uid in self.seen_uids:
                    continue
                self.seen_uids.add(uid)

                price_str = item.get("price", "0")
                try:
                    price = float(price_str.replace(",", ".").replace(" ", ""))
                except:
                    price = 0.0

                price_old_str = item.get("priceold", "") or "0"
                try:
                    price_old = float(price_old_str.replace(",", ".").replace(" ", ""))
                except:
                    price_old = 0.0

                qty_str = item.get("quantity", "0") or "0"
                try:
                    quantity = int(qty_str)
                except:
                    quantity = 0

                text = item.get("text", "") or ""
                text = re.sub(r'<[^>]+>', ' ', text)
                text = re.sub(r'\s+', ' ', text).strip()

                product = Product(
                    uid=uid,
                    title=item.get("title", ""),
                    sku=item.get("sku", "") or "",
                    price=price,
                    price_old=price_old,
                    quantity=quantity,
                    url=item.get("url", ""),
                    category=category_name,
                    description=text,
                )
                products.append(product)

            print(f"  Страница {slice_num}: +{len(items)} товаров (всего: {total})")

            if len(items) < PER_PAGE or slice_num * PER_PAGE >= total:
                break

            slice_num += 1

        return products

    def parse_all(self, categories: Dict[str, str] = None):
        """Парсить все категории"""
        categories = categories or CATEGORIES

        print(f"\n{'='*60}")
        print(f"Парсинг каталога Orizhka.ru")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Категорий: {len(categories)}")
        print(f"{'='*60}\n")

        for i, (storepart_uid, category_name) in enumerate(categories.items(), 1):
            print(f"[{i}/{len(categories)}] {category_name}")

            products = self.get_products(storepart_uid, category_name)
            self.products.extend(products)

            print(f"  Итого: {len(products)} товаров\n")

        # Удаляем дубликаты
        unique_products = {}
        for p in self.products:
            if p.uid not in unique_products:
                unique_products[p.uid] = p
        self.products = list(unique_products.values())

        print(f"\n{'='*60}")
        print(f"ИТОГО: {len(self.products)} уникальных товаров")
        print(f"Ошибок: {len(self.errors)}")
        print(f"{'='*60}")

    def save_to_database(self):
        """Сохранить в PostgreSQL"""
        if not self.products:
            print("[DB] Нет товаров для сохранения")
            return

        conn = get_db()
        cur = conn.cursor()

        saved = 0
        updated = 0

        for p in self.products:
            cur.execute("""
                INSERT INTO products (product_id, sku, name, price, old_price, availability, category, url, city_id, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, NOW())
                ON CONFLICT (product_id) DO UPDATE SET
                    name = EXCLUDED.name,
                    price = EXCLUDED.price,
                    old_price = EXCLUDED.old_price,
                    availability = EXCLUDED.availability,
                    category = EXCLUDED.category,
                    url = EXCLUDED.url,
                    updated_at = NOW()
                RETURNING (xmax = 0) AS inserted
            """, (
                p.uid,
                p.sku or None,
                p.title,
                p.price,
                p.price_old if p.price_old > 0 else None,
                p.quantity,
                p.category,
                p.url
            ))

            result = cur.fetchone()
            if result and result[0]:
                saved += 1
            else:
                updated += 1

        conn.commit()
        cur.close()
        conn.close()

        print(f"[DB] Сохранено: {saved} новых, {updated} обновлено")

    def save_to_new_schema(self):
        """
        Сохранение в новую схему БД v10: orizhka_nomenclature (с price) + orizhka_product_urls
        Single-URL: один URL на товар (outlet_id = NULL), price в nomenclature
        """
        if not self.products:
            print("[DB] Нет товаров для сохранения")
            return

        conn = get_db()
        cur = conn.cursor()

        # Создаём outlet если не существует
        cur.execute("""
            INSERT INTO outlets (code, city, name, is_active)
            VALUES ('orizhka-spb', 'Санкт-Петербург', 'Orizhka СПб', true)
            ON CONFLICT (code) DO NOTHING
        """)

        nom_inserted = 0
        nom_updated = 0
        urls_inserted = 0

        for p in self.products:
            # Используем sku как article, или uid если sku пустой
            article = p.sku.strip() if p.sku else p.uid
            if not article:
                continue

            name = p.title.strip()
            if not name:
                continue

            # URL товара
            product_url = p.url
            if product_url and not product_url.startswith("http"):
                product_url = BASE_URL + product_url

            category = p.category or None
            price = p.price

            # 1. UPSERT в orizhka_nomenclature (price в nomenclature)
            cur.execute("""
                INSERT INTO orizhka_nomenclature (name, article, category, brand, price, first_seen_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (article) DO UPDATE SET
                    name = EXCLUDED.name,
                    category = EXCLUDED.category,
                    brand = EXCLUDED.brand,
                    price = EXCLUDED.price,
                    updated_at = NOW()
                RETURNING id, (xmax = 0) as inserted
            """, (name, article, category, 'Apple', price))

            row = cur.fetchone()
            nomenclature_id = row[0]
            if row[1]:
                nom_inserted += 1
            else:
                nom_updated += 1

            # 2. INSERT в orizhka_product_urls (single-URL: outlet_id = NULL)
            if product_url:
                cur.execute("""
                    INSERT INTO orizhka_product_urls (nomenclature_id, outlet_id, url, updated_at)
                    VALUES (%s, NULL, %s, NOW())
                    ON CONFLICT (url) DO NOTHING
                """, (nomenclature_id, product_url))
                urls_inserted += 1

        conn.commit()
        cur.close()
        conn.close()

        print(f"\n=== Сохранено в БД (v10) ===")
        print(f"orizhka_nomenclature: +{nom_inserted} новых, ~{nom_updated} обновлено")
        print(f"orizhka_product_urls: {urls_inserted} URL")

    def save_json(self, filepath: str = None):
        """Сохранить в JSON"""
        filepath = filepath or PRODUCTS_JSON

        data = {
            "parsed_at": datetime.now().isoformat(),
            "total": len(self.products),
            "products": [p.to_dict() for p in self.products]
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"Сохранено в {filepath}")

    def save_excel(self, filepath: str = None):
        """Сохранить в Excel"""
        filepath = filepath or PRODUCTS_XLSX

        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"

        headers = ["UID", "Название", "Артикул", "Цена", "Старая цена", "Кол-во", "Категория", "URL"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, product in enumerate(self.products, 2):
            ws.cell(row=row, column=1, value=product.uid)
            ws.cell(row=row, column=2, value=product.title)
            ws.cell(row=row, column=3, value=product.sku)
            ws.cell(row=row, column=4, value=product.price)
            ws.cell(row=row, column=5, value=product.price_old if product.price_old > 0 else "")
            ws.cell(row=row, column=6, value=product.quantity)
            ws.cell(row=row, column=7, value=product.category)
            ws.cell(row=row, column=8, value=product.url)

        column_widths = [20, 60, 15, 12, 12, 10, 20, 80]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.auto_filter.ref = f"A1:H{len(self.products) + 1}"

        wb.save(filepath)
        print(f"Сохранено в {filepath}")

    def close(self):
        self.client.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def save_staging(products: List[Product]):
    """Сохранение сырых данных в orizhka_staging"""
    if not products:
        print("[DB] Нет товаров для сохранения в staging")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("TRUNCATE TABLE orizhka_staging")

        insert_sql = """
            INSERT INTO orizhka_staging (
                outlet_code, name, article, category,
                brand, price, old_price, url
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """

        for p in products:
            article = p.sku.strip() if p.sku else p.uid
            cur.execute(insert_sql, (
                'orizhka-spb',
                p.title,
                article,
                p.category or '',
                'Apple',
                p.price,
                p.price_old if p.price_old > 0 else None,
                (BASE_URL + p.url) if p.url and not p.url.startswith("http") else p.url,
            ))

        conn.commit()
        print(f"[DB] Сохранено в orizhka_staging: {len(products)} товаров")
    finally:
        cur.close()
        conn.close()


def process_staging():
    """Обработка staging → orizhka_nomenclature (с price) + orizhka_product_urls"""
    conn = get_db()
    cur = conn.cursor()
    try:
        # Создаём outlet если не существует
        cur.execute("""
            INSERT INTO zip_outlets (code, city, name, is_active)
            VALUES ('orizhka-spb', 'Санкт-Петербург', 'Orizhka СПб', true)
            ON CONFLICT (code) DO NOTHING
        """)

        # 1. UPSERT в orizhka_nomenclature (price в nomenclature)
        cur.execute("""
            INSERT INTO orizhka_nomenclature (name, article, category, brand, price, first_seen_at, updated_at)
            SELECT DISTINCT ON (article)
                name, article, category, brand, price, NOW(), NOW()
            FROM orizhka_staging
            WHERE article IS NOT NULL AND article != ''
            ON CONFLICT (article) DO UPDATE SET
                name = EXCLUDED.name,
                category = EXCLUDED.category,
                brand = EXCLUDED.brand,
                price = EXCLUDED.price,
                updated_at = NOW()
        """)
        nom_count = cur.rowcount
        print(f"[DB] orizhka_nomenclature: {nom_count} записей")

        # 2. INSERT в orizhka_product_urls (single-URL: outlet_id = NULL)
        cur.execute("""
            INSERT INTO orizhka_product_urls (nomenclature_id, outlet_id, url, updated_at)
            SELECT DISTINCT ON (s.url)
                n.id, NULL, s.url, NOW()
            FROM orizhka_staging s
            JOIN orizhka_nomenclature n ON n.article = s.article
            WHERE s.article IS NOT NULL AND s.article != ''
              AND s.url IS NOT NULL AND s.url != ''
            ON CONFLICT (url) DO NOTHING
        """)
        url_count = cur.rowcount
        print(f"[DB] orizhka_product_urls: {url_count} URL")

        conn.commit()
    finally:
        cur.close()
        conn.close()


def main():
    """Основная функция"""
    arg_parser = argparse.ArgumentParser(description='Парсер Orizhka.ru')
    arg_parser.add_argument('--all', action='store_true', help='Полный цикл: парсинг + staging + process')
    arg_parser.add_argument('--process', action='store_true', help='Только обработка staging')
    arg_parser.add_argument('--no-db', action='store_true', help='Без сохранения в БД')
    arg_parser.add_argument('--direct', action='store_true', help='Прямой UPSERT в nomenclature+prices (без staging)')
    arg_parser.add_argument('--old-schema', action='store_true', help='Использовать старую схему БД (products)')
    arg_parser.add_argument('--init-db', action='store_true', help='Только инициализация БД')
    args = arg_parser.parse_args()

    print("Orizhka.ru Parser v2.0")
    print("-" * 40)

    if args.init_db:
        init_db()
        return

    # Только обработка staging
    if args.process:
        print("Обработка staging...")
        process_staging()
        print("\nОбработка завершена!")
        return

    with OrizhkaParser(save_to_db=not args.no_db) as parser:
        parser.parse_all()

        if not args.no_db:
            if args.direct:
                # Прямой UPSERT (без staging)
                parser.save_to_new_schema()
            elif args.old_schema:
                parser.save_to_database()
            else:
                # Стандарт: staging
                save_staging(parser.products)
                if args.all:
                    process_staging()

        parser.save_json()
        parser.save_excel()

    print("\nПарсинг завершён!")


if __name__ == "__main__":
    main()
