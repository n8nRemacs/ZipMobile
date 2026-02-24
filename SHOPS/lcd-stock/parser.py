"""
Парсер LCD-Stock.ru - магазин дисплеев и запчастей
Тип: HTML парсер с наличием по магазинам
БД: db_lcdstock
"""

import httpx
import json
import time
import os
import re
import argparse
import psycopg2
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import dataclass, field
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# === КОНФИГУРАЦИЯ ===

BASE_URL = "https://lcd-stock.ru"

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
PRODUCTS_JSON = os.path.join(DATA_DIR, "products.json")
PRODUCTS_XLSX = os.path.join(DATA_DIR, "products.xlsx")

REQUEST_DELAY = 0.3  # Задержка между запросами
REQUEST_TIMEOUT = 30

# === КОНФИГУРАЦИЯ БД (Supabase) ===
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db_wrapper import get_db  # Автоматически маппит таблицы на новые имена

# Таблицы (полные имена)
TABLE_PRODUCTS = "lcdstock_products"
TABLE_STOCK = "lcdstock_stock"
TABLE_NOMENCLATURE = "lcdstock_nomenclature"
TABLE_PRODUCT_URLS = "lcdstock_product_urls"
TABLE_STAGING = "lcdstock_staging"
TABLE_OUTLETS = "zip_outlets"

# Категории
CATEGORIES = {
    "displei": "Дисплеи",
    "zadnie-kryshki": "Задние крышки",
    "akkumulyatory": "Аккумуляторы",
    "aksessuary": "Аксессуары",
    "shlejfy-platy": "Платы зарядки",
    "smart-chasy": "Смарт-часы",
}

# Магазины
OUTLETS = [
    {"name": "ТК Савеловский", "slug": "savelovskiy", "address": "Москва, ТК Савеловский"},
    {"name": "Склад Савеловский", "slug": "sklad-savelovskiy", "address": "Москва, Склад Савеловский"},
    {"name": "ТЦ Митинский радиорынок", "slug": "mitinskiy", "address": "Москва, ТЦ Митинский радиорынок"},
    {"name": "ТЦ Горбушкин Двор", "slug": "gorbushka", "address": "Москва, ТЦ Горбушкин Двор"},
    {"name": "ТРЦ Мегаберезка", "slug": "megaberezka", "address": "Москва, ТРЦ Мегаберезка"},
]


def init_db():
    """Создать таблицы"""
    conn = get_db()
    cur = conn.cursor()

    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_PRODUCTS} (
            id SERIAL PRIMARY KEY,
            product_id VARCHAR(255) UNIQUE NOT NULL,
            sku VARCHAR(100),
            name TEXT NOT NULL,
            price NUMERIC(10,2),
            old_price NUMERIC(10,2),
            category VARCHAR(200),
            brand VARCHAR(100),
            color VARCHAR(100),
            url TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS {TABLE_STOCK} (
            id SERIAL PRIMARY KEY,
            product_id VARCHAR(255) NOT NULL,
            outlet_id INTEGER,
            status VARCHAR(50),
            quantity INTEGER DEFAULT 0,
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(product_id, outlet_id)
        );

        CREATE INDEX IF NOT EXISTS idx_{SHOP_PREFIX}_products_category ON {TABLE_PRODUCTS}(category);
        CREATE INDEX IF NOT EXISTS idx_{SHOP_PREFIX}_products_updated ON {TABLE_PRODUCTS}(updated_at);
        CREATE INDEX IF NOT EXISTS idx_{SHOP_PREFIX}_stock_product ON {TABLE_STOCK}(product_id);
        CREATE INDEX IF NOT EXISTS idx_{SHOP_PREFIX}_stock_outlet ON {TABLE_STOCK}(outlet_id);
    """)

    # Добавляем магазины
    for outlet in OUTLETS:
        cur.execute("""
            INSERT INTO {TABLE_OUTLETS} (name, slug, address)
            VALUES (%s, %s, %s)
            ON CONFLICT (slug) DO NOTHING
        """, (outlet["name"], outlet["slug"], outlet["address"]))

    conn.commit()
    cur.close()
    conn.close()
    print("[DB] Таблицы инициализированы")


@dataclass
class StockInfo:
    """Наличие в магазине"""
    outlet_name: str
    status: str  # "много", "мало", "нет в наличии"
    quantity: int = 0  # 0=нет, 1=мало, 2=много


@dataclass
class Product:
    """Товар"""
    product_id: str
    name: str
    price: float
    old_price: float
    url: str
    category: str
    brand: str = ""
    sku: str = ""
    color: str = ""
    stock: List[StockInfo] = field(default_factory=list)

    @property
    def total_quantity(self) -> int:
        return sum(s.quantity for s in self.stock)

    def to_dict(self) -> dict:
        return {
            "product_id": self.product_id,
            "name": self.name,
            "sku": self.sku,
            "price": self.price,
            "old_price": self.old_price,
            "url": self.url,
            "category": self.category,
            "brand": self.brand,
            "color": self.color,
            "stock": [{"outlet": s.outlet_name, "status": s.status} for s in self.stock],
        }


class LcdStockParser:
    """Парсер LCD-Stock.ru"""

    def __init__(self, parse_stock: bool = True):
        self.client = httpx.Client(
            timeout=REQUEST_TIMEOUT,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "ru,en;q=0.9",
            },
            follow_redirects=True,
        )
        self.products: List[Product] = []
        self.seen_ids: set = set()
        self.errors: List[Dict] = []
        self.last_request = 0
        self.parse_stock = parse_stock
        self.known_urls: Dict[str, str] = {}  # url → article, для инкрементального режима
        self.stats_skipped: int = 0

        os.makedirs(DATA_DIR, exist_ok=True)

    def _rate_limit(self):
        elapsed = time.time() - self.last_request
        if elapsed < REQUEST_DELAY:
            time.sleep(REQUEST_DELAY - elapsed)
        self.last_request = time.time()

    def _get_page(self, url: str) -> Optional[BeautifulSoup]:
        self._rate_limit()
        try:
            response = self.client.get(url)
            response.raise_for_status()
            return BeautifulSoup(response.text, 'html.parser')
        except Exception as e:
            self.errors.append({"url": url, "error": str(e)})
            return None

    def _extract_brand(self, name: str) -> str:
        name_lower = name.lower()
        brands = ["apple", "samsung", "xiaomi", "huawei", "honor", "realme", "oppo", "vivo", "tecno", "infinix", "poco"]
        for brand in brands:
            if brand in name_lower:
                return brand.capitalize()
        if "iphone" in name_lower or "ipad" in name_lower:
            return "Apple"
        if "galaxy" in name_lower:
            return "Samsung"
        if "redmi" in name_lower:
            return "Xiaomi"
        return ""

    def _extract_product_id(self, url: str) -> str:
        path = urlparse(url).path
        filename = os.path.basename(path)
        return filename.replace(".html", "")

    def _parse_price(self, text: str) -> float:
        if not text:
            return 0.0
        cleaned = re.sub(r'[^\d]', '', text)
        try:
            return float(cleaned)
        except:
            return 0.0

    def _parse_stock_status(self, status_text: str) -> tuple:
        """Парсить статус наличия -> (status, quantity)"""
        status = status_text.strip().lower()
        if "нет" in status:
            return ("нет в наличии", 0)
        elif "мало" in status:
            return ("мало", 1)
        elif "много" in status or "есть" in status:
            return ("много", 2)
        else:
            return (status_text.strip(), 1)

    def _get_product_details(self, url: str) -> Dict:
        """Получить детали товара: наличие, артикул, цвет"""
        result = {
            "stock": [],
            "sku": "",
            "color": ""
        }

        soup = self._get_page(url)
        if not soup:
            return result

        # Артикул: <div class="tovar-card-info"><span>Артикул: xxx</span></div>
        # Может быть несколько таких блоков, ищем тот где есть артикул
        info_divs = soup.select(".tovar-card-info")
        for info_div in info_divs:
            info_text = info_div.get_text(strip=True)
            if "Артикул:" in info_text:
                sku_match = re.search(r'Артикул:\s*(.+)', info_text)
                if sku_match:
                    result["sku"] = sku_match.group(1).strip()
                    break

        # Ищем характеристики товара
        items = soup.select("li")

        for item in items:
            left = item.select_one(".list-features-list-left")
            right = item.select_one(".list-features-list-right")

            if left and right:
                label = left.get_text(strip=True)
                value = right.get_text(strip=True)

                # Цвет
                if label.lower() == "цвет":
                    result["color"] = value

                # Наличие в магазинах (с ссылкой)
                left_link = left.select_one("a")
                if left_link:
                    outlet_name = left_link.get_text(strip=True)
                    # Проверяем что это магазин из нашего списка
                    if any(o["name"] in outlet_name for o in OUTLETS):
                        status, quantity = self._parse_stock_status(value)
                        result["stock"].append(StockInfo(
                            outlet_name=outlet_name,
                            status=status,
                            quantity=quantity
                        ))

        return result

    def _load_known_urls(self) -> Dict[str, str]:
        """Загрузить известные URL→article из БД для инкрементального режима"""
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute(f"""
                SELECT pu.url, n.article
                FROM {TABLE_PRODUCT_URLS} pu
                JOIN {TABLE_NOMENCLATURE} n ON n.id = pu.nomenclature_id
            """)
            result = {row[0]: row[1] for row in cur.fetchall()}
            cur.close()
            conn.close()
            return result
        except Exception as e:
            print(f"[WARN] _load_known_urls: {e}")
            return {}

    def parse_category(self, slug: str, name: str) -> List[Product]:
        """Парсить категорию"""
        products = []
        page = 1
        product_urls = []

        # Сначала собираем все URL товаров
        while True:
            if page == 1:
                url = f"{BASE_URL}/catalog/{slug}.html"
            else:
                url = f"{BASE_URL}/catalog/{slug}.html?page={page}"

            soup = self._get_page(url)
            if not soup:
                break

            cards = soup.select(".card-product")
            if not cards:
                break

            for card in cards:
                try:
                    title_el = card.select_one(".card-product_title")
                    title = title_el.get_text(strip=True) if title_el else ""

                    link_el = card.select_one("a")
                    href = link_el.get("href", "") if link_el else ""
                    if href and not href.startswith("http"):
                        href = urljoin(BASE_URL, href)

                    product_id = self._extract_product_id(href)
                    if product_id in self.seen_ids:
                        continue
                    self.seen_ids.add(product_id)

                    price_el = card.select_one(".product-price")
                    price = self._parse_price(price_el.get_text() if price_el else "")

                    old_price_el = card.select_one(".product-price-old")
                    old_price = self._parse_price(old_price_el.get_text() if old_price_el else "")

                    brand = self._extract_brand(title)

                    product_urls.append({
                        "product_id": product_id,
                        "name": title,
                        "price": price,
                        "old_price": old_price,
                        "url": href,
                        "category": name,
                        "brand": brand,
                    })

                except Exception as e:
                    self.errors.append({"category": name, "error": str(e)})

            print(f"  Страница {page}: +{len(cards)} товаров")

            next_link = soup.select_one(f'a[href*="page={page + 1}"]')
            if not next_link:
                break

            page += 1

        # Теперь парсим детали для каждого товара
        total = len(product_urls)
        for i, p_data in enumerate(product_urls):
            stock = []
            sku = ""
            color = ""

            # Инкрементальный режим: пропускаем detail fetch для известных URL
            if self.known_urls and p_data["url"] in self.known_urls:
                sku = self.known_urls[p_data["url"]]  # article из БД
                self.stats_skipped += 1
            elif self.parse_stock:
                details = self._get_product_details(p_data["url"])
                stock = details["stock"]
                sku = details["sku"]
                color = details["color"]
                if (i + 1) % 50 == 0:
                    print(f"  Детали: {i + 1}/{total}")

            product = Product(
                product_id=p_data["product_id"],
                name=p_data["name"],
                price=p_data["price"],
                old_price=p_data["old_price"],
                url=p_data["url"],
                category=p_data["category"],
                brand=p_data["brand"],
                sku=sku,
                color=color,
                stock=stock,
            )
            products.append(product)

        return products

    def parse_all(self, categories: Dict[str, str] = None):
        """Парсить все категории"""
        categories = categories or CATEGORIES

        print(f"\n{'='*60}")
        print(f"Парсинг каталога LCD-Stock.ru")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Категорий: {len(categories)}")
        print(f"Парсинг наличия: {'Да' if self.parse_stock else 'Нет'}")
        print(f"{'='*60}\n")

        for i, (slug, name) in enumerate(categories.items(), 1):
            print(f"[{i}/{len(categories)}] {name}")

            products = self.parse_category(slug, name)
            self.products.extend(products)

            print(f"  Итого: {len(products)} товаров\n")

        print(f"\n{'='*60}")
        print(f"ИТОГО: {len(self.products)} товаров")
        if self.stats_skipped > 0:
            print(f"Пропущено (уже в БД): {self.stats_skipped}")
        print(f"Ошибок: {len(self.errors)}")
        print(f"{'='*60}")

    def save_to_database(self):
        """Сохранить в PostgreSQL"""
        if not self.products:
            print("[DB] Нет товаров для сохранения")
            return

        conn = get_db()
        cur = conn.cursor()

        # Получаем ID магазинов
        cur.execute("SELECT slug, id FROM outlets")
        outlet_ids = {row[0]: row[1] for row in cur.fetchall()}

        # Создаем маппинг названий -> slug
        outlet_name_to_slug = {o["name"]: o["slug"] for o in OUTLETS}

        saved = 0
        updated = 0
        stock_saved = 0

        for p in self.products:
            # Сохраняем товар
            cur.execute("""
                INSERT INTO products (product_id, sku, name, price, old_price, category, brand, color, url, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (product_id) DO UPDATE SET
                    sku = EXCLUDED.sku,
                    name = EXCLUDED.name,
                    price = EXCLUDED.price,
                    old_price = EXCLUDED.old_price,
                    category = EXCLUDED.category,
                    brand = EXCLUDED.brand,
                    color = EXCLUDED.color,
                    url = EXCLUDED.url,
                    updated_at = NOW()
                RETURNING (xmax = 0) AS inserted
            """, (
                p.product_id,
                p.sku or None,
                p.name,
                p.price,
                p.old_price if p.old_price > 0 else None,
                p.category,
                p.brand or None,
                p.color or None,
                p.url
            ))

            result = cur.fetchone()
            if result and result[0]:
                saved += 1
            else:
                updated += 1

            # Сохраняем наличие
            for s in p.stock:
                slug = outlet_name_to_slug.get(s.outlet_name)
                if slug and slug in outlet_ids:
                    cur.execute("""
                        INSERT INTO stock (product_id, outlet_id, status, quantity, updated_at)
                        VALUES (%s, %s, %s, %s, NOW())
                        ON CONFLICT (product_id, outlet_id) DO UPDATE SET
                            status = EXCLUDED.status,
                            quantity = EXCLUDED.quantity,
                            updated_at = NOW()
                    """, (p.product_id, outlet_ids[slug], s.status, s.quantity))
                    stock_saved += 1

        conn.commit()
        cur.close()
        conn.close()

        print(f"[DB] Товары: {saved} новых, {updated} обновлено")
        print(f"[DB] Наличие: {stock_saved} записей")

    def save_to_new_schema(self, full_mode: bool = False):
        """
        Сохранение в новую схему БД v10: lcd_nomenclature (с price) + lcd_product_urls
        Single-URL: один URL на товар (outlet_id = NULL), price в nomenclature
        """
        if not self.products:
            print("[DB] Нет товаров для сохранения")
            return

        conn = get_db()
        cur = conn.cursor()

        # Создаём outlets если не существуют
        for outlet in OUTLETS:
            outlet_code = f"lcd-{outlet['slug']}"
            cur.execute("""
                INSERT INTO outlets (code, city, name, address, is_active)
                VALUES (%s, %s, %s, %s, true)
                ON CONFLICT (code) DO NOTHING
            """, (outlet_code, 'Москва', outlet['name'], outlet['address']))

        nom_inserted = 0
        nom_updated = 0
        urls_inserted = 0

        _nom_update = (
            "name = EXCLUDED.name, category = EXCLUDED.category, "
            "brand = COALESCE(NULLIF(EXCLUDED.brand, ''), lcd_nomenclature.brand), "
            "color = COALESCE(NULLIF(EXCLUDED.color, ''), lcd_nomenclature.color), "
            "price = EXCLUDED.price, updated_at = NOW()"
            if full_mode else
            "price = EXCLUDED.price, updated_at = NOW()"
        )
        for p in self.products:
            # Используем sku как article, или product_id если sku пустой
            article = p.sku.strip() if p.sku else p.product_id
            if not article:
                continue

            name = p.name.strip()
            if not name:
                continue

            # URL товара
            product_url = p.url
            if product_url and not product_url.startswith("http"):
                product_url = BASE_URL + product_url

            category = p.category or None
            brand = p.brand or None
            color = p.color or None
            price = p.price

            # 1. UPSERT в lcd_nomenclature (price в nomenclature)
            cur.execute(f"""
                INSERT INTO lcd_nomenclature (name, article, category, brand, color, price, first_seen_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (article) DO UPDATE SET
                    {_nom_update}
                RETURNING id, (xmax = 0) as inserted
            """, (name, article, category, brand, color, price))

            row = cur.fetchone()
            nomenclature_id = row[0]
            if row[1]:
                nom_inserted += 1
            else:
                nom_updated += 1

            # 2. INSERT в lcd_product_urls (single-URL: outlet_id = NULL)
            if product_url:
                cur.execute("""
                    INSERT INTO lcd_product_urls (nomenclature_id, outlet_id, url, updated_at)
                    VALUES (%s, NULL, %s, NOW())
                    ON CONFLICT (url) DO NOTHING
                """, (nomenclature_id, product_url))
                urls_inserted += 1

        conn.commit()
        cur.close()
        conn.close()

        print(f"\n=== Сохранено в БД (v10) ===")
        print(f"lcd_nomenclature: +{nom_inserted} новых, ~{nom_updated} обновлено")
        print(f"lcd_product_urls: {urls_inserted} URL")

    def save_json(self, filepath: str = None):
        filepath = filepath or PRODUCTS_JSON

        data = {
            "parsed_at": datetime.now().isoformat(),
            "source": "lcd-stock.ru",
            "total": len(self.products),
            "outlets": [o["name"] for o in OUTLETS],
            "products": [p.to_dict() for p in self.products]
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"Сохранено в {filepath}")

    def save_excel(self, filepath: str = None):
        filepath = filepath or PRODUCTS_XLSX

        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"

        # Заголовки с магазинами
        headers = ["ID", "Артикул", "Название", "Бренд", "Цвет", "Цена", "Категория"] + [o["name"] for o in OUTLETS] + ["URL"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, p in enumerate(self.products, 2):
            ws.cell(row=row, column=1, value=p.product_id)
            ws.cell(row=row, column=2, value=p.sku)
            ws.cell(row=row, column=3, value=p.name)
            ws.cell(row=row, column=4, value=p.brand)
            ws.cell(row=row, column=5, value=p.color)
            ws.cell(row=row, column=6, value=p.price)
            ws.cell(row=row, column=7, value=p.category)

            # Наличие по магазинам
            stock_dict = {s.outlet_name: s.status for s in p.stock}
            for i, outlet in enumerate(OUTLETS):
                status = stock_dict.get(outlet["name"], "")
                cell = ws.cell(row=row, column=8 + i, value=status)
                # Цветовая подсветка
                if "нет" in status.lower():
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif "мало" in status.lower():
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif status:
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            ws.cell(row=row, column=8 + len(OUTLETS), value=p.url)

        # Ширина колонок
        ws.column_dimensions['A'].width = 45  # ID
        ws.column_dimensions['B'].width = 30  # Артикул
        ws.column_dimensions['C'].width = 55  # Название
        ws.column_dimensions['D'].width = 12  # Бренд
        ws.column_dimensions['E'].width = 15  # Цвет
        ws.column_dimensions['F'].width = 10  # Цена
        ws.column_dimensions['G'].width = 15  # Категория
        for i in range(len(OUTLETS)):
            ws.column_dimensions[get_column_letter(8 + i)].width = 18
        ws.column_dimensions[get_column_letter(8 + len(OUTLETS))].width = 80

        ws.auto_filter.ref = f"A1:{get_column_letter(8 + len(OUTLETS))}{len(self.products) + 1}"

        wb.save(filepath)
        print(f"Сохранено в {filepath}")

    def close(self):
        self.client.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def save_staging(products: List[Product]):
    """Сохранение сырых данных в lcdstock_staging"""
    if not products:
        print("[DB] Нет товаров для сохранения в staging")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(f"TRUNCATE TABLE {TABLE_STAGING}")

        insert_sql = f"""
            INSERT INTO {TABLE_STAGING} (
                outlet_code, name, article, category,
                brand, color, price, old_price, url
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        for p in products:
            article = p.sku.strip() if p.sku else p.product_id

            # Для каждого магазина создаём запись
            if p.stock:
                for s in p.stock:
                    outlet_name = s.outlet_name
                    outlet_code = None
                    for o in OUTLETS:
                        if o["name"] in outlet_name:
                            outlet_code = f"lcd-{o['slug']}"
                            break
                    if not outlet_code:
                        continue

                    cur.execute(insert_sql, (
                        outlet_code,
                        p.name,
                        article,
                        p.category or '',
                        p.brand or '',
                        p.color or '',
                        p.price,
                        p.old_price if p.old_price > 0 else None,
                        p.url,
                    ))
            else:
                # Без наличия — одна запись
                cur.execute(insert_sql, (
                    'lcd-savelovskiy',
                    p.name,
                    article,
                    p.category or '',
                    p.brand or '',
                    p.color or '',
                    p.price,
                    p.old_price if p.old_price > 0 else None,
                    p.url,
                ))

        conn.commit()
        print(f"[DB] Сохранено в {TABLE_STAGING}: {cur.rowcount} записей")
    finally:
        cur.close()
        conn.close()


def process_staging(full_mode: bool = False):
    """Обработка staging → lcdstock_nomenclature (с price) + lcdstock_product_urls"""
    conn = get_db()
    cur = conn.cursor()
    try:
        # Создаём outlets
        for outlet in OUTLETS:
            outlet_code = f"lcd-{outlet['slug']}"
            cur.execute(f"""
                INSERT INTO {TABLE_OUTLETS} (code, city, name, address, is_active)
                VALUES (%s, %s, %s, %s, true)
                ON CONFLICT (code) DO NOTHING
            """, (outlet_code, 'Москва', outlet['name'], outlet['address']))

        # 1. UPSERT в nomenclature (price в nomenclature)
        _nom_update = (
            f"name = EXCLUDED.name, category = EXCLUDED.category, "
            f"brand = COALESCE(NULLIF(EXCLUDED.brand, ''), {TABLE_NOMENCLATURE}.brand), "
            f"color = COALESCE(NULLIF(EXCLUDED.color, ''), {TABLE_NOMENCLATURE}.color), "
            f"price = EXCLUDED.price, updated_at = NOW()"
            if full_mode else
            "price = EXCLUDED.price, updated_at = NOW()"
        )
        cur.execute(f"""
            INSERT INTO {TABLE_NOMENCLATURE} (name, article, category, brand, color, price, first_seen_at, updated_at)
            SELECT DISTINCT ON (article)
                name, article, category, brand, color, price, NOW(), NOW()
            FROM {TABLE_STAGING}
            WHERE article IS NOT NULL AND article != ''
            ON CONFLICT (article) DO UPDATE SET
                {_nom_update}
        """)
        nom_count = cur.rowcount
        print(f"[DB] {TABLE_NOMENCLATURE}: {nom_count} записей")

        # 2. INSERT в product_urls (single-URL: outlet_id = NULL)
        cur.execute(f"""
            INSERT INTO {TABLE_PRODUCT_URLS} (nomenclature_id, outlet_id, url, updated_at)
            SELECT DISTINCT ON (s.url)
                n.id, NULL, s.url, NOW()
            FROM {TABLE_STAGING} s
            JOIN {TABLE_NOMENCLATURE} n ON n.article = s.article
            WHERE s.article IS NOT NULL AND s.article != ''
              AND s.url IS NOT NULL AND s.url != ''
            ON CONFLICT (url) DO NOTHING
        """)
        url_count = cur.rowcount
        print(f"[DB] {TABLE_PRODUCT_URLS}: {url_count} URL")

        conn.commit()
    finally:
        cur.close()
        conn.close()


def main():
    arg_parser = argparse.ArgumentParser(description='Парсер LCD-Stock.ru')
    arg_parser.add_argument('--all', action='store_true', help='Полный цикл: парсинг + staging + process')
    arg_parser.add_argument('--process', action='store_true', help='Только обработка staging')
    arg_parser.add_argument('--no-db', action='store_true', help='Без сохранения в БД')
    arg_parser.add_argument('--direct', action='store_true', help='Прямой UPSERT без staging')
    arg_parser.add_argument('--old-schema', action='store_true', help='Использовать старую схему БД (products + stock)')
    arg_parser.add_argument('--no-stock', action='store_true', help='Без парсинга наличия (быстрый режим)')
    arg_parser.add_argument('--init-db', action='store_true', help='Только инициализация БД')
    arg_parser.add_argument('--category', type=str, help='Только одна категория')
    arg_parser.add_argument('--full', action='store_true', help='Полный парсинг (detail fetch для всех товаров, игнорировать кэш из БД)')
    args = arg_parser.parse_args()

    print("LCD-Stock.ru Parser v3.0")
    print("-" * 40)

    if args.init_db:
        init_db()
        return

    # Только обработка staging
    if args.process:
        print("Обработка staging...")
        process_staging(full_mode=args.full)
        print("\nОбработка завершена!")
        return

    categories = CATEGORIES
    if args.category:
        if args.category in CATEGORIES:
            categories = {args.category: CATEGORIES[args.category]}
        else:
            print(f"Категория '{args.category}' не найдена")
            return

    with LcdStockParser(parse_stock=not args.no_stock) as parser:
        if not args.full and not args.no_db:
            parser.known_urls = parser._load_known_urls()
            print(f"[DB] Известных URL: {len(parser.known_urls)}, инкрементальный режим")
        elif args.full:
            print(f"[FULL] Полный парсинг — detail fetch для всех товаров")

        parser.parse_all(categories)

        if not args.no_db:
            if args.direct:
                # Прямой UPSERT (без staging)
                parser.save_to_new_schema(full_mode=args.full)
            elif args.old_schema:
                parser.save_to_database()
            else:
                # Стандарт: staging
                save_staging(parser.products)
                if args.all:
                    process_staging(full_mode=args.full)

        parser.save_json()
        parser.save_excel()

    print("\nПарсинг завершён!")


if __name__ == "__main__":
    main()
