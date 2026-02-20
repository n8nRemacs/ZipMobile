"""
Парсер GreenSpark.ru v3 - объединённая версия
- Оригинальная логика парсинга (работает)
- IP ротация при блокировках
- Xvfb для получения cookies (обход детекции headless)
- Telegram уведомления
- Пропуск спарсенных городов

Серверы:
- server-a: 85.198.98.104
- server-b-ip1: 155.212.221.189
- server-b-ip2: 217.114.14.17
- server-c: 155.212.221.67 (новый)
"""

import httpx
import json
import time
import os
import re
import sys
import argparse
import psycopg2
import subprocess
import asyncio
from datetime import datetime, timedelta

# Отключаем буферизацию stdout для немедленного вывода в nohup логи
sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)
from typing import Optional, List, Dict, Tuple
from urllib.parse import urlencode, unquote
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from config import (
    BASE_URL, API_URL, PRODUCTS_ENDPOINT,
    ROOT_CATEGORY, REQUEST_DELAY, REQUEST_TIMEOUT, PER_PAGE,
    DEFAULT_SHOP_ID, DATA_DIR,
    PRODUCTS_JSON, PRODUCTS_XLSX, ERRORS_LOG, CATEGORIES_JSON
)

# Telegram уведомления
try:
    from telegram_notifier import TelegramNotifier, get_notifier
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    print("[WARN] telegram_notifier не найден")

# === Конфигурация БД ===
DB_HOST = os.environ.get("DB_HOST", "85.198.98.104")
DB_PORT = int(os.environ.get("DB_PORT", 5433))
DB_NAME = os.environ.get("DB_NAME", "db_greenspark")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "Mi31415926pSss!")

# Файл с cookies
COOKIES_FILE = "cookies.json"


def get_db():
    """Подключение к БД"""
    return psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        sslmode="require"
    )


# === КОНФИГУРАЦИЯ СЕРВЕРОВ ===

@dataclass
class ServerEndpoint:
    """Точка выхода (сервер + IP)"""
    name: str
    ip: str
    ssh_host: str
    ssh_user: str = "root"
    proxy_port: int = 0  # Порт SOCKS5 прокси (0 = не использовать)
    is_local: bool = False
    banned_until: datetime = None
    ban_count: int = 0

    @property
    def is_banned(self) -> bool:
        if self.banned_until is None:
            return False
        return datetime.now() < self.banned_until

    def mark_banned(self, minutes: int = 20):
        self.banned_until = datetime.now() + timedelta(minutes=minutes)
        self.ban_count += 1

    def clear_ban(self):
        self.banned_until = None
        self.ban_count = 0

    @property
    def proxy_url(self) -> Optional[str]:
        """URL SOCKS5 прокси если настроен"""
        if self.proxy_port > 0:
            return f"socks5://127.0.0.1:{self.proxy_port}"
        return None


# Все доступные серверы (proxy_port назначается динамически при переключении)
ENDPOINTS = [
    ServerEndpoint(name="server-a", ip="85.198.98.104", ssh_host="85.198.98.104"),
    ServerEndpoint(name="server-b-ip1", ip="155.212.221.189", ssh_host="155.212.221.189"),
    ServerEndpoint(name="server-b-ip2", ip="217.114.14.17", ssh_host="155.212.221.189"),
    ServerEndpoint(name="server-c", ip="155.212.221.67", ssh_host="155.212.221.67"),
]

# Настройки ожидания
INITIAL_WAIT_MINUTES = 50      # Разбан через ~45 мин, ставим 50 с запасом
WAIT_INCREMENT_MINUTES = 10
MAX_WAIT_MINUTES = 70


class IPRotator:
    """Менеджер ротации IP"""

    # Базовый порт для SOCKS прокси (каждый сервер получит +1, +2, ...)
    BASE_PROXY_PORT = 10800

    def __init__(self, current_server: str = None, shop_id: str = "16344"):
        self.shop_id = shop_id
        self.endpoints = [ServerEndpoint(
            name=ep.name, ip=ep.ip, ssh_host=ep.ssh_host
        ) for ep in ENDPOINTS]
        self.current_index = 0
        self.current_wait_time = INITIAL_WAIT_MINUTES
        self.notifier = get_notifier() if TELEGRAM_AVAILABLE else None

        # SSH туннели для прокси
        self.ssh_tunnels: Dict[str, subprocess.Popen] = {}

        # Статистика
        self.stats = {
            "products_total": 0,
            "products_session": 0,
            "cities_done": 0,
            "cities_total": 0,
            "bans": 0,
            "start_time": datetime.now(),
        }

        # Определяем текущий сервер
        if current_server:
            for i, ep in enumerate(self.endpoints):
                if ep.name == current_server:
                    self.current_index = i
                    ep.is_local = True
                    break
        else:
            self._detect_current_server()

        print(f"[ROTATOR] Текущий: {self.current_endpoint.name} ({self.current_endpoint.ip})")

    def _detect_current_server(self):
        """Определить текущий сервер по внешнему IP"""
        try:
            response = httpx.get("https://api.ipify.org", timeout=5)
            my_ip = response.text.strip()
            for i, ep in enumerate(self.endpoints):
                if ep.ip == my_ip:
                    self.current_index = i
                    ep.is_local = True
                    print(f"[ROTATOR] Определён IP: {my_ip}")
                    return
        except:
            pass

    @property
    def current_endpoint(self) -> ServerEndpoint:
        return self.endpoints[self.current_index]

    @property
    def all_banned(self) -> bool:
        return all(ep.is_banned for ep in self.endpoints)

    def get_next_endpoint(self) -> Optional[ServerEndpoint]:
        """Получить следующий незабаненный endpoint"""
        for _ in range(len(self.endpoints)):
            self.current_index = (self.current_index + 1) % len(self.endpoints)
            if not self.current_endpoint.is_banned:
                return self.current_endpoint
        return None

    def start_ssh_tunnel(self, endpoint: ServerEndpoint) -> bool:
        """Запустить SSH туннель для SOCKS5 прокси к серверу"""
        if endpoint.is_local:
            endpoint.proxy_port = 0  # Локальный сервер не нужен прокси
            return True

        # Назначаем порт
        endpoint_idx = next((i for i, ep in enumerate(self.endpoints) if ep.name == endpoint.name), 0)
        port = self.BASE_PROXY_PORT + endpoint_idx
        endpoint.proxy_port = port

        # Останавливаем старый туннель если есть
        self.stop_ssh_tunnel(endpoint)

        print(f"[ROTATOR] Запуск SSH туннеля к {endpoint.name} на порту {port}...")

        try:
            # SSH SOCKS5 прокси: ssh -D port -N -f user@host
            cmd = [
                "ssh", "-D", str(port),
                "-o", "StrictHostKeyChecking=no",
                "-o", "ServerAliveInterval=60",
                "-o", "ExitOnForwardFailure=yes",
                "-N",  # Без выполнения команды
                f"{endpoint.ssh_user}@{endpoint.ssh_host}"
            ]

            process = subprocess.Popen(
                cmd,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE
            )

            # Ждём немного и проверяем что туннель запустился
            time.sleep(2)
            if process.poll() is not None:
                # Процесс завершился - ошибка
                stderr = process.stderr.read().decode() if process.stderr else ""
                print(f"[ROTATOR] Ошибка SSH туннеля: {stderr[:200]}")
                return False

            self.ssh_tunnels[endpoint.name] = process
            print(f"[ROTATOR] SSH туннель к {endpoint.name} запущен (SOCKS5 на порту {port})")
            return True

        except Exception as e:
            print(f"[ROTATOR] Исключение SSH туннеля: {e}")
            return False

    def stop_ssh_tunnel(self, endpoint: ServerEndpoint):
        """Остановить SSH туннель"""
        if endpoint.name in self.ssh_tunnels:
            process = self.ssh_tunnels[endpoint.name]
            try:
                process.terminate()
                process.wait(timeout=5)
            except:
                process.kill()
            del self.ssh_tunnels[endpoint.name]
            print(f"[ROTATOR] SSH туннель к {endpoint.name} остановлен")

    def stop_all_tunnels(self):
        """Остановить все SSH туннели"""
        for name in list(self.ssh_tunnels.keys()):
            ep = next((e for e in self.endpoints if e.name == name), None)
            if ep:
                self.stop_ssh_tunnel(ep)

    def get_cookies_via_xvfb(self, endpoint: ServerEndpoint = None) -> Optional[dict]:
        """Получить cookies через Xvfb"""
        endpoint = endpoint or self.current_endpoint
        print(f"[ROTATOR] Получение cookies для {endpoint.name}...")

        # Скрипт получения cookies (без f-string для избежания проблем с shell)
        cookie_script = '''
import asyncio
from playwright.async_api import async_playwright
import json
import sys

SHOP_ID = sys.argv[1] if len(sys.argv) > 1 else "16344"

async def get_cookies():
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=['--disable-blink-features=AutomationControlled', '--no-sandbox']
        )
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080},
            locale='ru-RU'
        )
        await context.add_cookies([
            {'name': 'magazine', 'value': SHOP_ID, 'domain': 'green-spark.ru', 'path': '/'},
            {'name': 'global_magazine', 'value': SHOP_ID, 'domain': 'green-spark.ru', 'path': '/'},
        ])
        page = await context.new_page()
        await page.add_init_script('Object.defineProperty(navigator, "webdriver", {get: () => undefined})')
        await page.goto('https://green-spark.ru/catalog/komplektuyushchie_dlya_remonta/', wait_until='domcontentloaded')
        await page.wait_for_timeout(5000)
        await page.goto('https://green-spark.ru/local/api/catalog/products/?path[]=komplektuyushchie_dlya_remonta&perPage=10', wait_until='domcontentloaded')
        await page.wait_for_timeout(2000)
        cookies = await context.cookies()
        cookies_dict = {c['name']: c['value'] for c in cookies}
        cookies_dict['magazine'] = SHOP_ID
        cookies_dict['global_magazine'] = SHOP_ID
        cookies_dict['catalog-per-page'] = '100'
        with open('cookies.json', 'w') as f:
            json.dump(cookies_dict, f, indent=2)
        print(f'OK:{len(cookies_dict)}')
        await browser.close()

asyncio.run(get_cookies())
'''
        # Сохраняем скрипт локально
        local_script = "/tmp/get_cookies_gs.py"
        with open(local_script, 'w') as f:
            f.write(cookie_script)

        if endpoint.is_local:
            cmd = f"cd /opt/parsers/GreenSpark && xvfb-run --auto-servernum python3 {local_script} {self.shop_id}"
        else:
            # Удалённый сервер: копируем скрипт через scp, потом выполняем
            remote_script = "/tmp/get_cookies_gs.py"
            scp_script = f"scp {local_script} {endpoint.ssh_user}@{endpoint.ssh_host}:{remote_script}"

            try:
                subprocess.run(scp_script, shell=True, timeout=30, check=True)
            except Exception as e:
                print(f"[ROTATOR] Ошибка копирования скрипта: {e}")
                return None

            cmd = f"ssh {endpoint.ssh_user}@{endpoint.ssh_host} 'cd /opt/parsers/GreenSpark && xvfb-run --auto-servernum python3 {remote_script} {self.shop_id}'"

        try:
            result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=120)

            if 'OK:' in result.stdout:
                print(f"[ROTATOR] Cookies получены на {endpoint.name}")

                # Копируем cookies если удалённый
                if not endpoint.is_local:
                    scp_cmd = f"scp {endpoint.ssh_user}@{endpoint.ssh_host}:/opt/parsers/GreenSpark/cookies.json ./cookies.json"
                    subprocess.run(scp_cmd, shell=True, timeout=30)

                return self._load_cookies()
            else:
                print(f"[ROTATOR] Ошибка: {result.stderr[:200] if result.stderr else result.stdout[:200]}")
                return None

        except subprocess.TimeoutExpired:
            print(f"[ROTATOR] Таймаут xvfb-run")
            return None
        except Exception as e:
            print(f"[ROTATOR] Исключение: {e}")
            return None

    def _load_cookies(self) -> Optional[dict]:
        """Загрузить cookies из файла"""
        try:
            with open(COOKIES_FILE, "r", encoding="utf-8") as f:
                cookies = json.load(f)
            cookies.pop("__meta__", None)
            return cookies
        except:
            return None

    def handle_ban(self, reason: str, current_city: str = None) -> Tuple[bool, Optional[dict]]:
        """Обработать бан текущего IP"""
        self.stats["bans"] += 1
        current = self.current_endpoint

        # Помечаем как забаненный
        current.mark_banned(self.current_wait_time)

        print(f"\n{'!'*60}")
        print(f"[BAN] {current.name} ({current.ip}) забанен")
        print(f"[BAN] Причина: {reason}")
        print(f"{'!'*60}\n")

        # Telegram
        if self.notifier:
            self.notifier.notify_ban(
                server_name=current.name,
                ip=current.ip,
                reason=reason,
                products_session=self.stats["products_session"],
                total_products=self.stats["products_total"],
                cities_done=self.stats["cities_done"],
                cities_total=self.stats["cities_total"],
                current_city=current_city
            )

        # Останавливаем туннель текущего сервера если есть
        self.stop_ssh_tunnel(current)

        # Пробуем другой IP
        next_ep = self.get_next_endpoint()

        if next_ep:
            print(f"[ROTATOR] Переключение на {next_ep.name} ({next_ep.ip})")

            if self.notifier:
                self.notifier.notify_ip_switch(
                    server_name=next_ep.name,
                    old_ip=current.ip,
                    new_ip=next_ep.ip,
                    reason=reason,
                    products_parsed=self.stats["products_session"],
                    total_products=self.stats["products_total"]
                )

            # Запускаем SSH туннель для прокси (если не локальный сервер)
            if not next_ep.is_local:
                if not self.start_ssh_tunnel(next_ep):
                    print(f"[ROTATOR] Не удалось создать SSH туннель к {next_ep.name}")
                    next_ep.mark_banned(self.current_wait_time)
                    return self.handle_ban(f"Не удалось создать SSH туннель к {next_ep.ip}", current_city)

            cookies = self.get_cookies_via_xvfb(next_ep)
            if cookies:
                self.stats["products_session"] = 0
                return True, cookies  # current_endpoint уже обновлён в get_next_endpoint
            else:
                self.stop_ssh_tunnel(next_ep)
                next_ep.mark_banned(self.current_wait_time)
                return self.handle_ban(f"Не удалось получить cookies для {next_ep.ip}", current_city)
        else:
            return self._wait_and_retry(current_city)

    def _wait_and_retry(self, current_city: str = None) -> Tuple[bool, Optional[dict]]:
        """Ожидание когда все забанены"""
        wait_minutes = self.current_wait_time

        if self.notifier:
            servers_info = [{"name": ep.name, "ip": ep.ip} for ep in self.endpoints]
            self.notifier.notify_all_banned(
                servers=servers_info,
                wait_minutes=wait_minutes,
                total_products=self.stats["products_total"],
                cities_done=self.stats["cities_done"],
                cities_total=self.stats["cities_total"]
            )

        print(f"\n{'='*60}")
        print(f"[ROTATOR] ВСЕ IP ЗАБАНЕНЫ")
        print(f"[ROTATOR] Ожидание {wait_minutes} минут...")
        print(f"{'='*60}\n")

        time.sleep(wait_minutes * 60)

        # Пробуем первый endpoint
        for ep in self.endpoints:
            ep.clear_ban()

        self.current_index = 0
        cookies = self.get_cookies_via_xvfb(self.current_endpoint)

        if cookies:
            if self.notifier:
                self.notifier.notify_resume(
                    server_name=self.current_endpoint.name,
                    ip=self.current_endpoint.ip,
                    success=True,
                    total_products=self.stats["products_total"],
                    cities_done=self.stats["cities_done"],
                    cities_total=self.stats["cities_total"],
                    wait_time_was=wait_minutes
                )
            self.current_wait_time = INITIAL_WAIT_MINUTES
            self.stats["products_session"] = 0
            return True, cookies
        else:
            self.current_wait_time = min(
                self.current_wait_time + WAIT_INCREMENT_MINUTES,
                MAX_WAIT_MINUTES
            )
            self.current_endpoint.mark_banned(self.current_wait_time)
            return self._wait_and_retry(current_city)


class GreenSparkParser:
    """Парсер каталога GreenSpark с IP ротацией"""

    def __init__(self, shop_id: int = DEFAULT_SHOP_ID, rotator: IPRotator = None,
                 incremental_save: bool = True):
        self.shop_id = shop_id
        self.rotator = rotator
        self.client = None
        self.delay = REQUEST_DELAY
        self.last_request = 0
        self.products: List[Dict] = []
        self.categories: Dict[str, str] = {}
        self.errors: List[Dict] = []
        self.seen_ids: set = set()
        self.current_city: str = None
        self.current_city_id: int = None
        self.blocked = False

        # Инкрементальное сохранение
        self.incremental_save = incremental_save
        self.unsaved_products: List[Dict] = []
        self.total_saved = 0

        os.makedirs(DATA_DIR, exist_ok=True)

    def _maybe_save_incremental(self):
        """Сохранить если накопилось достаточно товаров"""
        if not self.incremental_save:
            return

        if len(self.unsaved_products) >= SAVE_EVERY_N_PRODUCTS:
            saved = save_products_incremental(self.unsaved_products, verbose=True)
            self.total_saved += saved
            self.unsaved_products = []

    def _flush_unsaved(self):
        """Сохранить оставшиеся товары"""
        if self.unsaved_products:
            saved = save_products_incremental(self.unsaved_products, verbose=True)
            self.total_saved += saved
            self.unsaved_products = []

    def _notify_block(self, reason: str):
        """Отправить Telegram при детекте блокировки"""
        if self.rotator and self.rotator.notifier:
            server_name = self.rotator.current_endpoint.name if self.rotator.current_endpoint else "unknown"
            ip = self.rotator.current_endpoint.ip if self.rotator.current_endpoint else "unknown"
            self.rotator.notifier.notify_ban(
                server_name=server_name,
                ip=ip,
                reason=reason,
                products_session=len(self.products),
                total_products=self.total_saved,
                cities_done=self.rotator.stats.get("cities_done", 0),
                cities_total=self.rotator.stats.get("cities_total", 0),
                current_city=self.current_city
            )

    def init_client(self, cookies: dict = None, proxy_url: str = None):
        """Инициализировать HTTP клиент с cookies и опциональным прокси"""
        if cookies is None:
            cookies = self._load_cookies()

        default_ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36"
        user_agent = unquote(cookies.get("__jua_", "")) or default_ua

        if self.client:
            self.client.close()

        # Получаем proxy_url из rotator если не передан
        if proxy_url is None and self.rotator:
            proxy_url = self.rotator.current_endpoint.proxy_url

        client_kwargs = {
            "timeout": REQUEST_TIMEOUT,
            "headers": {
                "User-Agent": user_agent,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "ru,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
            },
            "cookies": cookies,
            "follow_redirects": True,
        }

        # Добавляем прокси если есть
        if proxy_url:
            client_kwargs["proxy"] = proxy_url
            print(f"[CLIENT] Используем прокси: {proxy_url}")

        self.client = httpx.Client(**client_kwargs)
        self.blocked = False

    def _load_cookies(self) -> dict:
        """Загрузить cookies из файла"""
        if os.path.exists(COOKIES_FILE):
            try:
                with open(COOKIES_FILE, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)
                cookies.pop("__meta__", None)
                print(f"Cookies загружены из {COOKIES_FILE}")
                return cookies
            except Exception as e:
                print(f"Ошибка загрузки cookies: {e}")

        return {
            "magazine": str(self.shop_id),
            "global_magazine": str(self.shop_id),
            "catalog-per-page": str(PER_PAGE),
        }

    def _rate_limit(self):
        """Ограничение частоты запросов"""
        elapsed = time.time() - self.last_request
        if elapsed < self.delay:
            time.sleep(self.delay - elapsed)
        self.last_request = time.time()

    def load_cities(self) -> List[Dict]:
        """Загрузить список городов"""
        cities_path = os.path.join(os.path.dirname(__file__), "data", "greenspark_cities.json")
        with open(cities_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def get_parsed_city_ids(self) -> set:
        """Получить ID городов которые уже спарсены (имеют товары в greenspark_prices)"""
        conn = get_db()
        cur = conn.cursor()
        try:
            cur.execute("""
                SELECT DISTINCT SUBSTRING(o.code FROM 'greenspark-(.+)')::int as city_id
                FROM outlets o
                INNER JOIN greenspark_prices gp ON o.id = gp.outlet_id
                WHERE o.code LIKE 'greenspark-%'
            """)
            return set(row[0] for row in cur.fetchall() if row[0])
        finally:
            cur.close()
            conn.close()

    def set_city(self, city_id: int, city_name: str):
        """Установить город для парсинга"""
        self.client.cookies.set("magazine", str(city_id))
        self.client.cookies.set("global_magazine", str(city_id))
        self.current_city = city_name
        self.current_city_id = city_id

    def _build_path_params(self, path_parts: List[str]) -> str:
        """Построить параметры path[] для URL"""
        return "&".join([f"path[]={part}" for part in path_parts])

    def get_category_data(self, path_parts: List[str], page: int = 1) -> Optional[dict]:
        """Получить данные категории через API"""
        self._rate_limit()

        path_params = self._build_path_params(path_parts)
        url = f"{API_URL}{PRODUCTS_ENDPOINT}?{path_params}&orderBy=quantity&orderDirection=desc&perPage={PER_PAGE}&page={page}"

        try:
            response = self.client.get(url)

            # Проверяем блокировку
            content_type = response.headers.get('content-type', '')
            if 'application/json' not in content_type:
                print(f"[BLOCK] Получен не JSON ответ: {content_type}")
                self.blocked = True
                self._notify_block(f"Не JSON: {content_type[:50]}")
                return None

            if response.status_code == 403:
                print(f"[BLOCK] HTTP 403 - IP заблокирован")
                self.blocked = True
                self._notify_block("HTTP 403")
                return None

            response.raise_for_status()
            return response.json()

        except httpx.HTTPStatusError as e:
            if e.response.status_code == 403:
                self.blocked = True
            self.errors.append({
                "path": "/".join(path_parts),
                "error": f"HTTP {e.response.status_code}",
                "time": datetime.now().isoformat()
            })
            return None
        except Exception as e:
            self.errors.append({
                "path": "/".join(path_parts),
                "error": str(e),
                "time": datetime.now().isoformat()
            })
            return None

    def extract_article(self, picture_url: str) -> str:
        """Извлечь артикул из URL картинки"""
        match = re.search(r'gs-(\d+)', picture_url, re.IGNORECASE)
        if match:
            return f"GS-{match.group(1)}"

        match = re.search(r'ip-(\d+)', picture_url, re.IGNORECASE)
        if match:
            return f"ИП-{match.group(1)}"

        return ""

    def fetch_article_from_api(self, product_url: str) -> str:
        """Получить артикул через детальный API"""
        self._rate_limit()
        try:
            match = re.search(r'/catalog/(.+?)(?:\.html)?/?$', product_url)
            if not match:
                return ""

            path_str = match.group(1).rstrip('/')
            path_parts = path_str.split('/')
            path_params = self._build_path_params(path_parts)
            api_url = f"{API_URL}catalog/detail/?{path_params}"

            response = self.client.get(api_url)
            if response.status_code != 200:
                return ""

            content_type = response.headers.get('content-type', '')
            if 'application/json' not in content_type:
                return ""

            data = response.json()
            product = data.get("product", {})
            article = product.get("article", "")

            return article.strip() if article else ""

        except:
            return ""

    def fetch_article_from_page(self, product_url: str) -> str:
        """Получить артикул (API + HTML fallback)"""
        article = self.fetch_article_from_api(product_url)
        if article:
            return article

        self._rate_limit()
        try:
            response = self.client.get(product_url)
            if response.status_code != 200:
                return ""

            html = response.text
            # Формат 1: GS-00001234 (буквы-тире-цифры)
            match = re.search(r'Артикул[:\s]*([А-ЯA-Zа-яa-z]{2,3}-\d+)', html, re.IGNORECASE)
            if match:
                return match.group(1).upper()
            # Формат 2: 00000000656 (чистые цифры, 8+ символов)
            match = re.search(r'Артикул[:\s]*(\d{8,})', html, re.IGNORECASE)
            if match:
                return match.group(1)
            return ""
        except:
            return ""

    def reparse_missing_articles(self):
        """Допарсинг артикулов через единую номенклатуру greenspark_nomenclature"""
        missing = [(i, p) for i, p in enumerate(self.products) if not p.get("article")]

        if not missing:
            print("Все товары имеют артикулы")
            return

        print(f"\n{'='*60}")
        print(f"Допарсинг артикулов: {len(missing)} товаров без артикула")
        print(f"{'='*60}")

        # === ШАГ 1: Batch поиск по БД ===
        print(f"\n[Шаг 1] Поиск артикулов в БД (batch)...")

        # Собираем все URL для batch запроса
        urls_to_lookup = []
        url_to_indices = {}  # url -> list of product indices
        for i, product in missing:
            url = product.get("url", "")
            if url:
                urls_to_lookup.append(url)
                if url not in url_to_indices:
                    url_to_indices[url] = []
                url_to_indices[url].append(i)

        print(f"  URL для поиска: {len(urls_to_lookup)}")

        # Batch запрос к БД
        db_articles = self._batch_lookup_articles(urls_to_lookup)
        print(f"  Найдено в БД: {len(db_articles)} артикулов")

        # Применяем найденные артикулы
        from_db = 0
        for url, article in db_articles.items():
            if article and url in url_to_indices:
                for idx in url_to_indices[url]:
                    self.products[idx]["article"] = article
                    from_db += 1

        print(f"  Применено из БД: {from_db}")

        # === ШАГ 2: Определяем что нужно допарсить ===
        still_missing = [(i, p) for i, p in enumerate(self.products) if not p.get("article")]

        print(f"\n[Шаг 2] Товары для HTTP допарсинга: {len(still_missing)}")

        if still_missing:
            # Сохраняем в Excel для анализа
            excel_path = f"{DATA_DIR}/missing_articles.xlsx"
            self._save_missing_to_excel(still_missing, excel_path)
            print(f"\n  Список сохранён в: {excel_path}")

            # Выводим первые 20 в консоль
            print(f"\n--- Первые 20 товаров без артикула ---")
            for idx, (i, product) in enumerate(still_missing[:20]):
                url = product.get("url", "")
                name = product.get("name", "")[:60]
                print(f"  {idx+1}. {name}")
                print(f"     {url}")

            if len(still_missing) > 20:
                print(f"  ... и ещё {len(still_missing) - 20} товаров (см. Excel)")
            print("--- Конец списка ---\n")

        # === ШАГ 3: HTTP допарсинг ===
        if not still_missing:
            print("\n[Шаг 3] HTTP допарсинг не требуется - все артикулы найдены в БД!")
            print(f"\nИтого: найдено {from_db} артикулов из БД")
            return

        print(f"\n[Шаг 3] HTTP допарсинг {len(still_missing)} товаров...")

        from_http = 0
        for idx, (i, product) in enumerate(still_missing):
            if self.blocked:
                print("[BLOCK] Допарсинг прерван из-за блокировки")
                break

            url = product.get("url", "")
            if not url:
                continue

            article = self.fetch_article_from_page(url)

            if article:
                self.products[i]["article"] = article
                from_http += 1
                self._save_article_to_db(product, article)

            if (idx + 1) % 50 == 0:
                print(f"  HTTP: {idx + 1}/{len(still_missing)}, найдено: {from_http}")

        print(f"\n{'='*60}")
        print(f"Допарсинг завершён:")
        print(f"  - Из БД: {from_db}")
        print(f"  - Из HTTP: {from_http}")
        print(f"  - Всего: {from_db + from_http}")
        print(f"{'='*60}")

    def _batch_lookup_articles(self, urls: List[str]) -> Dict[str, str]:
        """Batch поиск артикулов по списку URL"""
        if not urls:
            return {}

        try:
            conn = get_db()
            cur = conn.cursor()

            # Используем ANY для batch запроса (исключаем NULL и пустые строки)
            cur.execute(
                "SELECT product_url, article FROM greenspark_nomenclature WHERE product_url = ANY(%s) AND article IS NOT NULL AND article != ''",
                (urls,)
            )

            result = {row[0]: row[1] for row in cur.fetchall()}

            cur.close()
            conn.close()

            return result
        except Exception as e:
            print(f"[BATCH LOOKUP ERROR] {e}")
            return {}

    def _save_missing_to_excel(self, missing_items: List[tuple], filepath: str):
        """Сохранить список товаров без артикула в Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Missing Articles"

            # Заголовки
            headers = ["№", "Название", "URL", "Категория", "Цена"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            # Данные
            for row_idx, (i, product) in enumerate(missing_items, 2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                ws.cell(row=row_idx, column=2, value=product.get("name", ""))
                ws.cell(row=row_idx, column=3, value=product.get("url", ""))
                ws.cell(row=row_idx, column=4, value=product.get("category", ""))
                ws.cell(row=row_idx, column=5, value=product.get("price", 0))

            # Ширина колонок
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 80
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 12

            wb.save(filepath)
            print(f"  Сохранено {len(missing_items)} товаров в {filepath}")
        except Exception as e:
            print(f"[EXCEL ERROR] {e}")

    def _lookup_article(self, name: str, url: str) -> Optional[str]:
        """Проверить артикул в единой номенклатуре. Возвращает: артикул или None"""
        try:
            conn = get_db()
            cur = conn.cursor()
            # Ищем только по URL (он уникален), name может отличаться между городами
            cur.execute(
                "SELECT article FROM greenspark_nomenclature WHERE product_url = %s",
                (url,)
            )
            row = cur.fetchone()
            cur.close()
            conn.close()

            if row is None:
                return None  # Товар не в номенклатуре (новый)
            return row[0]  # Артикул или None (если ещё не найден)
        except Exception as e:
            print(f"[LOOKUP ERROR] {e}")
            return None

    def _save_to_lookup(self, name: str, url: str, article: Optional[str]):
        """Обновить артикул в единой номенклатуре"""
        try:
            conn = get_db()
            cur = conn.cursor()
            # Ищем только по URL (он уникален)
            cur.execute("""
                UPDATE greenspark_nomenclature
                SET article = %s, updated_at = NOW()
                WHERE product_url = %s AND article IS NULL
            """, (article, url))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            print(f"[LOOKUP SAVE ERROR] {e}")

    def _save_article_to_db(self, product: dict, article: str):
        """Обновить артикул в единой номенклатуре (запись уже существует по url)"""
        try:
            conn = get_db()
            cur = conn.cursor()

            url = product.get("url", "")

            # UPDATE артикула только по URL (он уникален)
            cur.execute("""
                UPDATE greenspark_nomenclature
                SET article = %s, updated_at = NOW()
                WHERE product_url = %s
            """, (article, url))

            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            print(f"[DB ERROR] Ошибка обновления артикула {article}: {e}")

    def extract_breadcrumbs_path(self, breadcrumbs: List[dict]) -> tuple:
        """Извлечь путь из хлебных крошек"""
        slug_parts = []
        name_parts = []

        for crumb in breadcrumbs:
            url = crumb.get("url", "")
            name = crumb.get("name", "")

            match = re.search(r'/catalog/(.+?)/?$', url)
            if match:
                full_path = match.group(1).rstrip('/')
                last_part = full_path.split('/')[-1]
                slug_parts.append(last_part)
            name_parts.append(name)

        return "/".join(slug_parts), " / ".join(name_parts)

    def extract_product_info(self, product: dict, category_slug: str, category_name: str) -> Optional[Dict]:
        """Извлечь информацию о товаре"""
        try:
            product_id = product.get("id")

            if product_id in self.seen_ids:
                return None
            self.seen_ids.add(product_id)

            url_path = product.get("url", "")
            full_url = BASE_URL + url_path if url_path else ""

            article = product.get("article", "").strip()
            if not article:
                picture = product.get("picture", {})
                picture_url = picture.get("original", "") or picture.get("webp", "")
                article = self.extract_article(picture_url)

            prices = product.get("prices", [])
            price_retail = 0.0
            price_green5 = 0.0

            for p in prices:
                name = p.get("name", "").lower()
                price_value = p.get("price", 0) or 0

                if "розница" in name:
                    price_retail = float(price_value)
                elif "грин 5" in name:
                    price_green5 = float(price_value)

            return {
                "url": full_url,
                "category": category_name,
                "article": article,
                "name": product.get("name", ""),
                "price": price_retail,
                "price_wholesale": price_green5,
            }

        except Exception as e:
            self.errors.append({
                "product_id": str(product.get("id")),
                "error": str(e),
                "time": datetime.now().isoformat()
            })
            return None

    def crawl_category(self, path_parts: List[str], depth: int = 0):
        """Рекурсивный обход категории"""
        if self.blocked:
            return

        path_str = "/".join(path_parts)
        indent = "  " * depth

        print(f"{indent}[{depth}] Обход: {path_str}")

        data = self.get_category_data(path_parts, page=1)
        if not data:
            print(f"{indent}    [SKIP] Не удалось получить данные")
            return

        section_meta = data.get("sectionMeta", {})
        breadcrumbs = section_meta.get("breadcrumbs", [])
        category_slug, category_name = self.extract_breadcrumbs_path(breadcrumbs)

        if category_slug:
            self.categories[category_slug] = category_name

        subsections = data.get("subsections", [])

        if subsections:
            print(f"{indent}    Подкатегорий: {len(subsections)}")
            for sub in subsections:
                if self.blocked:
                    break
                sub_url = sub.get("url", "")
                match = re.search(r'/catalog/(.+?)/?$', sub_url)
                if match:
                    sub_path = match.group(1).rstrip('/').split('/')
                    self.crawl_category(sub_path, depth + 1)
        else:
            self._collect_products_from_category(data, path_parts, category_slug, category_name, depth)

    def _collect_products_from_category(self, first_page_data: dict, path_parts: List[str],
                                        category_slug: str, category_name: str, depth: int):
        """Собрать все товары из категории"""
        if self.blocked:
            return

        indent = "  " * depth

        products_data = first_page_data.get("products", {})
        products_list = products_data.get("data", [])
        meta = products_data.get("meta", {})
        total_pages = meta.get("pageCount", 1)
        total_products = meta.get("total", 0)

        print(f"{indent}    Товаров: {total_products}, страниц: {total_pages}")

        page_count = 0
        for product in products_list:
            info = self.extract_product_info(product, category_slug, category_name)
            if info:
                # Добавляем city_id сразу для инкрементального сохранения
                info["city_id"] = self.current_city_id
                info["city_name"] = self.current_city
                self.products.append(info)
                self.unsaved_products.append(info)
                page_count += 1

        print(f"{indent}    Страница 1/{total_pages}: +{page_count} товаров")
        self._maybe_save_incremental()

        for page in range(2, total_pages + 1):
            if self.blocked:
                break

            data = self.get_category_data(path_parts, page=page)
            if not data:
                continue

            products_data = data.get("products", {})
            products_list = products_data.get("data", [])

            page_count = 0
            for product in products_list:
                info = self.extract_product_info(product, category_slug, category_name)
                if info:
                    info["city_id"] = self.current_city_id
                    info["city_name"] = self.current_city
                    self.products.append(info)
                    self.unsaved_products.append(info)
                    page_count += 1

            print(f"{indent}    Страница {page}/{total_pages}: +{page_count} товаров")
            self._maybe_save_incremental()

    def parse_catalog(self, start_category: str = None, reparse_articles: bool = True):
        """Парсит каталог"""
        start = start_category or ROOT_CATEGORY

        print(f"\n{'='*60}")
        print(f"Парсинг каталога GreenSpark.ru")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Корневая категория: {start}")
        print(f"{'='*60}\n")

        print("[Этап 1] Обход каталога и сбор товаров\n")
        path_parts = start.split('/')
        self.crawl_category(path_parts)

        print(f"\n{'='*60}")
        print(f"Этап 1 завершён: {len(self.products)} товаров")
        print(f"Категорий: {len(self.categories)}")
        print(f"Ошибок: {len(self.errors)}")
        print(f"{'='*60}")

        if reparse_articles and not self.blocked:
            print("\n[Этап 2] Допарсинг артикулов")
            self.reparse_missing_articles()

        # Сохраняем оставшиеся товары
        self._flush_unsaved()

        missing_articles = len([p for p in self.products if not p.get("article")])
        print(f"\n{'='*60}")
        print(f"ИТОГО: {len(self.products)} товаров")
        print(f"С артикулами: {len(self.products) - missing_articles}")
        print(f"Без артикулов: {missing_articles}")
        if self.incremental_save:
            print(f"Сохранено в БД: {self.total_saved}")
        print(f"{'='*60}")

    def parse_city(self, city_id: int, city_name: str, start_category: str = None,
                   reparse_articles: bool = True) -> int:
        """Парсить один город, возвращает количество товаров"""
        self.products = []
        self.seen_ids = set()
        self.blocked = False
        self.unsaved_products = []
        self.total_saved = 0

        self.set_city(city_id, city_name)
        self.parse_catalog(start_category, reparse_articles)

        # city_id уже добавлен в _collect_products_from_category
        return len(self.products)

    def close(self):
        if self.client:
            self.client.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def ensure_outlets_for_cities():
    """Создаёт outlets для всех городов"""
    cities_path = os.path.join(os.path.dirname(__file__), "data", "greenspark_cities.json")
    with open(cities_path, 'r', encoding='utf-8') as f:
        cities = json.load(f)

    conn = get_db()
    cur = conn.cursor()
    try:
        for city in cities:
            city_id = city["set_city"]
            city_name = city["name"]
            outlet_code = f"greenspark-{city_id}"

            cur.execute("""
                INSERT INTO outlets (code, city, name, is_active)
                VALUES (%s, %s, %s, true)
                ON CONFLICT (code) DO UPDATE SET city = EXCLUDED.city, name = EXCLUDED.name
            """, (outlet_code, city_name, f"GreenSpark {city_name}"))

        conn.commit()
        print(f"Создано/обновлено {len(cities)} outlets")
    finally:
        cur.close()
        conn.close()


# Инкрементальное сохранение каждые N товаров
# 200 товаров парсятся за ~3-4 сек, сохранение batch UPSERT ~0.1-0.2 сек
# Замедление ~3-5%, но защита от потери данных при падении
SAVE_EVERY_N_PRODUCTS = 200


def save_products_incremental(products: List[Dict], verbose: bool = True):
    """Инкрементальное сохранение ВСЕХ товаров в единую номенклатуру (по name+url)"""
    if not products:
        return 0

    conn = get_db()
    cur = conn.cursor()
    saved = 0

    try:
        for p in products:
            name = p.get("name", "").strip()
            url = p.get("url", "").strip()
            if not name or not url:
                continue

            article = p.get("article", "").strip() or None  # NULL если пустой
            outlet_code = f"greenspark-{p.get('city_id')}" if p.get("city_id") else "greenspark-online"

            # UPSERT в greenspark_nomenclature (по url, артикул может быть NULL)
            cur.execute("""
                INSERT INTO greenspark_nomenclature (name, product_url, article, category, first_seen_at, updated_at)
                VALUES (%s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (product_url) DO UPDATE SET
                    name = EXCLUDED.name,
                    article = COALESCE(EXCLUDED.article, greenspark_nomenclature.article),
                    category = EXCLUDED.category,
                    updated_at = NOW()
                RETURNING id
            """, (name, url, article, p.get("category", "")))

            nom_row = cur.fetchone()
            if not nom_row:
                continue
            nom_id = nom_row[0]

            # Получаем outlet_id
            cur.execute("SELECT id FROM outlets WHERE code = %s", (outlet_code,))
            outlet_row = cur.fetchone()
            if not outlet_row:
                continue
            outlet_id = outlet_row[0]

            # UPSERT в greenspark_prices
            cur.execute("""
                INSERT INTO greenspark_prices (nomenclature_id, outlet_id, price, price_wholesale, updated_at)
                VALUES (%s, %s, %s, %s, NOW())
                ON CONFLICT (nomenclature_id, outlet_id) DO UPDATE SET
                    price = EXCLUDED.price, price_wholesale = EXCLUDED.price_wholesale,
                    updated_at = NOW()
            """, (nom_id, outlet_id, p.get("price", 0), p.get("price_wholesale", 0)))

            saved += 1

        conn.commit()
        if verbose:
            print(f"    [DB] Сохранено {saved} товаров")
        return saved

    except Exception as e:
        conn.rollback()
        print(f"    [DB ERROR] {e}")
        return 0
    finally:
        cur.close()
        conn.close()


def save_products_to_db(products: List[Dict]):
    """Сохранить ВСЕ товары в единую номенклатуру (batch режим)"""
    if not products:
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        saved_nom = 0
        saved_prices = 0

        for p in products:
            name = p.get("name", "").strip()
            url = p.get("url", "").strip()
            if not name or not url:
                continue

            article = p.get("article", "").strip() or None
            outlet_code = f"greenspark-{p.get('city_id')}" if p.get("city_id") else "greenspark-online"

            # UPSERT в greenspark_nomenclature (по url)
            cur.execute("""
                INSERT INTO greenspark_nomenclature (name, product_url, article, category, first_seen_at, updated_at)
                VALUES (%s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (product_url) DO UPDATE SET
                    name = EXCLUDED.name,
                    article = COALESCE(EXCLUDED.article, greenspark_nomenclature.article),
                    category = EXCLUDED.category,
                    updated_at = NOW()
                RETURNING id
            """, (name, url, article, p.get("category", "")))

            nom_row = cur.fetchone()
            if not nom_row:
                continue
            nom_id = nom_row[0]
            saved_nom += 1

            # Получаем outlet_id
            cur.execute("SELECT id FROM outlets WHERE code = %s", (outlet_code,))
            outlet_row = cur.fetchone()
            if not outlet_row:
                continue
            outlet_id = outlet_row[0]

            # UPSERT в greenspark_prices
            cur.execute("""
                INSERT INTO greenspark_prices (nomenclature_id, outlet_id, price, price_wholesale, updated_at)
                VALUES (%s, %s, %s, %s, NOW())
                ON CONFLICT (nomenclature_id, outlet_id) DO UPDATE SET
                    price = EXCLUDED.price, price_wholesale = EXCLUDED.price_wholesale,
                    updated_at = NOW()
            """, (nom_id, outlet_id, p.get("price", 0), p.get("price_wholesale", 0)))
            saved_prices += 1

        conn.commit()
        print(f"Сохранено: {saved_nom} товаров в номенклатуру, {saved_prices} цен")

    finally:
        cur.close()
        conn.close()


def main():
    arg_parser = argparse.ArgumentParser(description='Парсер GreenSpark.ru v3')
    arg_parser.add_argument('--server', type=str, help='Имя текущего сервера')
    arg_parser.add_argument('--all-cities', action='store_true', help='Парсить все города')
    arg_parser.add_argument('--city', type=str, help='Парсить конкретный город (имя)')
    arg_parser.add_argument('--skip-parsed', action='store_true', help='Пропускать спарсенные города')
    arg_parser.add_argument('--category', type=str, help='Стартовая категория')
    arg_parser.add_argument('--no-reparse', action='store_true', help='Без допарсинга артикулов')
    arg_parser.add_argument('--no-db', action='store_true', help='Без сохранения в БД')
    arg_parser.add_argument('--use-staging', action='store_true',
                           help='Использовать staging режим (Supabase через parser.py)')
    arg_parser.add_argument('--process', action='store_true',
                           help='Только process_staging (Supabase)')
    args = arg_parser.parse_args()

    # Только process_staging (через parser.py → Supabase)
    if args.process:
        from parser import process_staging as gs_process_staging
        print("[*] Processing staging (Supabase)...")
        gs_process_staging()
        return

    # Инициализация
    rotator = IPRotator(current_server=args.server)

    # Получаем cookies
    print("[INIT] Получение cookies...")
    cookies = rotator.get_cookies_via_xvfb()

    if not cookies:
        print("[ERROR] Не удалось получить cookies!")
        return

    parser = GreenSparkParser(rotator=rotator)
    parser.init_client(cookies)

    if args.all_cities or args.city:
        # Мультигородской парсинг
        cities = parser.load_cities()

        # Фильтр по конкретному городу
        if args.city:
            cities = [c for c in cities if c["name"].lower() == args.city.lower()]
            if not cities:
                print(f"[ERROR] Город '{args.city}' не найден!")
                return

        parsed_ids = parser.get_parsed_city_ids() if args.skip_parsed else set()

        # Фильтруем спарсенные
        cities_to_parse = [c for c in cities if c["set_city"] not in parsed_ids]

        print(f"\n{'='*60}")
        print(f"МУЛЬТИГОРОДСКОЙ ПАРСИНГ GreenSpark.ru v3")
        print(f"Всего городов: {len(cities)}")
        print(f"Уже спарсено: {len(parsed_ids)}")
        print(f"К парсингу: {len(cities_to_parse)}")
        print(f"{'='*60}\n")

        rotator.stats["cities_total"] = len(cities_to_parse)
        ensure_outlets_for_cities()

        for i, city in enumerate(cities_to_parse, 1):
            city_id = city["set_city"]
            city_name = city["name"]

            print(f"\n[{i}/{len(cities_to_parse)}] === ГОРОД: {city_name} (ID: {city_id}) ===\n")

            products_count = parser.parse_city(
                city_id, city_name,
                start_category=args.category,
                reparse_articles=not args.no_reparse
            )

            # Проверка блокировки
            if parser.blocked:
                print(f"[BLOCK] Город {city_name} заблокирован, переключаем IP...")
                success, new_cookies = rotator.handle_ban(
                    reason=f"Блокировка при парсинге {city_name}",
                    current_city=city_name
                )
                if success:
                    parser.init_client(new_cookies)
                    # Повторяем город
                    products_count = parser.parse_city(
                        city_id, city_name,
                        start_category=args.category,
                        reparse_articles=not args.no_reparse
                    )
                else:
                    print("[FATAL] Не удалось восстановиться после блокировки")
                    break

            # Сохраняем город
            if products_count > 0 and not args.no_db:
                if args.use_staging:
                    from parser import save_staging as gs_save_staging, process_staging as gs_process_staging
                    gs_save_staging(parser.products, multi_city=True, append=True)
                    gs_process_staging()
                else:
                    save_products_to_db(parser.products)

            rotator.stats["cities_done"] += 1
            rotator.stats["products_total"] += products_count

            print(f"Город {city_name}: {products_count} товаров")

        # Завершение
        if rotator.notifier:
            rotator.notifier.notify_complete(
                total_products=rotator.stats["products_total"],
                cities_done=rotator.stats["cities_done"],
                duration_minutes=int((datetime.now() - rotator.stats["start_time"]).total_seconds() / 60),
                errors=len(parser.errors)
            )

    else:
        # Одиночный город
        parser.parse_catalog(start_category=args.category, reparse_articles=not args.no_reparse)

        if not args.no_db:
            if args.use_staging:
                from parser import save_staging as gs_save_staging, process_staging as gs_process_staging
                gs_save_staging(parser.products)
                gs_process_staging()
            else:
                save_products_to_db(parser.products)

    parser.close()

    # Останавливаем все SSH туннели
    rotator.stop_all_tunnels()

    print("\nПарсинг завершён!")


if __name__ == "__main__":
    main()
