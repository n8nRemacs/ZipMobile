"""
Парсер GreenSpark.ru - обход каталога комплектующих для ремонта
Работает через API green-spark.ru/local/api/catalog/products/

База данных: db_greenspark
Таблицы: staging, outlets, nomenclature, current_prices, price_history
"""

import httpx
import json
import time
import os
import re
import argparse
import psycopg2
from datetime import datetime
from typing import Optional, List, Dict
from urllib.parse import urlencode, unquote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import subprocess
import random

# Координатор для эстафеты между серверами
try:
    from coordinator import ParserCoordinator
    COORDINATOR_AVAILABLE = True
except ImportError:
    COORDINATOR_AVAILABLE = False
    ParserCoordinator = None

from config import (
    BASE_URL, API_URL, PRODUCTS_ENDPOINT,
    ROOT_CATEGORY, REQUEST_DELAY, REQUEST_TIMEOUT, PER_PAGE,
    DEFAULT_SHOP_ID, DATA_DIR,
    PRODUCTS_JSON, PRODUCTS_XLSX, ERRORS_LOG, CATEGORIES_JSON
)

# === Настройки устойчивости ===
COOKIE_REFRESH_THRESHOLD = 3000  # Обновлять cookies каждые N товаров
EMPTY_RESPONSES_BEFORE_BLOCK = 5  # Сколько пустых ответов = блокировка
BLOCK_WAIT_MINUTES = 18  # Время ожидания при блокировке (чтобы не наложились блокировки IP)
DIAGNOSTICS_LOG = os.path.join(DATA_DIR, "diagnostics.json")  # Лог диагностики

# === Настройки мульти-IP ===
AVAILABLE_IPS = []  # Список доступных IP для переключения (заполняется через --ips)

# === Настройки прокси-ротации ===
PROXIES_FILE = "proxies.txt"  # Файл с прокси (ip:port на каждой строке)
PROXY_SWITCH_THRESHOLD = 3  # После скольких ошибок переключаться на следующий прокси

# Список User-Agent для ротации (разные браузеры/версии)
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Edg/120.0.0.0",
]

# === Конфигурация БД (Supabase) ===
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db_wrapper import get_db  # Автоматически маппит таблицы на новые имена

# Файл с cookies из браузера
COOKIES_FILE = "cookies.json"


class GreenSparkCatalogParser:
    """Парсер каталога GreenSpark через обход категорий"""

    def __init__(self, shop_id: int = DEFAULT_SHOP_ID, proxy: str = None, ips: list = None,
                 coordinator: 'ParserCoordinator' = None, server_name: str = None,
                 proxy_file: str = None, use_proxy: bool = False):
        self.shop_id = shop_id
        self.proxy = proxy  # Одиночный прокси (старый способ)
        self.use_proxy = use_proxy  # Режим ротации прокси

        # Координатор для эстафеты между серверами
        self.coordinator = coordinator
        self.server_name = server_name

        # Мульти-IP (legacy)
        self.available_ips = ips or []
        self.current_ip_index = 0
        self.current_ip = self.available_ips[0] if self.available_ips else None
        self.ip_block_times = {}  # {ip: timestamp} - время последней блокировки

        # Ротация прокси (новый способ)
        self.proxies = []
        self.current_proxy_idx = 0
        self.proxy_errors = 0  # Счётчик ошибок на текущем прокси
        if use_proxy:
            self._load_proxies(proxy_file or PROXIES_FILE)

        # Загружаем cookies из файла если есть
        cookies = self._load_cookies(shop_id)
        self.cookies = cookies

        # Извлекаем User-Agent из cookies (важно для обхода JS-защиты)
        self.default_ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 YaBrowser/24.12.0.0 Safari/537.36"
        self.user_agent = unquote(cookies.get("__jua_", "")) or self.default_ua

        # Создаём HTTP клиент
        self._create_client()

        self.delay = REQUEST_DELAY
        self.last_request = 0
        self.products: List[Dict] = []
        self.categories: Dict[str, str] = {}
        self.errors: List[Dict] = []
        self.seen_ids: set = set()

        # Мультигород
        self.current_city: str = None
        self.current_city_id: int = None

        # === Диагностика и устойчивость ===
        self.products_since_refresh = 0
        self.consecutive_empty = 0
        self.is_blocked = False
        self.diagnostics: List[Dict] = []
        self.total_requests = 0
        self.failed_requests = 0
        self.block_events: List[Dict] = []
        self.incremental_save = False

        os.makedirs(DATA_DIR, exist_ok=True)

    def _load_proxies(self, filename: str):
        """Загрузить прокси из файла"""
        if not os.path.exists(filename):
            print(f"[PROXY] Файл {filename} не найден, работаем без прокси")
            self.use_proxy = False
            return

        with open(filename, 'r') as f:
            self.proxies = [line.strip() for line in f if line.strip()]

        if self.proxies:
            print(f"[PROXY] Загружено {len(self.proxies)} прокси из {filename}")
        else:
            print(f"[PROXY] Файл {filename} пуст, работаем без прокси")
            self.use_proxy = False

    def _get_current_proxy(self) -> Optional[dict]:
        """Получить текущий прокси в формате для requests/httpx"""
        if not self.use_proxy or not self.proxies:
            return None

        proxy = self.proxies[self.current_proxy_idx]
        return f"http://{proxy}"

    def _switch_proxy(self) -> bool:
        """Переключиться на следующий прокси"""
        if not self.use_proxy or len(self.proxies) < 2:
            return False

        old_proxy = self.proxies[self.current_proxy_idx]
        self.current_proxy_idx = (self.current_proxy_idx + 1) % len(self.proxies)
        new_proxy = self.proxies[self.current_proxy_idx]

        self.proxy_errors = 0
        print(f"[PROXY] Переключение: {old_proxy} -> {new_proxy} (#{self.current_proxy_idx + 1}/{len(self.proxies)})")

        # Меняем User-Agent при смене прокси
        self.user_agent = self._rotate_user_agent()

        # Пересоздаём клиент с новым прокси
        try:
            self.client.close()
        except:
            pass

        self._create_client()

        # Восстанавливаем город
        if self.current_city_id:
            self.set_city(self.current_city_id, self.current_city)

        self.log_diagnostic("proxy_switch", "switched", {
            "from_proxy": old_proxy,
            "to_proxy": new_proxy,
            "proxy_idx": self.current_proxy_idx
        })

        return True

    def _create_client(self):
        """Создаёт HTTP клиент с текущими настройками IP/прокси"""
        # Настройки транспорта
        transport = None

        # Приоритет: 1) ротация прокси, 2) одиночный прокси, 3) IP
        if self.use_proxy and self.proxies:
            proxy_url = self._get_current_proxy()
            transport = httpx.HTTPTransport(proxy=proxy_url)
            print(f"[PROXY] Используется прокси #{self.current_proxy_idx + 1}: {self.proxies[self.current_proxy_idx]}")
        elif self.proxy:
            transport = httpx.HTTPTransport(proxy=self.proxy)
            print(f"Используется прокси: {self.proxy}")
        elif self.current_ip:
            transport = httpx.HTTPTransport(local_address=self.current_ip)
            print(f"Используется IP: {self.current_ip}")

        self.client = httpx.Client(
            timeout=REQUEST_TIMEOUT,
            headers={
                "User-Agent": self.user_agent,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "ru,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
            },
            cookies=self.cookies,
            follow_redirects=True,
            transport=transport,
        )

    def _rotate_user_agent(self) -> str:
        """Выбирает новый случайный User-Agent (отличающийся от текущего)."""
        available = [ua for ua in USER_AGENTS if ua != self.user_agent]
        if available:
            new_ua = random.choice(available)
        else:
            new_ua = random.choice(USER_AGENTS)
        return new_ua

    def switch_ip(self) -> bool:
        """Переключается на следующий IP. Возвращает True если переключение успешно."""
        if len(self.available_ips) < 2:
            return False

        old_ip = self.current_ip
        old_ua = self.user_agent
        self.current_ip_index = (self.current_ip_index + 1) % len(self.available_ips)
        self.current_ip = self.available_ips[self.current_ip_index]

        # Меняем отпечаток (User-Agent) при смене IP
        self.user_agent = self._rotate_user_agent()

        print(f"\n[IP SWITCH] Переключение: {old_ip} -> {self.current_ip}")
        print(f"[IP SWITCH] Новый User-Agent: {self.user_agent[:60]}...")
        self.log_diagnostic("ip_switch", "switched", {
            "from_ip": old_ip,
            "to_ip": self.current_ip,
            "new_user_agent": self.user_agent[:60]
        })

        # Закрываем старый клиент и создаём новый с новым User-Agent
        try:
            self.client.close()
        except:
            pass

        self._create_client()

        # Восстанавливаем город
        if self.current_city_id:
            self.set_city(self.current_city_id, self.current_city)

        return True

    def _load_cookies(self, shop_id: int) -> dict:
        """Загрузить cookies из файла или использовать базовые"""
        if os.path.exists(COOKIES_FILE):
            try:
                with open(COOKIES_FILE, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)
                print(f"Cookies загружены из {COOKIES_FILE}")
                return cookies
            except Exception as e:
                print(f"Ошибка загрузки cookies: {e}")

        # Базовые cookies
        return {
            "magazine": str(shop_id),
            "global_magazine": str(shop_id),
            "catalog-per-page": str(PER_PAGE),
        }

    def _rate_limit(self):
        """Ограничение частоты запросов"""
        elapsed = time.time() - self.last_request
        if elapsed < self.delay:
            time.sleep(self.delay - elapsed)
        self.last_request = time.time()

    def refresh_cookies(self):
        """Обновляет cookies через get_cookies.py"""
        print(f"\n{'='*50}")
        print(f"[COOKIES] Обновление cookies (спаршено {self.products_since_refresh} товаров)")
        print(f"{'='*50}")

        try:
            result = subprocess.run(
                ["python3", "get_cookies.py", "--headless"],
                capture_output=True, text=True, timeout=120, cwd=os.path.dirname(__file__)
            )

            if result.returncode == 0:
                # Перезагружаем cookies
                cookies = self._load_cookies(self.shop_id)
                for name, value in cookies.items():
                    self.client.cookies.set(name, value)

                # Восстанавливаем город
                if self.current_city_id:
                    self.set_city(self.current_city_id, self.current_city)

                self.products_since_refresh = 0
                self.consecutive_empty = 0
                self.is_blocked = False
                print("[COOKIES] Cookies обновлены успешно!")

                self.log_diagnostic("cookie_refresh", "success", {"products_parsed": self.products_since_refresh})
                return True
            else:
                print(f"[COOKIES] Ошибка: {result.stderr[:200]}")
                self.log_diagnostic("cookie_refresh", "failed", {"error": result.stderr[:500]})
                return False

        except subprocess.TimeoutExpired:
            print("[COOKIES] Таймаут обновления cookies (120s)")
            self.log_diagnostic("cookie_refresh", "timeout", {})
            return False
        except Exception as e:
            print(f"[COOKIES] Исключение: {e}")
            self.log_diagnostic("cookie_refresh", "exception", {"error": str(e)})
            return False

    def check_cookie_refresh(self):
        """Проверяет нужно ли обновить cookies"""
        if self.products_since_refresh >= COOKIE_REFRESH_THRESHOLD:
            self.refresh_cookies()

    def log_diagnostic(self, event_type: str, status: str, details: dict):
        """Логирует событие диагностики"""
        entry = {
            "time": datetime.now().isoformat(),
            "event": event_type,
            "status": status,
            "city": self.current_city,
            "city_id": self.current_city_id,
            "total_products": len(self.products),
            "total_requests": self.total_requests,
            "failed_requests": self.failed_requests,
            **details
        }
        self.diagnostics.append(entry)

        # Сохраняем в файл сразу (для отладки)
        try:
            with open(DIAGNOSTICS_LOG, 'w', encoding='utf-8') as f:
                json.dump(self.diagnostics, f, ensure_ascii=False, indent=2)
        except:
            pass

    def handle_blocking(self, reason: str):
        """Обрабатывает блокировку сайтом"""
        self.is_blocked = True
        block_event = {
            "time": datetime.now().isoformat(),
            "reason": reason,
            "city": self.current_city,
            "city_id": self.current_city_id,
            "consecutive_empty": self.consecutive_empty,
            "products_count": len(self.products)
        }
        self.block_events.append(block_event)

        print(f"\n{'!'*60}")
        print(f"[BLOCK] Обнаружена блокировка: {reason}")
        print(f"[BLOCK] Город: {self.current_city}, пустых ответов подряд: {self.consecutive_empty}")
        print(f"{'!'*60}")

        self.log_diagnostic("blocking_detected", reason, block_event)

        # === ПРИОРИТЕТ 1: РОТАЦИЯ ПРОКСИ ===
        if self.use_proxy and len(self.proxies) >= 2:
            print(f"[BLOCK] Режим ротации прокси: переключаемся на следующий прокси...")
            if self._switch_proxy():
                self.consecutive_empty = 0
                self.is_blocked = False
                # Небольшая пауза после смены прокси
                time.sleep(2)
                return True

        # === ПРИОРИТЕТ 2: КООРДИНАТОР (эстафета между серверами) ===
        if self.coordinator and not self.use_proxy:
            print(f"[BLOCK] Режим эстафеты: передаём управление другому серверу...")
            try:
                # Сообщаем о бане (NOTIFY другим серверам)
                self.coordinator.on_banned(BLOCK_WAIT_MINUTES)
                # Запускаем другой сервер
                self.coordinator.trigger_other_server()
                print(f"[BLOCK] Эстафета передана. Завершаем работу этого сервера.")
                # Выбрасываем исключение чтобы остановить парсинг
                raise SystemExit(f"Сервер забанен, эстафета передана. Выход.")
            except SystemExit:
                raise
            except Exception as e:
                print(f"[BLOCK] Ошибка координатора: {e}. Продолжаем стандартную обработку.")

        # === ПРИОРИТЕТ 3: РОТАЦИЯ IP (legacy) ===
        # Любая блокировка = блокировка по IP (403, Not JSON, таймауты)
        # Запоминаем время блокировки текущего IP
        if self.current_ip:
            self.ip_block_times[self.current_ip] = time.time()

        # Если есть второй IP - переключаемся
        if len(self.available_ips) >= 2:
            print(f"[BLOCK] Блокировка по IP. Переключаемся на другой IP...")
            self.switch_ip()

            # Проверяем нужно ли ждать для нового IP
            new_ip = self.current_ip
            if new_ip in self.ip_block_times:
                elapsed = time.time() - self.ip_block_times[new_ip]
                wait_needed = BLOCK_WAIT_MINUTES * 60 - elapsed
                if wait_needed > 0:
                    wait_mins = int(wait_needed / 60) + 1
                    print(f"[BLOCK] IP {new_ip} был заблокирован {int(elapsed/60)} мин назад. Ждём ещё {wait_mins} мин...")
                    self.log_diagnostic("ip_cooldown", "waiting", {"wait_minutes": wait_mins, "ip": new_ip})
                    time.sleep(wait_needed)
                else:
                    print(f"[BLOCK] IP {new_ip} разблокирован (прошло {int(elapsed/60)} мин)")

            # Обновляем cookies на новом IP
            print("[BLOCK] Получаем новые cookies...")
            self.refresh_cookies()
            self.consecutive_empty = 0
            self.is_blocked = False
            return True

        # Нет второго IP - пробуем обновить cookies
        print("[BLOCK] Попытка обновить cookies...")
        if self.refresh_cookies():
            print("[BLOCK] Cookies обновлены, пробуем продолжить...")
            self.consecutive_empty = 0
            return True

        # Если не помогло - ждём
        print(f"[BLOCK] Ожидание {BLOCK_WAIT_MINUTES} минут...")
        self.log_diagnostic("blocking_wait", "started", {"wait_minutes": BLOCK_WAIT_MINUTES})

        time.sleep(BLOCK_WAIT_MINUTES * 60)

        # После ожидания обновляем cookies
        print("[BLOCK] Ожидание завершено, обновляем cookies...")
        self.refresh_cookies()
        self.consecutive_empty = 0
        self.is_blocked = False

        self.log_diagnostic("blocking_wait", "completed", {})
        return True

    def load_cities(self) -> List[Dict]:
        """Загружает список городов из JSON файла"""
        cities_path = os.path.join(os.path.dirname(__file__), "data", "greenspark_cities.json")
        with open(cities_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def set_city(self, city_id: int, city_name: str):
        """Устанавливает город для парсинга через cookies"""
        self.client.cookies.set("magazine", str(city_id))
        self.client.cookies.set("global_magazine", str(city_id))
        self.current_city = city_name
        self.current_city_id = city_id

    def parse_all_cities(self, start_category: str = None, reparse_articles: bool = True,
                         save_after_each: bool = False, process_db: bool = False,
                         resume_from_city: int = None, skip_parsed: bool = False,
                         incremental_save: bool = False):
        """Парсит каталог для всех городов

        Args:
            save_after_each: Сохранять в БД после каждого города (защита от потери данных)
            process_db: Обрабатывать staging после каждого города
            resume_from_city: ID города с которого продолжить (пропустить предыдущие)
            skip_parsed: Пропускать города у которых уже есть данные в БД
            incremental_save: Сохранять в staging после каждой категории (защита от бана)
        """
        self.incremental_save = incremental_save
        cities = self.load_cities()
        all_products = []

        # Определяем с какого города начать
        start_index = 0
        if resume_from_city:
            for i, city in enumerate(cities):
                if city["set_city"] == resume_from_city:
                    start_index = i
                    print(f"[RESUME] Продолжение с города {city['name']} (ID: {resume_from_city})")
                    break

        # Получаем список уже спарсенных городов
        parsed_city_ids = set()
        if skip_parsed:
            parsed_city_ids = get_parsed_city_ids()
            print(f"[SKIP] Уже спарсено городов: {len(parsed_city_ids)}")

        print(f"\n{'='*60}")
        print(f"МУЛЬТИГОРОДСКОЙ ПАРСИНГ GreenSpark.ru")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Всего городов: {len(cities)}")
        print(f"Начать с индекса: {start_index}")
        print(f"Пропуск спарсенных: {skip_parsed}")
        print(f"Сохранение после каждого города: {save_after_each}")
        print(f"Инкрементальное сохранение: {incremental_save}")
        print(f"Cookie refresh каждые: {COOKIE_REFRESH_THRESHOLD} товаров")
        print(f"{'='*60}\n")

        # Очищаем staging только если не продолжаем
        if save_after_each and not resume_from_city and not skip_parsed:
            ensure_outlets_for_cities()
            clear_staging()
        elif save_after_each:
            ensure_outlets_for_cities()

        # Проверяем cookies в начале - делаем тестовый запрос
        print("[CHECK] Проверка cookies...")
        test_data = self.get_category_data([ROOT_CATEGORY], page=1)
        if not test_data:
            print("[CHECK] Cookies невалидны, обновляем...")
            self.refresh_cookies()
            # Повторяем тестовый запрос
            test_data = self.get_category_data([ROOT_CATEGORY], page=1)
            if not test_data:
                print("[ERROR] Не удалось получить данные после обновления cookies!")
                print("[ERROR] Возможно IP заблокирован. Ждём 15 минут...")
                self.handle_blocking("Initial check failed")
        else:
            print("[CHECK] Cookies валидны, продолжаем парсинг")
        self.consecutive_empty = 0  # Сбрасываем после проверки

        cities_to_parse = cities[start_index:]
        total_cities = len(cities_to_parse)

        for i, city in enumerate(cities_to_parse, 1):
            city_id = city["set_city"]
            city_name = city["name"]

            # Пропускаем уже спарсенные
            if skip_parsed and city_id in parsed_city_ids:
                print(f"\n[{i}/{total_cities}] === ГОРОД: {city_name} (ID: {city_id}) === [SKIP - уже спарсен]")
                continue

            print(f"\n[{i}/{total_cities}] === ГОРОД: {city_name} (ID: {city_id}) ===\n")

            # Сбрасываем состояние для нового города
            self.products = []
            self.seen_ids = set()
            # НЕ сбрасываем consecutive_empty - мониторим между городами тоже

            # Устанавливаем город
            self.set_city(city_id, city_name)

            # Логируем начало парсинга города
            self.log_diagnostic("city_start", "started", {"city": city_name, "city_id": city_id})

            # Парсим каталог для города
            self.parse_catalog(start_category, reparse_articles)

            # Добавляем city_id и city_name к каждому товару
            for p in self.products:
                p["city_id"] = city_id
                p["city_name"] = city_name

            all_products.extend(self.products)
            print(f"Город {city_name}: {len(self.products)} товаров")

            # Логируем завершение города
            self.log_diagnostic("city_complete", "success" if self.products else "empty",
                              {"city": city_name, "city_id": city_id, "products": len(self.products)})

            # Сохраняем после каждого города
            if save_after_each and self.products:
                print(f"  [SAVE] Сохранение {len(self.products)} товаров города {city_name}...")
                save_staging(self.products, multi_city=True, append=True)
                if process_db:
                    process_staging()
                print(f"  [SAVE] Сохранено! Всего в staging: {get_staging_count()} записей")

        self.products = all_products

        print(f"\n{'='*60}")
        print(f"ИТОГО ПО ВСЕМ ГОРОДАМ: {len(self.products)} записей")
        print(f"Блокировок: {len(self.block_events)}")
        print(f"Всего запросов: {self.total_requests}, неудачных: {self.failed_requests}")
        print(f"{'='*60}")

        # Сохраняем финальную диагностику
        self.save_diagnostics()

        return all_products

    def save_diagnostics(self):
        """Сохраняет полную диагностику в файл"""
        report = {
            "generated_at": datetime.now().isoformat(),
            "total_products": len(self.products),
            "total_requests": self.total_requests,
            "failed_requests": self.failed_requests,
            "block_events": self.block_events,
            "diagnostics": self.diagnostics
        }
        try:
            with open(DIAGNOSTICS_LOG, 'w', encoding='utf-8') as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
            print(f"Диагностика сохранена в {DIAGNOSTICS_LOG}")
        except Exception as e:
            print(f"Ошибка сохранения диагностики: {e}")

    def _build_path_params(self, path_parts: List[str]) -> str:
        """Построить параметры path[] для URL"""
        params = []
        for part in path_parts:
            params.append(f"path[]={part}")
        return "&".join(params)

    def get_category_data(self, path_parts: List[str], page: int = 1) -> Optional[dict]:
        """Получить данные категории через API с диагностикой"""
        self._rate_limit()
        self.total_requests += 1

        path_params = self._build_path_params(path_parts)
        url = f"{API_URL}{PRODUCTS_ENDPOINT}?{path_params}&orderBy=quantity&orderDirection=desc&perPage={PER_PAGE}&page={page}"
        path_str = "/".join(path_parts)

        try:
            response = self.client.get(url)
            status_code = response.status_code

            # Детальная диагностика по статус-коду
            if status_code == 403:
                self.failed_requests += 1
                self.consecutive_empty += 1
                self.proxy_errors += 1
                error_info = {
                    "path": path_str,
                    "error": "HTTP 403 Forbidden - доступ запрещён (возможно блокировка)",
                    "status_code": 403,
                    "time": datetime.now().isoformat()
                }
                self.errors.append(error_info)
                self.log_diagnostic("request_error", "forbidden", error_info)

                # При ротации прокси - сразу переключаемся на 403
                if self.use_proxy and self.proxy_errors >= PROXY_SWITCH_THRESHOLD:
                    self._switch_proxy()
                    self.consecutive_empty = 0
                elif self.consecutive_empty >= EMPTY_RESPONSES_BEFORE_BLOCK:
                    self.handle_blocking("HTTP 403 - доступ заблокирован")
                return None

            elif status_code == 429:
                self.failed_requests += 1
                self.consecutive_empty += 1
                self.proxy_errors += 1
                error_info = {
                    "path": path_str,
                    "error": "HTTP 429 Too Many Requests - слишком много запросов",
                    "status_code": 429,
                    "time": datetime.now().isoformat()
                }
                self.errors.append(error_info)
                self.log_diagnostic("request_error", "rate_limited", error_info)

                # При 429 сразу переключаем прокси (если доступны)
                if self.use_proxy:
                    self._switch_proxy()
                    self.consecutive_empty = 0
                    time.sleep(1)  # Небольшая пауза
                else:
                    self.handle_blocking("HTTP 429 - rate limit")
                return None

            elif status_code >= 500:
                self.failed_requests += 1
                error_info = {
                    "path": path_str,
                    "error": f"HTTP {status_code} - ошибка сервера",
                    "status_code": status_code,
                    "time": datetime.now().isoformat()
                }
                self.errors.append(error_info)
                self.log_diagnostic("request_error", "server_error", error_info)
                return None

            response.raise_for_status()

            # Проверяем что это JSON
            content_type = response.headers.get('content-type', '')
            if 'application/json' not in content_type:
                self.failed_requests += 1
                self.consecutive_empty += 1
                error_info = {
                    "path": path_str,
                    "error": f"Not JSON response: {content_type} - возможно JS-защита или блокировка",
                    "content_type": content_type,
                    "body_preview": response.text[:200] if response.text else "",
                    "time": datetime.now().isoformat()
                }
                self.errors.append(error_info)
                self.log_diagnostic("request_error", "not_json", error_info)

                if self.consecutive_empty >= EMPTY_RESPONSES_BEFORE_BLOCK:
                    self.handle_blocking("Not JSON - JS защита или блокировка")
                return None

            data = response.json()

            # Проверяем есть ли данные
            products = data.get("products", {}).get("data", [])
            if not products and not data.get("subsections"):
                self.consecutive_empty += 1
                if self.consecutive_empty >= EMPTY_RESPONSES_BEFORE_BLOCK:
                    self.handle_blocking("Много пустых ответов подряд")
            else:
                self.consecutive_empty = 0  # Сбрасываем счётчик при успехе
                self.proxy_errors = 0  # Сбрасываем счётчик ошибок прокси при успехе

            return data

        except httpx.TimeoutException as e:
            self.failed_requests += 1
            self.consecutive_empty += 1
            self.proxy_errors += 1
            error_info = {
                "path": path_str,
                "error": f"Timeout: {str(e)}",
                "error_type": "timeout",
                "time": datetime.now().isoformat()
            }
            self.errors.append(error_info)
            self.log_diagnostic("request_error", "timeout", error_info)

            # При прокси-ротации переключаемся после нескольких таймаутов
            if self.use_proxy and self.proxy_errors >= PROXY_SWITCH_THRESHOLD:
                self._switch_proxy()
                self.consecutive_empty = 0
            elif self.consecutive_empty >= EMPTY_RESPONSES_BEFORE_BLOCK:
                self.handle_blocking("Таймауты - возможно блокировка")
            return None

        except httpx.ConnectError as e:
            self.failed_requests += 1
            self.consecutive_empty += 1
            self.proxy_errors += 1
            error_info = {
                "path": path_str,
                "error": f"Connection error: {str(e)}",
                "error_type": "connection",
                "time": datetime.now().isoformat()
            }
            self.errors.append(error_info)
            self.log_diagnostic("request_error", "connection_error", error_info)

            # При прокси-ротации сразу переключаемся на ошибку подключения
            if self.use_proxy:
                self._switch_proxy()
                self.consecutive_empty = 0
            elif self.consecutive_empty >= EMPTY_RESPONSES_BEFORE_BLOCK:
                self.handle_blocking("Ошибки подключения")
            return None

        except httpx.HTTPStatusError as e:
            self.failed_requests += 1
            error_info = {
                "path": path_str,
                "error": f"HTTP {e.response.status_code}",
                "status_code": e.response.status_code,
                "time": datetime.now().isoformat()
            }
            self.errors.append(error_info)
            self.log_diagnostic("request_error", "http_error", error_info)
            return None

        except Exception as e:
            self.failed_requests += 1
            error_info = {
                "path": path_str,
                "error": str(e),
                "error_type": type(e).__name__,
                "time": datetime.now().isoformat()
            }
            self.errors.append(error_info)
            self.log_diagnostic("request_error", "exception", error_info)
            return None

    def extract_article(self, picture_url: str) -> str:
        """Извлечь артикул из URL картинки"""
        # gs-00006970_1.jpg -> GS-00006970
        match = re.search(r'gs-(\d+)', picture_url, re.IGNORECASE)
        if match:
            return f"GS-{match.group(1)}"
        return ""

    def fetch_article_from_api(self, product_url: str) -> str:
        """Получить артикул через детальный API"""
        self._rate_limit()
        try:
            # Извлекаем path из URL: /catalog/a/b/c/product.html -> ["a", "b", "c", "product"]
            match = re.search(r'/catalog/(.+?)(?:\.html)?/?$', product_url)
            if not match:
                return ""

            path_str = match.group(1).rstrip('/')
            path_parts = path_str.split('/')

            # Строим URL для detail API
            path_params = self._build_path_params(path_parts)
            api_url = f"{API_URL}catalog/detail/?{path_params}"

            response = self.client.get(api_url)
            if response.status_code != 200:
                return ""

            # Проверяем Content-Type
            content_type = response.headers.get('content-type', '')
            if 'application/json' not in content_type:
                return ""

            data = response.json()
            product = data.get("product", {})
            article = product.get("article", "")

            return article.strip() if article else ""

        except Exception as e:
            return ""

    def fetch_article_from_page(self, product_url: str) -> str:
        """Получить артикул (сначала API, потом HTML fallback)"""
        # Сначала пробуем API
        article = self.fetch_article_from_api(product_url)
        if article:
            return article

        # Fallback: парсим HTML
        self._rate_limit()
        try:
            response = self.client.get(product_url)
            if response.status_code != 200:
                return ""

            html = response.text
            # Ищем паттерн "Артикул: XXX-XXXXX" (русские и латинские буквы)
            match = re.search(r'Артикул[:\s]*([А-ЯA-Zа-яa-z]{2,3}-\d+)', html, re.IGNORECASE)
            if match:
                return match.group(1).upper()
            return ""
        except Exception as e:
            return ""

    def reparse_missing_articles(self):
        """Допарсинг артикулов для товаров без артикула"""
        missing = [(i, p) for i, p in enumerate(self.products) if not p.get("article")]

        if not missing:
            print("Все товары имеют артикулы")
            return

        print(f"\n{'='*60}")
        print(f"Допарсинг артикулов: {len(missing)} товаров без артикула")
        print(f"{'='*60}\n")

        found = 0
        for idx, (i, product) in enumerate(missing):
            url = product.get("url", "")
            if not url:
                continue

            article = self.fetch_article_from_page(url)
            if article:
                self.products[i]["article"] = article
                found += 1

            # Прогресс каждые 100 товаров
            if (idx + 1) % 100 == 0:
                print(f"  Обработано {idx + 1}/{len(missing)}, найдено артикулов: {found}")

        print(f"\nДопарсинг завершён: найдено {found} артикулов из {len(missing)}")

    def extract_breadcrumbs_path(self, breadcrumbs: List[dict]) -> tuple:
        """Извлечь путь из хлебных крошек"""
        slug_parts = []
        name_parts = []

        for crumb in breadcrumbs:
            url = crumb.get("url", "")
            name = crumb.get("name", "")

            # Извлекаем slug из URL: /catalog/komplektuyushchie_dlya_remonta/ -> komplektuyushchie_dlya_remonta
            match = re.search(r'/catalog/(.+?)/?$', url)
            if match:
                # Берём только последнюю часть пути
                full_path = match.group(1).rstrip('/')
                last_part = full_path.split('/')[-1]
                slug_parts.append(last_part)
            name_parts.append(name)

        return "/".join(slug_parts), " / ".join(name_parts)

    def extract_product_info(self, product: dict, category_slug: str, category_name: str) -> Optional[Dict]:
        """Извлечь информацию о товаре"""
        try:
            product_id = product.get("id")

            # Пропускаем дубликаты
            if product_id in self.seen_ids:
                return None
            self.seen_ids.add(product_id)

            # URL товара
            url_path = product.get("url", "")
            full_url = BASE_URL + url_path if url_path else ""

            # Артикул из картинки
            picture = product.get("picture", {})
            picture_url = picture.get("original", "") or picture.get("webp", "")
            article = self.extract_article(picture_url)

            # Цены
            prices = product.get("prices", [])
            price_retail = 0.0
            price_green5 = 0.0

            for p in prices:
                name = p.get("name", "").lower()
                price_value = p.get("price", 0) or 0

                if "розница" in name:
                    price_retail = float(price_value)
                elif "грин 5" in name:
                    price_green5 = float(price_value)

            return {
                "url": full_url,
                "category": category_name,
                "article": article,
                "name": product.get("name", ""),
                "price": price_retail,
                "price_wholesale": price_green5,
            }

        except Exception as e:
            self.errors.append({
                "product_id": str(product.get("id")),
                "error": str(e),
                "time": datetime.now().isoformat()
            })
            return None

    def crawl_category(self, path_parts: List[str], depth: int = 0):
        """Рекурсивный обход категории"""
        path_str = "/".join(path_parts)
        indent = "  " * depth

        print(f"{indent}[{depth}] Обход: {path_str}")

        # Получаем данные первой страницы
        data = self.get_category_data(path_parts, page=1)
        if not data:
            print(f"{indent}    [SKIP] Не удалось получить данные")
            return

        # Извлекаем информацию о категории
        section_meta = data.get("sectionMeta", {})
        breadcrumbs = section_meta.get("breadcrumbs", [])
        category_slug, category_name = self.extract_breadcrumbs_path(breadcrumbs)

        # Сохраняем маппинг
        if category_slug:
            self.categories[category_slug] = category_name

        # Проверяем подкатегории
        subsections = data.get("subsections", [])

        if subsections:
            # Есть подкатегории - рекурсивно обходим
            print(f"{indent}    Подкатегорий: {len(subsections)}")
            for sub in subsections:
                sub_url = sub.get("url", "")
                # Извлекаем путь: /catalog/a/b/c/ -> ["a", "b", "c"]
                match = re.search(r'/catalog/(.+?)/?$', sub_url)
                if match:
                    sub_path = match.group(1).rstrip('/').split('/')
                    self.crawl_category(sub_path, depth + 1)
        else:
            # Конечная категория - собираем товары
            self._collect_products_from_category(data, path_parts, category_slug, category_name, depth)

    def _collect_products_from_category(self, first_page_data: dict, path_parts: List[str],
                                        category_slug: str, category_name: str, depth: int):
        """Собрать все товары из категории (все страницы)"""
        indent = "  " * depth

        products_data = first_page_data.get("products", {})
        products_list = products_data.get("data", [])
        meta = products_data.get("meta", {})
        total_pages = meta.get("pageCount", 1)
        total_products = meta.get("total", 0)

        # Проверяем прогресс если есть координатор (для продолжения после бана)
        start_page = 1
        if self.coordinator and self.current_city:
            progress = self.coordinator.get_progress(self.current_city)
            if category_slug in progress:
                cat_progress = progress[category_slug]
                if cat_progress["status"] == "done":
                    print(f"{indent}    [SKIP] Категория уже спарсена")
                    return
                start_page = cat_progress["current_page"] + 1
                if start_page > 1:
                    print(f"{indent}    [RESUME] Продолжаем с страницы {start_page}")

        print(f"{indent}    Товаров: {total_products}, страниц: {total_pages}")

        category_products = []  # Товары этой категории для инкрементального сохранения

        # Обрабатываем первую страницу (если не пропускаем)
        if start_page <= 1:
            page_count = 0
            page_products = []
            for product in products_list:
                info = self.extract_product_info(product, category_slug, category_name)
                if info:
                    self.products.append(info)
                    category_products.append(info)
                    page_products.append(info)
                    self.products_since_refresh += 1
                    page_count += 1

            print(f"{indent}    Страница 1/{total_pages}: +{page_count} товаров")
            self.check_cookie_refresh()

            # Сохраняем после каждой страницы (для защиты от бана)
            if self.incremental_save and page_products and self.current_city_id:
                for p in page_products:
                    p["city_id"] = self.current_city_id
                    p["city_name"] = self.current_city
                save_staging(page_products, multi_city=True, append=True)

            # Сохраняем прогресс в координатор
            if self.coordinator and self.current_city:
                self.coordinator.save_progress(self.current_city, category_slug, 1, total_pages, page_count)

        # Обрабатываем остальные страницы
        for page in range(max(2, start_page), total_pages + 1):
            data = self.get_category_data(path_parts, page=page)
            if not data:
                continue

            products_data = data.get("products", {})
            products_list = products_data.get("data", [])

            page_count = 0
            page_products = []
            for product in products_list:
                info = self.extract_product_info(product, category_slug, category_name)
                if info:
                    self.products.append(info)
                    category_products.append(info)
                    page_products.append(info)
                    self.products_since_refresh += 1
                    page_count += 1

            print(f"{indent}    Страница {page}/{total_pages}: +{page_count} товаров")
            self.check_cookie_refresh()

            # Сохраняем после каждой страницы (для защиты от бана)
            if self.incremental_save and page_products and self.current_city_id:
                for p in page_products:
                    p["city_id"] = self.current_city_id
                    p["city_name"] = self.current_city
                save_staging(page_products, multi_city=True, append=True)

            # Сохраняем прогресс в координатор
            if self.coordinator and self.current_city:
                self.coordinator.save_progress(self.current_city, category_slug, page, total_pages, page_count)

        # Помечаем категорию как завершённую
        if self.coordinator and self.current_city:
            self.coordinator.mark_category_done(self.current_city, category_slug)

        # Логируем итог (без дублирования сохранения - уже сохранили постранично)
        if self.incremental_save and category_products:
            print(f"{indent}    [SAVE] Всего по категории: {len(category_products)} товаров")

    def parse_catalog(self, start_category: str = None, reparse_articles: bool = True):
        """Парсит весь каталог начиная с указанной категории"""
        start = start_category or ROOT_CATEGORY

        print(f"\n{'='*60}")
        print(f"Парсинг каталога GreenSpark.ru")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Корневая категория: {start}")
        print(f"{'='*60}\n")

        # Этап 1: Обход каталога
        print("[Этап 1] Обход каталога и сбор товаров\n")
        path_parts = start.split('/')
        self.crawl_category(path_parts)

        print(f"\n{'='*60}")
        print(f"Этап 1 завершён: {len(self.products)} товаров")
        print(f"Категорий: {len(self.categories)}")
        print(f"Ошибок: {len(self.errors)}")
        print(f"{'='*60}")

        # Этап 2: Допарсинг артикулов
        if reparse_articles:
            print("\n[Этап 2] Допарсинг артикулов")
            self.reparse_missing_articles()

        # Итого
        missing_articles = len([p for p in self.products if not p.get("article")])
        print(f"\n{'='*60}")
        print(f"ИТОГО: {len(self.products)} товаров")
        print(f"С артикулами: {len(self.products) - missing_articles}")
        print(f"Без артикулов: {missing_articles}")
        print(f"{'='*60}")

    def save_to_json(self, filename: str = None):
        """Сохранить в JSON"""
        filename = filename or PRODUCTS_JSON

        data = {
            "source": "greenspark.ru",
            "date": datetime.now().isoformat(),
            "total": len(self.products),
            "products": self.products,
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"Сохранено в {filename}: {len(self.products)} товаров")

    def save_to_excel(self, filename: str = None):
        """Сохранить в Excel"""
        filename = filename or PRODUCTS_XLSX

        if not self.products:
            print("Нет товаров для сохранения в Excel")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"

        # Заголовки
        headers = [
            "Артикул",
            "Наименование",
            "Цена",
            "Опт",
            "Категория",
            "URL"
        ]

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Записываем данные
        for row, product in enumerate(self.products, 2):
            ws.cell(row=row, column=1, value=product.get("article", ""))
            ws.cell(row=row, column=2, value=product.get("name", ""))
            ws.cell(row=row, column=3, value=product.get("price", 0))
            ws.cell(row=row, column=4, value=product.get("price_wholesale", 0))
            ws.cell(row=row, column=5, value=product.get("category", ""))
            ws.cell(row=row, column=6, value=product.get("url", ""))

        # Автоширина колонок
        column_widths = [15, 70, 12, 12, 50, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(filename)
        print(f"Сохранено в {filename}: {len(self.products)} товаров")

    def save_categories(self, filename: str = None):
        """Сохранить маппинг категорий"""
        filename = filename or CATEGORIES_JSON

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.categories, f, ensure_ascii=False, indent=2)

        print(f"Категории сохранены в {filename}")

    def save_errors(self, filename: str = None):
        """Сохранить лог ошибок"""
        filename = filename or ERRORS_LOG

        if self.errors:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.errors, f, ensure_ascii=False, indent=2)
            print(f"Ошибки сохранены в {filename}")

    def close(self):
        """Закрыть соединение"""
        self.client.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def ensure_outlet():
    """Создаёт outlet для GreenSpark если не существует"""
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO outlets (code, city, name, is_active)
            VALUES ('greenspark-online', 'Москва', 'GreenSpark Online', true)
            ON CONFLICT (code) DO NOTHING
        """)
        conn.commit()
    finally:
        cur.close()
        conn.close()


def ensure_outlets_for_cities():
    """Создаёт outlets для всех городов GreenSpark"""
    cities_path = os.path.join(os.path.dirname(__file__), "data", "greenspark_cities.json")
    with open(cities_path, 'r', encoding='utf-8') as f:
        cities = json.load(f)

    conn = get_db()
    cur = conn.cursor()
    try:
        for city in cities:
            city_id = city["set_city"]
            city_name = city["name"]
            outlet_code = f"greenspark-{city_id}"

            cur.execute("""
                INSERT INTO outlets (code, city, name, is_active)
                VALUES (%s, %s, %s, true)
                ON CONFLICT (code) DO UPDATE SET city = EXCLUDED.city, name = EXCLUDED.name
            """, (outlet_code, city_name, f"GreenSpark {city_name}"))

        conn.commit()
        print(f"Создано/обновлено {len(cities)} outlets для GreenSpark")
    finally:
        cur.close()
        conn.close()


def clear_staging():
    """Очистить staging таблицу"""
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("TRUNCATE TABLE staging")
        conn.commit()
        print("Staging очищен")
    finally:
        cur.close()
        conn.close()


def get_staging_count() -> int:
    """Получить количество записей в staging"""
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM staging")
        return cur.fetchone()[0]
    finally:
        cur.close()
        conn.close()


def get_parsed_city_ids() -> set:
    """Получить ID городов у которых есть данные в current_prices"""
    conn = get_db()
    cur = conn.cursor()
    try:
        # Извлекаем city_id из outlet_code (greenspark-{city_id})
        cur.execute("""
            SELECT DISTINCT
                CAST(REPLACE(o.code, 'greenspark-', '') AS INTEGER) as city_id
            FROM current_prices cp
            JOIN outlets o ON o.id = cp.outlet_id
            WHERE o.code LIKE 'greenspark-%'
              AND o.code != 'greenspark-online'
        """)
        return {row[0] for row in cur.fetchall()}
    except Exception as e:
        print(f"Ошибка получения спарсенных городов: {e}")
        return set()
    finally:
        cur.close()
        conn.close()


def save_staging(products: List[Dict], multi_city: bool = False, append: bool = False):
    """Сохранение товаров в staging таблицу

    Args:
        append: Если True, не очищать staging (добавлять к существующим)
    """
    if not products:
        print("Нет товаров для сохранения в staging")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        # Очищаем staging только если не append режим
        if not append:
            cur.execute("TRUNCATE TABLE staging")

        # Вставляем товары
        insert_sql = """
            INSERT INTO staging (
                outlet_code, name, article, category,
                price, price_wholesale, url
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """

        for p in products:
            # Определяем outlet_code в зависимости от режима
            if multi_city and p.get("city_id"):
                outlet_code = f"greenspark-{p.get('city_id')}"
            else:
                outlet_code = "greenspark-online"

            cur.execute(insert_sql, (
                outlet_code,
                p.get("name", ""),
                p.get("article", ""),
                p.get("category", ""),
                p.get("price", 0),
                p.get("price_wholesale", 0),
                p.get("url", ""),
            ))

        conn.commit()
        print(f"Сохранено в staging: {len(products)} товаров")
    finally:
        cur.close()
        conn.close()


def process_staging():
    """Обработка staging: UPSERT в nomenclature и current_prices"""
    conn = get_db()
    cur = conn.cursor()
    try:
        # Убеждаемся что outlet существует
        ensure_outlet()

        # 1. UPSERT в nomenclature
        cur.execute("""
            INSERT INTO nomenclature (article, name, category, first_seen_at, updated_at)
            SELECT DISTINCT ON (article)
                article, name, category, NOW(), NOW()
            FROM staging
            WHERE article IS NOT NULL AND article != ''
            ON CONFLICT (article) DO UPDATE SET
                name = EXCLUDED.name,
                category = EXCLUDED.category,
                updated_at = NOW()
        """)
        nom_count = cur.rowcount
        print(f"Nomenclature: {nom_count} записей обновлено/добавлено")

        # 2. UPSERT в current_prices
        cur.execute("""
            INSERT INTO current_prices (nomenclature_id, outlet_id, price, price_wholesale, updated_at)
            SELECT DISTINCT ON (n.id, o.id)
                n.id, o.id, s.price, s.price_wholesale, NOW()
            FROM staging s
            JOIN nomenclature n ON n.article = s.article
            JOIN outlets o ON o.code = s.outlet_code
            WHERE s.article IS NOT NULL AND s.article != ''
            ON CONFLICT (nomenclature_id, outlet_id) DO UPDATE SET
                price = EXCLUDED.price,
                price_wholesale = EXCLUDED.price_wholesale,
                updated_at = NOW()
        """)
        price_count = cur.rowcount
        print(f"Current prices: {price_count} записей обновлено/добавлено")

        conn.commit()

        # Статистика
        cur.execute("SELECT COUNT(*) FROM nomenclature")
        total_nom = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM current_prices")
        total_prices = cur.fetchone()[0]

        print(f"\n=== Итого в БД ===")
        print(f"Номенклатура: {total_nom}")
        print(f"Цены: {total_prices}")

    finally:
        cur.close()
        conn.close()


def save_to_db(products: List[Dict], multi_city: bool = False):
    """
    Сохранение в новую схему БД: greenspark_nomenclature + greenspark_prices
    UPSERT по article (уникальный ключ товара)
    product_url сохраняется в greenspark_prices
    """
    if not products:
        print("Нет товаров для сохранения")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        # Создаём outlets если нужно
        if multi_city:
            ensure_outlets_for_cities()
        else:
            ensure_outlet()

        # Кэш outlet_id по outlet_code
        cur.execute("SELECT code, id FROM outlets WHERE code LIKE 'greenspark-%'")
        outlet_cache = {row[0]: row[1] for row in cur.fetchall()}

        nom_inserted = 0
        nom_updated = 0
        prices_upserted = 0

        for p in products:
            article = p.get("article", "").strip()
            if not article:
                continue

            name = p.get("name", "").strip()
            if not name:
                continue

            # Определяем outlet
            if multi_city and p.get("city_id"):
                outlet_code = f"greenspark-{p.get('city_id')}"
            else:
                outlet_code = "greenspark-online"

            outlet_id = outlet_cache.get(outlet_code)
            if not outlet_id:
                continue

            # URL товара
            product_url = p.get("url", "")
            if not product_url:
                product_url = f"https://green-spark.ru/product/{article}"

            category = p.get("category", "").strip() or None

            # 1. UPSERT в greenspark_nomenclature (без product_url)
            cur.execute("""
                INSERT INTO greenspark_nomenclature (name, article, category, first_seen_at, updated_at)
                VALUES (%s, %s, %s, NOW(), NOW())
                ON CONFLICT (article) DO UPDATE SET
                    name = EXCLUDED.name,
                    category = EXCLUDED.category,
                    updated_at = NOW()
                RETURNING id, (xmax = 0) as inserted
            """, (name, article, category))

            row = cur.fetchone()
            nomenclature_id = row[0]
            if row[1]:
                nom_inserted += 1
            else:
                nom_updated += 1

            # 2. UPSERT в greenspark_prices (с product_url)
            price = p.get("price", 0)
            price_wholesale = p.get("price_wholesale", 0)

            cur.execute("""
                INSERT INTO greenspark_prices (nomenclature_id, outlet_id, price, price_wholesale, product_url, updated_at)
                VALUES (%s, %s, %s, %s, %s, NOW())
                ON CONFLICT (nomenclature_id, outlet_id) DO UPDATE SET
                    price = EXCLUDED.price,
                    price_wholesale = EXCLUDED.price_wholesale,
                    product_url = EXCLUDED.product_url,
                    updated_at = NOW()
            """, (nomenclature_id, outlet_id, price, price_wholesale, product_url))
            prices_upserted += 1

        conn.commit()

        print(f"\n=== Сохранено в БД (новая схема) ===")
        print(f"greenspark_nomenclature: +{nom_inserted} новых, ~{nom_updated} обновлено")
        print(f"greenspark_prices: {prices_upserted} записей")

        # Итоговая статистика
        cur.execute("SELECT COUNT(*) FROM greenspark_nomenclature")
        total_nom = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM greenspark_prices")
        total_prices = cur.fetchone()[0]
        print(f"\nИтого в БД: {total_nom} товаров, {total_prices} цен")

    finally:
        cur.close()
        conn.close()


def main():
    arg_parser = argparse.ArgumentParser(description='Парсер каталога GreenSpark.ru')
    arg_parser.add_argument('--category', '-c', type=str, default=None,
                           help='Категория для парсинга (по умолчанию: komplektuyushchie_dlya_remonta)')
    arg_parser.add_argument('--shop', '-s', type=int, default=DEFAULT_SHOP_ID,
                           help=f'ID магазина (по умолчанию: {DEFAULT_SHOP_ID})')
    arg_parser.add_argument('--no-reparse', action='store_true',
                           help='Пропустить допарсинг артикулов')
    arg_parser.add_argument('--all', action='store_true',
                           help='Полный парсинг: сбор + сохранение в БД + обработка')
    arg_parser.add_argument('--all-cities', action='store_true',
                           help='Парсить все города (60 городов)')
    arg_parser.add_argument('--save-each', action='store_true',
                           help='Сохранять в БД после каждого города (защита от потери данных)')
    arg_parser.add_argument('--process', action='store_true',
                           help='Только обработка staging (без парсинга)')
    arg_parser.add_argument('--no-db', action='store_true',
                           help='Не сохранять в БД (только JSON/Excel)')
    arg_parser.add_argument('--old-schema', action='store_true',
                           help='Использовать старую схему БД (staging -> nomenclature)')
    arg_parser.add_argument('--resume-from', type=int, default=None,
                           help='ID города с которого продолжить парсинг')
    arg_parser.add_argument('--skip-parsed', action='store_true',
                           help='Пропускать города у которых уже есть данные в БД')
    arg_parser.add_argument('--incremental', action='store_true',
                           help='Сохранять в БД после каждой категории (защита от бана)')
    arg_parser.add_argument('--proxy', type=str, default=None,
                           help='HTTP прокси (например: http://proxy:8080)')
    arg_parser.add_argument('--proxy-file', type=str, default=None,
                           help='Файл с прокси для ротации (ip:port на каждой строке)')
    arg_parser.add_argument('--ips', type=str, default=None,
                           help='Список IP для переключения через запятую (например: 155.212.221.189,217.114.14.17)')
    # Аргументы для координатора (эстафета между серверами)
    arg_parser.add_argument('--coordinated', action='store_true',
                           help='Режим координации: брать города из очереди, передавать эстафету при бане')
    arg_parser.add_argument('--server', type=str, default=os.environ.get('PARSER_SERVER_NAME', 'server-a'),
                           help='Имя сервера для координации (по умолчанию: server-a)')
    args = arg_parser.parse_args()

    # Парсим список IP
    ips_list = []
    if args.ips:
        ips_list = [ip.strip() for ip in args.ips.split(',')]
        print(f"Доступные IP для переключения: {ips_list}")

    # Только обработка staging
    if args.process:
        print("Обработка staging...")
        process_staging()
        print("\nОбработка завершена!")
        return

    # === РЕЖИМ КООРДИНАЦИИ (эстафета между серверами) ===
    if args.coordinated:
        if not COORDINATOR_AVAILABLE:
            print("[ERROR] Координатор недоступен. Установите coordinator.py")
            return

        print(f"\n{'='*60}")
        print(f"РЕЖИМ КООРДИНАЦИИ - Сервер: {args.server}")
        print(f"{'='*60}\n")

        coordinator = ParserCoordinator(args.server)
        coordinator.register_server()

        # Цикл: берём города из очереди пока есть
        while True:
            result = coordinator.take_city()
            if not result:
                print("[COORD] Очередь пуста. Завершаем работу.")
                break

            city_name, city_id = result
            print(f"\n[COORD] Парсим город: {city_name} (ID: {city_id})")

            try:
                with GreenSparkCatalogParser(
                    shop_id=args.shop,
                    proxy=args.proxy,
                    ips=ips_list,
                    coordinator=coordinator,
                    server_name=args.server,
                    proxy_file=args.proxy_file,
                    use_proxy=bool(args.proxy_file)
                ) as parser:
                    parser.incremental_save = True  # Всегда сохраняем постранично
                    parser.set_city(city_id, city_name)
                    parser.parse_catalog(
                        start_category=args.category,
                        reparse_articles=not args.no_reparse
                    )

                    # Город завершён
                    coordinator.complete_city(city_name)
                    print(f"[COORD] Город {city_name} завершён: {len(parser.products)} товаров")

            except SystemExit as e:
                # Эстафета передана другому серверу
                print(f"[COORD] {e}")
                break
            except Exception as e:
                print(f"[COORD] Ошибка при парсинге города {city_name}: {e}")
                continue

        coordinator.close()
        print("\n[COORD] Координированный парсинг завершён!")
        return

    # === ОБЫЧНЫЙ РЕЖИМ (без координации) ===
    coordinator = None
    if COORDINATOR_AVAILABLE and args.server:
        # Создаём координатор для сохранения прогресса (но без очереди)
        try:
            coordinator = ParserCoordinator(args.server)
        except Exception as e:
            print(f"[WARN] Не удалось создать координатор: {e}")
            coordinator = None

    with GreenSparkCatalogParser(
        shop_id=args.shop,
        proxy=args.proxy,
        ips=ips_list,
        coordinator=coordinator,
        server_name=args.server,
        proxy_file=args.proxy_file,
        use_proxy=bool(args.proxy_file)
    ) as parser:
        if args.all_cities:
            # Мультигородской парсинг
            parser.parse_all_cities(
                start_category=args.category,
                reparse_articles=not args.no_reparse,
                save_after_each=args.save_each and not args.no_db,
                process_db=args.all and args.save_each,
                resume_from_city=args.resume_from,
                skip_parsed=args.skip_parsed,
                incremental_save=args.incremental and not args.no_db
            )
        else:
            # Одиночный город
            parser.parse_catalog(start_category=args.category, reparse_articles=not args.no_reparse)

        # Сохраняем в файлы
        parser.save_to_json()
        parser.save_to_excel()
        parser.save_categories()
        parser.save_errors()

        # Сохраняем в БД
        if not args.no_db:
            if args.old_schema:
                # LEGACY: staging -> nomenclature -> current_prices
                if args.all_cities:
                    ensure_outlets_for_cities()
                    save_staging(parser.products, multi_city=True)
                else:
                    save_staging(parser.products, multi_city=False)

                if args.all:
                    process_staging()
            else:
                # НОВАЯ СХЕМА: greenspark_nomenclature + greenspark_prices
                save_to_db(parser.products, multi_city=args.all_cities)

    if coordinator:
        coordinator.close()

    print("\nПарсинг завершён!")


if __name__ == "__main__":
    main()
