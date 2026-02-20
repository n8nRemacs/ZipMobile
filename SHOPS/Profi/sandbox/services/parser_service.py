"""
Async обёртка ProfiParser — скачивание Excel + парсинг → staging → nomenclature + prices
"""
from __future__ import annotations
import asyncio
import json
import os
import sys
import tempfile
from uuid import UUID

import httpx

# Add parent dirs so we can import ProfiParser and configs
_profi_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _profi_dir not in sys.path:
    sys.path.insert(0, os.path.join(_profi_dir, "Profi"))
    sys.path.insert(0, _profi_dir)

from price_lists_config import PRICE_LISTS, get_info_by_outlet
from fetch_price_lists import fetch_price_lists

from ..db import get_pool
from ..tasks import update_task
from ..config import settings


async def _download_file(client: httpx.AsyncClient, url: str) -> str | None:
    """Download Excel file, return temp path."""
    try:
        resp = await client.get(url, follow_redirects=True)
        resp.raise_for_status()
        suffix = ".xls" if ".xls" in url.lower() and ".xlsx" not in url.lower() else ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
            f.write(resp.content)
            return f.name
    except Exception:
        return None


def _parse_excel_sync(file_path: str, city: str, shop: str, outlet_code: str) -> list[dict]:
    """Sync Excel parsing using openpyxl with font-size category detection."""
    from openpyxl import load_workbook
    import re

    FONT_SIZE_BRAND = 11
    FONT_SIZE_MODEL = 10
    FONT_SIZE_PART_TYPE = 9

    products = []
    current_brand = ""
    current_model = ""
    current_part_type = ""

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Find header row
        header_row = None
        for r in range(1, min(ws.max_row + 1, 20)):
            for c in range(1, min(ws.max_column or 50, 50) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "наимен" in v.lower():
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            return products

        # Resolve columns
        col_map = {}
        for c in range(1, (ws.max_column or 50) + 1):
            v = ws.cell(header_row, c).value
            if not isinstance(v, str):
                continue
            vl = v.strip().lower().replace("ё", "е")
            if "наимен" in vl:
                col_map["name"] = c
            elif "артик" in vl:
                col_map["article"] = c
            elif "цен" in vl or "розн" in vl:
                col_map["price"] = c

        name_col = col_map.get("name")
        article_col = col_map.get("article")
        price_col = col_map.get("price")

        if not name_col:
            return products

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            name_cell = row[name_col - 1] if name_col and len(row) >= name_col else None
            if not name_cell or not name_cell.value:
                continue

            name_val = str(name_cell.value).strip()
            if not name_val:
                continue

            # Detect font size for category
            font_sz = None
            try:
                if name_cell.font and name_cell.font.sz:
                    font_sz = int(round(float(name_cell.font.sz)))
            except Exception:
                pass

            if font_sz and font_sz >= FONT_SIZE_BRAND:
                current_brand = name_val
                current_model = ""
                current_part_type = ""
                continue
            elif font_sz and font_sz == FONT_SIZE_MODEL:
                current_model = name_val
                current_part_type = ""
                continue
            elif font_sz and font_sz == FONT_SIZE_PART_TYPE:
                current_part_type = name_val
                continue

            # Product row
            article = ""
            if article_col and len(row) >= article_col:
                av = row[article_col - 1].value
                if av:
                    article = str(av).strip()

            price = 0.0
            if price_col and len(row) >= price_col:
                pv = row[price_col - 1].value
                if pv is not None:
                    price_match = re.search(r'([\d\s]+[.,]?\d*)', str(pv))
                    if price_match:
                        try:
                            price = float(price_match.group(1).replace(' ', '').replace(',', '.'))
                        except Exception:
                            pass

            products.append({
                "article": article,
                "name": name_val,
                "brand_raw": current_brand,
                "model_raw": current_model,
                "part_type_raw": current_part_type,
                "price": price,
                "city": city,
                "shop": shop,
                "outlet_code": outlet_code,
            })

        wb.close()
    except Exception:
        pass
    finally:
        try:
            os.unlink(file_path)
        except Exception:
            pass

    return products


async def _save_products(products: list[dict]):
    """UPSERT products into profi_nomenclature + profi_prices."""
    if not products:
        return

    pool = get_pool()

    # Upsert nomenclature
    for p in products:
        if not p.get("article"):
            continue
        await pool.execute(
            """INSERT INTO profi_nomenclature (article, name, brand_raw, model_raw, part_type_raw)
               VALUES ($1, $2, $3, $4, $5)
               ON CONFLICT (article) DO UPDATE SET
                   name = EXCLUDED.name,
                   brand_raw = EXCLUDED.brand_raw,
                   model_raw = EXCLUDED.model_raw,
                   part_type_raw = EXCLUDED.part_type_raw,
                   updated_at = NOW()""",
            p["article"], p["name"], p["brand_raw"], p["model_raw"], p["part_type_raw"],
        )

    # Upsert prices
    for p in products:
        if not p.get("article") or not p.get("price"):
            continue
        await pool.execute(
            """INSERT INTO profi_prices (article, outlet_code, city, price)
               VALUES ($1, $2, $3, $4)
               ON CONFLICT (article, outlet_code) DO UPDATE SET
                   price = EXCLUDED.price,
                   updated_at = NOW()""",
            p["article"], p.get("outlet_code", ""), p.get("city", ""), p["price"],
        )


async def parse_outlet(task_id: UUID, outlet_code: str):
    """Parse a single outlet by code."""
    info = get_info_by_outlet(outlet_code)
    if not info:
        raise ValueError(f"Unknown outlet: {outlet_code}")

    await update_task(task_id, progress={"message": f"Downloading {info['shop']}..."})

    async with httpx.AsyncClient(timeout=settings.http_timeout) as client:
        file_path = await _download_file(client, info["url"])

    if not file_path:
        raise RuntimeError(f"Failed to download {info['url']}")

    await update_task(task_id, progress={"message": f"Parsing {info['shop']}..."})
    products = await asyncio.to_thread(
        _parse_excel_sync, file_path, info["city"], info["shop"], outlet_code
    )

    await update_task(task_id, progress={"message": f"Saving {len(products)} products..."})
    await _save_products(products)
    await update_task(task_id, result={"products_count": len(products), "outlet": outlet_code})


async def parse_all_outlets(task_id: UUID):
    """Parse all configured outlets."""
    sem = asyncio.Semaphore(settings.max_concurrent_downloads)
    total = len(PRICE_LISTS)
    completed = 0
    total_products = 0

    async def _process_one(pl: dict):
        nonlocal completed, total_products
        async with sem:
            async with httpx.AsyncClient(timeout=settings.http_timeout) as client:
                file_path = await _download_file(client, pl["url"])

            if file_path:
                products = await asyncio.to_thread(
                    _parse_excel_sync, file_path,
                    pl["city"], pl["shop"], pl.get("outlet_code", ""),
                )
                await _save_products(products)
                total_products += len(products)

            completed += 1
            await update_task(task_id, progress={
                "current": completed, "total": total,
                "message": f"{pl['city']} - {pl['shop']}",
            })

    tasks = [_process_one(pl) for pl in PRICE_LISTS]
    await asyncio.gather(*tasks, return_exceptions=True)
    await update_task(task_id, result={"outlets": total, "products": total_products})


async def parse_dynamic(task_id: UUID):
    """Discover URLs dynamically from siriust.ru and parse."""
    await update_task(task_id, progress={"message": "Discovering URLs..."})
    price_lists = await asyncio.to_thread(fetch_price_lists)

    sem = asyncio.Semaphore(settings.max_concurrent_downloads)
    total = len(price_lists)
    completed = 0
    total_products = 0

    async def _process_one(pl: dict):
        nonlocal completed, total_products
        async with sem:
            async with httpx.AsyncClient(timeout=settings.http_timeout) as client:
                file_path = await _download_file(client, pl["url"])

            if file_path:
                outlet_code = pl.get("outlet_code", pl["url"].split("/")[-1].replace(".xls", ""))
                products = await asyncio.to_thread(
                    _parse_excel_sync, file_path,
                    pl["city"], pl["shop"], outlet_code,
                )
                await _save_products(products)
                total_products += len(products)

            completed += 1
            await update_task(task_id, progress={
                "current": completed, "total": total,
                "message": f"{pl['city']} - {pl['shop']}",
            })

    tasks = [_process_one(pl) for pl in price_lists]
    await asyncio.gather(*tasks, return_exceptions=True)
    await update_task(task_id, result={
        "outlets_discovered": total,
        "products": total_products,
    })
