"""
Парсер прайс-листов Profi (siriust.ru) - все города

База данных: db_profi
Таблицы: staging, outlets, nomenclature, current_prices, price_history

Поддерживает динамическую загрузку списка прайс-листов с siriust.ru
"""

import json
import csv
import re
import os
import tempfile
import subprocess
import psycopg2
import argparse
import httpx
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import load_workbook

from config import DATA_DIR, PRODUCTS_CSV, PRODUCTS_JSON
from price_lists_config import PRICE_LISTS, get_cities
from fetch_price_lists import fetch_price_lists, extract_city_from_url

# === Конфигурация БД (Supabase) ===
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db_wrapper import get_db  # Автоматически маппит таблицы на новые имена

# === Размеры шрифта для определения категорий ===
FONT_SIZE_BRAND = 11      # Бренд
FONT_SIZE_MODEL = 10      # Модель
FONT_SIZE_PART_TYPE = 9   # Тип запчасти


class ProfiParser:
    """Парсер прайс-листов Profi для всех городов"""

    def __init__(self):
        self.products: List[Dict] = []
        self.outlets_parsed: List[Dict] = []
        self.errors: List[Dict] = []
        os.makedirs(DATA_DIR, exist_ok=True)

    def download_price_list(self, url: str) -> Optional[str]:
        """Скачивает прайс-лист и возвращает путь к временному файлу"""
        try:
            with httpx.Client(timeout=60, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()

                # Сохраняем во временный файл
                suffix = ".xls" if ".xls" in url.lower() else ".xlsx"
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
                    f.write(response.content)
                    return f.name
        except Exception as e:
            self.errors.append({
                "url": url,
                "error": str(e),
                "time": datetime.now().isoformat()
            })
            return None

    def parse_price(self, price_text) -> float:
        """Парсит цену из текста"""
        if price_text is None:
            return 0.0
        price_match = re.search(r'([\d\s]+[.,]?\d*)', str(price_text))
        if price_match:
            price_str = price_match.group(1).replace(' ', '').replace(',', '.')
            try:
                return float(price_str)
            except:
                return 0.0
        return 0.0

    def _get_font_size(self, cell) -> Optional[int]:
        """Получает размер шрифта ячейки"""
        try:
            if cell and cell.font and cell.font.sz:
                return int(round(float(cell.font.sz)))
        except:
            pass
        return None

    def _canon(self, s: str) -> str:
        """Канонизация строки для поиска заголовков"""
        s = s.strip().lower().replace("ё", "е")
        return re.sub(r"[^a-z0-9а-я]+", "", s)

    def _find_header_row(self, ws) -> Optional[int]:
        """Ищет строку с заголовками (содержит 'наимен')"""
        for r in range(1, min(ws.max_row + 1, 20)):
            for c in range(1, min(ws.max_column, 50) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "наимен" in v.lower():
                    return r
        return None

    def _resolve_columns(self, ws, header_row: int) -> Dict[str, Optional[int]]:
        """Определяет индексы колонок по заголовкам"""
        headers_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip():
                headers_map[self._canon(v)] = c

        # Синонимы для поиска колонок
        targets = {
            "name": ["наимен", "наименование", "name"],
            "article": ["артикул", "код", "code", "sku"],
            "barcode": ["штрихкод", "штрихкод", "barcode", "ean"],
            "price": ["цена", "розница", "retail", "price"],
            "unit": ["ед", "единица", "unit"],
            "quantity": ["количество", "наличие", "остаток", "stock", "qty"],
            "warehouse": ["склад", "warehouse"],
            "note": ["примечание", "комментарий", "comment", "note"],
        }

        resolved = {}
        for target, keys in targets.items():
            found = None
            for key in keys:
                for hdr_canon, col in headers_map.items():
                    if key in hdr_canon:
                        found = col
                        break
                if found:
                    break
            resolved[target] = found

        return resolved

    def _row_is_empty(self, ws, row: int, cols: List[Optional[int]]) -> bool:
        """Проверяет пустая ли строка по ключевым колонкам"""
        for col in cols:
            if col:
                v = ws.cell(row, col).value
                if v not in (None, "", " "):
                    return False
        return True

    def _convert_xls_to_xlsx(self, xls_path: str) -> Optional[str]:
        """Конвертирует .xls в .xlsx через LibreOffice"""
        try:
            # Создаём временную директорию для выходного файла
            out_dir = tempfile.mkdtemp()

            # Конвертируем через soffice
            result = subprocess.run([
                'soffice', '--headless', '--convert-to', 'xlsx',
                '--outdir', out_dir, xls_path
            ], capture_output=True, text=True, timeout=60)

            if result.returncode != 0:
                return None

            # Находим сконвертированный файл
            base_name = os.path.splitext(os.path.basename(xls_path))[0]
            xlsx_path = os.path.join(out_dir, f"{base_name}.xlsx")

            if os.path.exists(xlsx_path):
                return xlsx_path
            return None
        except Exception as e:
            return None

    def parse_excel_file(self, file_path: str, city: str, shop: str, outlet_code: str) -> List[Dict]:
        """Парсит Excel файл с разбором категорий по размеру шрифта"""
        products = []
        converted_path = None

        # Если файл .xls - конвертируем в .xlsx
        if file_path.lower().endswith('.xls') and not file_path.lower().endswith('.xlsx'):
            converted_path = self._convert_xls_to_xlsx(file_path)
            if converted_path:
                file_path = converted_path
            else:
                self.errors.append({
                    "file": file_path,
                    "error": "Failed to convert .xls to .xlsx",
                    "time": datetime.now().isoformat()
                })
                return []

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            self.errors.append({
                "file": file_path,
                "error": f"Failed to read Excel: {e}",
                "time": datetime.now().isoformat()
            })
            if converted_path:
                try:
                    os.unlink(converted_path)
                    os.rmdir(os.path.dirname(converted_path))
                except:
                    pass
            return []

        # Находим строку с заголовками
        header_row = self._find_header_row(ws)
        if not header_row:
            self.errors.append({
                "file": file_path,
                "error": "Header row not found",
                "time": datetime.now().isoformat()
            })
            return []

        # Определяем колонки
        cols = self._resolve_columns(ws, header_row)
        name_col = cols.get("name")
        if not name_col:
            self.errors.append({
                "file": file_path,
                "error": "Name column not found",
                "time": datetime.now().isoformat()
            })
            return []

        article_col = cols.get("article")
        price_col = cols.get("price")
        quantity_col = cols.get("quantity")

        # Колонки для проверки пустых строк
        check_cols = [name_col, article_col, price_col, quantity_col]

        # Текущие категории (определяются по размеру шрифта)
        current_brand = None
        current_model = None
        current_part_type = None

        for r in range(header_row + 1, ws.max_row + 1):
            # Пропускаем пустые строки
            if self._row_is_empty(ws, r, check_cols):
                continue

            name_cell = ws.cell(r, name_col)
            name_val = name_cell.value
            if not name_val:
                continue
            name_val = str(name_val).strip()
            if not name_val:
                continue

            # Определяем тип строки по размеру шрифта
            font_size = self._get_font_size(name_cell)

            if font_size == FONT_SIZE_BRAND:
                current_brand = name_val
                current_model = None  # Сброс модели при новом бренде
                current_part_type = None
                continue
            elif font_size == FONT_SIZE_MODEL:
                current_model = name_val
                current_part_type = None  # Сброс типа при новой модели
                continue
            elif font_size == FONT_SIZE_PART_TYPE:
                current_part_type = name_val
                continue

            # Это товар - извлекаем данные
            article = ""
            if article_col:
                article_val = ws.cell(r, article_col).value
                if article_val:
                    article = str(article_val).strip()

            price = 0.0
            if price_col:
                price = self.parse_price(ws.cell(r, price_col).value)

            # Формируем категорию из иерархии
            category_parts = [p for p in [current_brand, current_model, current_part_type] if p]
            category = " / ".join(category_parts) if category_parts else ""

            products.append({
                'product_id': article,
                'article': article,
                'name': name_val,
                'price': price,
                'category': category,
                'brand': current_brand or "",
                'model': current_model or "",
                'part_type': current_part_type or "",
                'city': city,
                'shop': shop,
                'outlet_code': outlet_code,
                'url': '',
            })

        try:
            wb.close()
        except:
            pass

        # Удаляем временный сконвертированный файл
        if converted_path:
            try:
                os.unlink(converted_path)
                os.rmdir(os.path.dirname(converted_path))
            except:
                pass

        return products

    def parse_single_outlet(self, price_list: Dict) -> int:
        """Парсит один прайс-лист (один outlet)"""
        url = price_list.get("url", "")
        city = price_list.get("city", "")
        shop = price_list.get("shop", "")
        outlet_code = price_list.get("outlet_code", "")

        if not outlet_code:
            # Генерируем outlet_code из URL
            filename = url.split("/")[-1].replace(".xls", "").replace(".xlsx", "")
            outlet_code = f"profi-{filename.lower().replace(' ', '-')}"

        print(f"  [{city}] {shop}...")

        # Скачиваем файл
        file_path = self.download_price_list(url)
        if not file_path:
            print(f"    [SKIP] Не удалось скачать")
            return 0

        try:
            # Парсим
            products = self.parse_excel_file(file_path, city, shop, outlet_code)
            self.products.extend(products)

            self.outlets_parsed.append({
                "city": city,
                "shop": shop,
                "outlet_code": outlet_code,
                "products_count": len(products)
            })

            print(f"    +{len(products)} товаров")
            return len(products)
        finally:
            # Удаляем временный файл
            try:
                os.unlink(file_path)
            except:
                pass

    def parse_all_outlets(self, use_dynamic: bool = False):
        """Парсит все прайс-листы (все города)"""
        print(f"\n{'='*60}")
        print(f"Парсинг прайс-листов Profi (siriust.ru)")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}\n")

        # Получаем список прайс-листов
        if use_dynamic:
            print("[Режим] Динамическая загрузка списка с сайта\n")
            price_lists = fetch_price_lists()
            # Добавляем outlet_code
            for pl in price_lists:
                filename = pl["url"].split("/")[-1].replace(".xls", "").replace(".xlsx", "")
                pl["outlet_code"] = f"profi-{filename.lower().replace(' ', '-').replace('%20', '-')}"
        else:
            print(f"[Режим] Статический список ({len(PRICE_LISTS)} прайс-листов)\n")
            price_lists = PRICE_LISTS

        total_products = 0
        for pl in price_lists:
            count = self.parse_single_outlet(pl)
            total_products += count

        print(f"\n{'='*60}")
        print(f"ИТОГО: {total_products} товаров из {len(self.outlets_parsed)} торговых точек")
        print(f"Ошибок: {len(self.errors)}")
        print(f"{'='*60}")

    def save_to_csv(self, filename: str = None):
        """Сохраняет товары в CSV"""
        filename = filename or PRODUCTS_CSV

        if not self.products:
            print("Нет товаров для сохранения")
            return

        fieldnames = ["article", "name", "price", "category", "city", "shop", "outlet_code"]

        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';', extrasaction='ignore')
            writer.writeheader()
            writer.writerows(self.products)

        print(f"Сохранено в {filename}: {len(self.products)} товаров")

    def save_to_json(self, filename: str = None):
        """Сохраняет товары в JSON"""
        filename = filename or PRODUCTS_JSON

        data = {
            "source": "siriust.ru (Profi)",
            "date": datetime.now().isoformat(),
            "total": len(self.products),
            "outlets": self.outlets_parsed,
            "products": self.products,
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"Сохранено в {filename}")

    def print_stats(self):
        """Выводит статистику"""
        print(f"\n{'='*60}")
        print("СТАТИСТИКА")
        print(f"{'='*60}")

        # По городам
        city_counts = {}
        for p in self.products:
            city = p.get("city", "Неизвестно")
            city_counts[city] = city_counts.get(city, 0) + 1

        print(f"\nГородов: {len(city_counts)}")
        print("\nТоп-10 городов по товарам:")
        for city, count in sorted(city_counts.items(), key=lambda x: -x[1])[:10]:
            print(f"  {count:5d} | {city}")


        # Ценовой диапазон
        prices = [p.get("price", 0) for p in self.products if p.get("price", 0) > 0]
        if prices:
            print(f"\nЦены: от {min(prices):.0f} до {max(prices):.0f} руб")
            print(f"Средняя: {sum(prices)/len(prices):.0f} руб")


def ensure_outlets(outlets: List[Dict]):
    """Создаёт outlets для всех торговых точек Profi"""
    if not outlets:
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        for outlet in outlets:
            cur.execute("""
                INSERT INTO outlets (code, city, name, is_active)
                VALUES (%s, %s, %s, true)
                ON CONFLICT (code) DO UPDATE SET
                    city = EXCLUDED.city,
                    name = EXCLUDED.name
            """, (outlet["outlet_code"], outlet["city"], outlet["shop"]))
        conn.commit()
        print(f"Создано/обновлено {len(outlets)} outlets для Profi")
    finally:
        cur.close()
        conn.close()


def save_staging(products: List[Dict]):
    """Сохранение товаров в staging таблицу"""
    if not products:
        print("Нет товаров для сохранения в staging")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        # Очищаем staging для Profi
        cur.execute("DELETE FROM staging WHERE outlet_code LIKE 'profi-%'")

        insert_sql = """
            INSERT INTO staging (
                outlet_code, name, article, category,
                brand, model, part_type,
                price, url
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        for p in products:
            cur.execute(insert_sql, (
                p.get("outlet_code", ""),
                p.get("name", ""),
                p.get("article", ""),
                p.get("category", ""),
                p.get("brand", ""),
                p.get("model", ""),
                p.get("part_type", ""),
                p.get("price", 0),
                p.get("url", ""),
            ))

        conn.commit()
        print(f"Сохранено в staging: {len(products)} товаров")
    finally:
        cur.close()
        conn.close()


def process_staging(full_mode: bool = False):
    """Обработка staging: UPSERT в nomenclature и current_prices (LEGACY)"""
    conn = get_db()
    cur = conn.cursor()
    try:
        # 1. UPSERT в nomenclature с brand/model/part_type
        _nom_update = (
            "name = EXCLUDED.name, "
            "category = EXCLUDED.category, "
            "brand = EXCLUDED.brand, "
            "model = EXCLUDED.model, "
            "part_type = EXCLUDED.part_type, "
            "updated_at = NOW()"
            if full_mode else
            "updated_at = NOW()"
        )
        cur.execute(f"""
            INSERT INTO nomenclature (article, name, category, brand, model, part_type,
                                      first_seen_at, updated_at)
            SELECT DISTINCT ON (article)
                article, name, category, brand, model, part_type,
                NOW(), NOW()
            FROM staging
            WHERE article IS NOT NULL AND article != ''
            ON CONFLICT (article) DO UPDATE SET
                {_nom_update}
        """)
        nom_count = cur.rowcount
        print(f"Nomenclature: {nom_count} записей обновлено/добавлено")

        # 2. UPSERT в current_prices
        cur.execute("""
            INSERT INTO current_prices (nomenclature_id, outlet_id, price, updated_at)
            SELECT DISTINCT ON (n.id, o.id)
                n.id, o.id, s.price, NOW()
            FROM staging s
            JOIN nomenclature n ON n.article = s.article
            JOIN outlets o ON o.code = s.outlet_code
            WHERE s.article IS NOT NULL AND s.article != ''
            ON CONFLICT (nomenclature_id, outlet_id) DO UPDATE SET
                price = EXCLUDED.price,
                updated_at = NOW()
        """)
        price_count = cur.rowcount
        print(f"Current prices: {price_count} записей обновлено/добавлено")

        conn.commit()

        # Статистика
        cur.execute("SELECT COUNT(*) FROM nomenclature")
        total_nom = cur.fetchone()[0]

        print(f"\n=== Итого в БД ===")
        print(f"Номенклатура: {total_nom}")

    finally:
        cur.close()
        conn.close()


def save_to_db(products: List[Dict], outlets: List[Dict], full_mode: bool = False):
    """
    Сохранение в новую схему БД v10: profi_nomenclature (с price) + profi_product_urls
    Single-URL: один URL на товар (outlet_id = NULL), price в nomenclature
    """
    if not products:
        print("Нет товаров для сохранения")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        # Создаём outlets
        for outlet in outlets:
            cur.execute("""
                INSERT INTO outlets (code, city, name, is_active)
                VALUES (%s, %s, %s, true)
                ON CONFLICT (code) DO UPDATE SET
                    city = EXCLUDED.city,
                    name = EXCLUDED.name
            """, (outlet["outlet_code"], outlet["city"], outlet["shop"]))
        conn.commit()
        print(f"Создано/обновлено {len(outlets)} outlets для Profi")

        nom_inserted = 0
        nom_updated = 0
        urls_inserted = 0

        _nom_update = (
            "name = EXCLUDED.name, "
            "category = EXCLUDED.category, "
            "brand = EXCLUDED.brand, "
            "model = EXCLUDED.model, "
            "part_type = EXCLUDED.part_type, "
            "price = EXCLUDED.price, "
            "updated_at = NOW()"
            if full_mode else
            "price = EXCLUDED.price, "
            "updated_at = NOW()"
        )
        for p in products:
            # Генерируем product_url из артикула (у Profi нет реальных URL товаров)
            article = p.get("article", "")
            if not article:
                continue

            product_url = f"https://siriust.ru/product/{article}"

            # 1. UPSERT в profi_nomenclature (price в nomenclature)
            price = p.get("price", 0)

            cur.execute(f"""
                INSERT INTO profi_nomenclature (name, article, category, brand, model, part_type, price, first_seen_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (article) DO UPDATE SET
                    {_nom_update}
                RETURNING id, (xmax = 0) as inserted
            """, (
                p.get("name", ""),
                article,
                p.get("category", ""),
                p.get("brand", ""),
                p.get("model", ""),
                p.get("part_type", ""),
                price
            ))

            row = cur.fetchone()
            nomenclature_id = row[0]
            if row[1]:
                nom_inserted += 1
            else:
                nom_updated += 1

            # 2. INSERT в profi_product_urls (single-URL: outlet_id = NULL)
            cur.execute("""
                INSERT INTO profi_product_urls (nomenclature_id, outlet_id, url, updated_at)
                VALUES (%s, NULL, %s, NOW())
                ON CONFLICT (url) DO NOTHING
            """, (nomenclature_id, product_url))
            urls_inserted += 1

        conn.commit()

        print(f"\n=== Сохранено в БД (v10) ===")
        print(f"profi_nomenclature: +{nom_inserted} новых, ~{nom_updated} обновлено")
        print(f"profi_product_urls: {urls_inserted} URL")

        # Итоговая статистика
        cur.execute("SELECT COUNT(*) FROM profi_nomenclature")
        total_nom = cur.fetchone()[0]
        print(f"\nИтого в БД: {total_nom} товаров")

    finally:
        cur.close()
        conn.close()


def main():
    arg_parser = argparse.ArgumentParser(description='Парсер прайс-листов Profi (siriust.ru)')
    arg_parser.add_argument('--all', action='store_true',
                           help='Полный парсинг всех городов + сохранение в БД')
    arg_parser.add_argument('--dynamic', action='store_true',
                           help='Динамически получить список прайс-листов с сайта')
    arg_parser.add_argument('--process', action='store_true',
                           help='Только обработка staging (LEGACY, без парсинга)')
    arg_parser.add_argument('--old-schema', action='store_true',
                           help='Использовать старую схему БД (staging -> nomenclature)')
    arg_parser.add_argument('--no-db', action='store_true',
                           help='Не сохранять в БД (только CSV/JSON)')
    arg_parser.add_argument('--city', '-c', type=str, default=None,
                           help='Парсить только указанный город')
    arg_parser.add_argument('--full', action='store_true',
                           help='Полный парсинг (UPSERT и так полный для этого парсера)')
    args = arg_parser.parse_args()

    # Только обработка staging (LEGACY)
    if args.process:
        print("Обработка staging (LEGACY)...")
        process_staging(full_mode=args.full)
        print("\nОбработка завершена!")
        return

    # Парсинг
    parser = ProfiParser()

    if args.city:
        # Парсим только один город
        city_lower = args.city.lower()
        price_lists = [pl for pl in PRICE_LISTS if pl["city"].lower() == city_lower]
        if not price_lists:
            print(f"Город '{args.city}' не найден. Доступные города:")
            for city in sorted(set(pl["city"] for pl in PRICE_LISTS)):
                print(f"  - {city}")
            return

        print(f"Парсинг города: {args.city}")
        for pl in price_lists:
            parser.parse_single_outlet(pl)
    else:
        # Парсим все города
        parser.parse_all_outlets(use_dynamic=args.dynamic)

    parser.print_stats()
    parser.save_to_csv()
    parser.save_to_json()

    # Сохранение в БД
    if not args.no_db:
        if args.old_schema:
            # LEGACY: staging -> nomenclature -> current_prices
            ensure_outlets(parser.outlets_parsed)
            save_staging(parser.products)
            if args.all:
                process_staging(full_mode=args.full)
        else:
            # НОВАЯ СХЕМА: profi_nomenclature + profi_prices
            save_to_db(parser.products, parser.outlets_parsed, full_mode=args.full)

    print("\nПарсинг завершён!")


if __name__ == "__main__":
    main()
