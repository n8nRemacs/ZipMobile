#!/usr/bin/env python3
"""
Парсер прайс-листа Profi для n8n
Выводит JSON на stdout для интеграции с n8n

Использование:
  python parse_price_n8n.py /path/to/file.xlsx

На выходе JSON массив товаров
"""
import sys
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


def canon(s: str) -> str:
    """Канонизация строки для сравнения"""
    s = s.strip().lower().replace("ё", "е")
    return re.sub(r"[^a-z0-9а-я]+", "", s)


def cell_font_size(cell):
    """Получить размер шрифта ячейки"""
    try:
        return int(round(float(cell.font.sz))) if cell and cell.font and cell.font.sz else None
    except:
        return None


def find_header_row(ws):
    """Найти строку заголовка (содержит 'наимен')"""
    for r in range(1, min(ws.max_row + 1, 50)):
        for c in range(1, min(ws.max_column, 50) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "наимен" in v.lower():
                return r
    return None


def row_empty_fast(ws, r, cols):
    """Быстрая проверка пустой строки по ключевым колонкам"""
    for idx in cols:
        if idx:
            v = ws.cell(r, idx).value
            if v not in (None, "", " "):
                return False
    return True


def clean_value(v):
    """Преобразовать значение для JSON"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return str(v).strip() if str(v).strip() else None


def parse_price(src_path: str):
    """Парсинг Excel и вывод JSON"""
    SRC = Path(src_path)

    if not SRC.exists():
        return {"error": f"File not found: {SRC}"}

    # Открываем с поддержкой стилей (data_only=False для доступа к font)
    wb = load_workbook(SRC, data_only=False)
    ws = wb.active

    # Ищем заголовок
    header_row = find_header_row(ws)
    if not header_row:
        return {"error": "Header row with 'Наименование' not found"}

    # Карта заголовков
    headers_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            headers_map[v.strip()] = c
    canon_map = {canon(k): v for k, v in headers_map.items()}

    # Синонимы колонок
    targets = {
        "name": ["наимен", "наименование", "name"],
        "article": ["артикул", "код", "code", "sku"],
        "barcode": ["штрихкод", "штрих-код", "barcode", "ean"],
        "price_rub": ["цена", "розница", "retail", "price"],
        "unit": ["ед", "единица", "unit"],
        "stock": ["количество", "наличие", "остаток", "stock", "qty"],
        "warehouse": ["склад", "warehouse"],
        "note": ["примечание", "комментарий", "comment", "note"],
    }

    # Резолвим колонки
    resolved = {}
    for target, keys in targets.items():
        found = None
        for key in keys:
            for hdr_canon, col in canon_map.items():
                if key in hdr_canon:
                    found = col
                    break
            if found:
                break
        resolved[target] = found

    name_col = resolved["name"]
    if not name_col:
        return {"error": "Column 'Наименование' not found"}

    empty_check_cols = [name_col] + [c for c in resolved.values() if c]

    # Парсим данные
    current_brand = current_model = current_type = None
    products = []

    for r in range(header_row + 1, ws.max_row + 1):
        if row_empty_fast(ws, r, empty_check_cols):
            break

        name_cell = ws.cell(r, name_col)
        name_val = str(name_cell.value).strip() if name_cell.value else None
        if not name_val:
            continue

        fsize = cell_font_size(name_cell)

        # Иерархия по размеру шрифта
        if fsize == 11:
            current_brand = name_val
            current_model = None
            current_type = None
            continue
        elif fsize == 10:
            current_model = name_val
            current_type = None
            continue
        elif fsize == 9:
            current_type = name_val
            continue
        else:
            # Это товар - собираем все поля
            product = {
                "name": name_val,
                "article": clean_value(ws.cell(r, resolved["article"]).value) if resolved["article"] else None,
                "barcode": clean_value(ws.cell(r, resolved["barcode"]).value) if resolved["barcode"] else None,
                "price_rub": clean_value(ws.cell(r, resolved["price_rub"]).value) if resolved["price_rub"] else None,
                "note": clean_value(ws.cell(r, resolved["note"]).value) if resolved["note"] else None,
                "brand_attr": current_brand,
                "model_attr": current_model,
                "part_type_attr": current_type,
            }

            # Stock info → note (stock columns removed in v9.0)
            stock_val = clean_value(ws.cell(r, resolved["stock"]).value) if resolved["stock"] else None
            if stock_val:
                product["note"] = f"{product['note'] or ''} | Stock: {stock_val}".strip(" |")

            products.append(product)

    return products


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: parse_price_n8n.py <file.xlsx>"}))
        sys.exit(1)

    result = parse_price(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False))
