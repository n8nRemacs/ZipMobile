"""
Парсер Profi - ЭТАП 1: Сбор СЫРЫХ данных

Парсит Excel файлы и сохраняет в profi_nomenclature_all БЕЗ нормализации.
Определяет категории по размеру шрифта (font-size 11/10/9).

После парсинга данные обрабатываются в n8n Upload.json (ЭТАП 2).
"""

import json
import re
import os
import tempfile
import subprocess
import argparse
import httpx
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import load_workbook

from config import DATA_DIR, PRODUCTS_CSV, PRODUCTS_JSON
from price_lists_config import PRICE_LISTS
from fetch_price_lists import fetch_price_lists

# === Конфигурация БД (Supabase) ===
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db_wrapper import get_db

# === Размеры шрифта для определения категорий ===
FONT_SIZE_BRAND = 11      # Бренд
FONT_SIZE_MODEL = 10      # Модель
FONT_SIZE_PART_TYPE = 9   # Тип запчасти


class ProfiParserRaw:
    """Парсер - только сбор СЫРЫХ данных, БЕЗ нормализации"""

    def __init__(self):
        self.products: List[Dict] = []
        self.outlets_parsed: List[Dict] = []
        self.errors: List[Dict] = []
        self.stats = {'total_parsed': 0}
        os.makedirs(DATA_DIR, exist_ok=True)

    def download_price_list(self, url: str) -> Optional[str]:
        """Скачивает прайс-лист и возвращает путь к временному файлу"""
        try:
            with httpx.Client(timeout=60, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()

                suffix = ".xls" if ".xls" in url.lower() else ".xlsx"
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
                    f.write(response.content)
                    return f.name
        except Exception as e:
            self.errors.append({
                "url": url,
                "error": str(e),
                "time": datetime.now().isoformat()
            })
            return None

    def parse_price(self, price_text) -> float:
        """Парсит цену из текста"""
        if price_text is None:
            return 0.0
        price_match = re.search(r'([\d\s]+[.,]?\d*)', str(price_text))
        if price_match:
            price_str = price_match.group(1).replace(' ', '').replace(',', '.')
            try:
                return float(price_str)
            except:
                return 0.0
        return 0.0

    def _get_font_size(self, cell) -> Optional[int]:
        """Получает размер шрифта ячейки"""
        try:
            if cell and cell.font and cell.font.sz:
                return int(round(float(cell.font.sz)))
        except:
            pass
        return None

    def _canon(self, s: str) -> str:
        """Канонизация строки для поиска заголовков"""
        s = s.strip().lower().replace("ё", "е")
        return re.sub(r"[^a-z0-9а-я]+", "", s)

    def _find_header_row(self, ws) -> Optional[int]:
        """Ищет строку с заголовками (содержит 'наимен')"""
        for r in range(1, min(ws.max_row + 1, 20)):
            for c in range(1, min(ws.max_column, 50) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "наимен" in v.lower():
                    return r
        return None

    def _resolve_columns(self, ws, header_row: int) -> Dict[str, Optional[int]]:
        """Определяет индексы колонок по заголовкам"""
        headers_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip():
                headers_map[self._canon(v)] = c

        targets = {
            "name": ["наимен", "наименование", "name"],
            "article": ["артикул", "код", "code", "sku"],
            "barcode": ["штрихкод", "barcode", "ean"],
            "price": ["цена", "розница", "retail", "price"],
            "quantity": ["количество", "наличие", "остаток", "stock", "qty"],
        }

        resolved = {}
        for target, keys in targets.items():
            found = None
            for key in keys:
                for hdr_canon, col in headers_map.items():
                    if key in hdr_canon:
                        found = col
                        break
                if found:
                    break
            resolved[target] = found

        return resolved

    def _row_is_empty(self, ws, row: int, cols: List[Optional[int]]) -> bool:
        """Проверяет пустая ли строка по ключевым колонкам"""
        for col in cols:
            if col:
                v = ws.cell(row, col).value
                if v not in (None, "", " "):
                    return False
        return True

    def _convert_xls_to_xlsx(self, xls_path: str) -> Optional[str]:
        """Конвертирует .xls в .xlsx через LibreOffice"""
        try:
            out_dir = tempfile.mkdtemp()

            result = subprocess.run([
                'soffice', '--headless', '--convert-to', 'xlsx',
                '--outdir', out_dir, xls_path
            ], capture_output=True, text=True, timeout=60)

            if result.returncode != 0:
                return None

            base_name = os.path.splitext(os.path.basename(xls_path))[0]
            xlsx_path = os.path.join(out_dir, f"{base_name}.xlsx")

            if os.path.exists(xlsx_path):
                return xlsx_path
            return None
        except Exception as e:
            return None

    def parse_excel_file(self, file_path: str, city: str, shop: str, outlet_code: str) -> List[Dict]:
        """Парсит Excel файл - только сбор СЫРЫХ данных"""
        products = []
        converted_path = None

        # Конвертация .xls -> .xlsx
        if file_path.lower().endswith('.xls') and not file_path.lower().endswith('.xlsx'):
            converted_path = self._convert_xls_to_xlsx(file_path)
            if converted_path:
                file_path = converted_path
            else:
                self.errors.append({
                    "file": file_path,
                    "error": "Failed to convert .xls to .xlsx",
                    "time": datetime.now().isoformat()
                })
                return []

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            self.errors.append({
                "file": file_path,
                "error": f"Failed to read Excel: {e}",
                "time": datetime.now().isoformat()
            })
            if converted_path:
                try:
                    os.unlink(converted_path)
                    os.rmdir(os.path.dirname(converted_path))
                except:
                    pass
            return []

        # Находим заголовки
        header_row = self._find_header_row(ws)
        if not header_row:
            self.errors.append({
                "file": file_path,
                "error": "Header row not found",
                "time": datetime.now().isoformat()
            })
            return []

        # Определяем колонки
        cols = self._resolve_columns(ws, header_row)
        name_col = cols.get("name")
        if not name_col:
            self.errors.append({
                "file": file_path,
                "error": "Name column not found",
                "time": datetime.now().isoformat()
            })
            return []

        article_col = cols.get("article")
        barcode_col = cols.get("barcode")
        price_col = cols.get("price")
        quantity_col = cols.get("quantity")
        check_cols = [name_col, article_col, price_col, quantity_col]

        # Текущие категории (определяются по font-size)
        current_brand_raw = None
        current_model_raw = None
        current_part_type_raw = None

        for r in range(header_row + 1, ws.max_row + 1):
            if self._row_is_empty(ws, r, check_cols):
                continue

            name_cell = ws.cell(r, name_col)
            name_val = name_cell.value
            if not name_val:
                continue
            name_val = str(name_val).strip()
            if not name_val:
                continue

            # Определяем тип строки по размеру шрифта
            font_size = self._get_font_size(name_cell)

            if font_size == FONT_SIZE_BRAND:
                current_brand_raw = name_val
                current_model_raw = None
                current_part_type_raw = None
                continue
            elif font_size == FONT_SIZE_MODEL:
                current_model_raw = name_val
                current_part_type_raw = None
                continue
            elif font_size == FONT_SIZE_PART_TYPE:
                current_part_type_raw = name_val
                continue

            # Это товар - сохраняем КАК ЕСТЬ
            self.stats['total_parsed'] += 1

            # Извлекаем данные
            article = ""
            if article_col:
                article_val = ws.cell(r, article_col).value
                if article_val:
                    article = str(article_val).strip()

            barcode = ""
            if barcode_col:
                barcode_val = ws.cell(r, barcode_col).value
                if barcode_val:
                    barcode = str(barcode_val).strip()

            price = 0.0
            if price_col:
                price = self.parse_price(ws.cell(r, price_col).value)

            in_stock = False
            if quantity_col:
                qty_val = ws.cell(r, quantity_col).value
                if qty_val:
                    qty_str = str(qty_val).strip().lower()
                    in_stock = qty_str == '*' or 'наличи' in qty_str or qty_str not in ('', '0', 'нет')

            # Формируем категорию (для отладки)
            category_parts = [p for p in [current_brand_raw, current_model_raw, current_part_type_raw] if p]
            category = " / ".join(category_parts) if category_parts else ""

            # Сохраняем СЫРЫЕ данные
            products.append({
                'article': article,
                'barcode': barcode,
                'name': name_val,
                'price': price,
                'in_stock': in_stock,

                # Поля из Excel
                'brand': current_brand_raw or "",
                'model': current_model_raw or "",
                'part_type': current_part_type_raw or "",
                'category': category,

                # Метаданные
                'city': city,
                'shop': shop,
                'outlet_code': outlet_code,
            })

        try:
            wb.close()
        except:
            pass

        # Удаляем временный файл
        if converted_path:
            try:
                os.unlink(converted_path)
                os.rmdir(os.path.dirname(converted_path))
            except:
                pass

        return products

    def parse_single_outlet(self, price_list: Dict) -> int:
        """Парсит один прайс-лист"""
        url = price_list.get("url", "")
        city = price_list.get("city", "")
        shop = price_list.get("shop", "")
        outlet_code = price_list.get("outlet_code", "")

        if not outlet_code:
            filename = url.split("/")[-1].replace(".xls", "").replace(".xlsx", "")
            outlet_code = f"profi-{filename.lower().replace(' ', '-')}"

        print(f"  [{city}] {shop}...")

        # Скачиваем
        file_path = self.download_price_list(url)
        if not file_path:
            print(f"    [SKIP] Не удалось скачать")
            return 0

        try:
            # Парсим
            products = self.parse_excel_file(file_path, city, shop, outlet_code)
            self.products.extend(products)

            self.outlets_parsed.append({
                "city": city,
                "shop": shop,
                "outlet_code": outlet_code,
                "products_count": len(products)
            })

            print(f"    +{len(products)} товаров")
            return len(products)
        finally:
            try:
                os.unlink(file_path)
            except:
                pass

    def parse_all_outlets(self, use_dynamic: bool = False):
        """Парсит все прайс-листы"""
        print(f"\n{'='*60}")
        print(f"Парсер Profi - ЭТАП 1: Сбор СЫРЫХ данных")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}\n")

        if use_dynamic:
            print("[Режим] Динамическая загрузка списка с сайта\n")
            price_lists = fetch_price_lists()
            for pl in price_lists:
                filename = pl["url"].split("/")[-1].replace(".xls", "").replace(".xlsx", "")
                pl["outlet_code"] = f"profi-{filename.lower().replace(' ', '-').replace('%20', '-')}"
        else:
            print(f"[Режим] Статический список ({len(PRICE_LISTS)} прайс-листов)\n")
            price_lists = PRICE_LISTS

        total_products = 0
        for pl in price_lists:
            count = self.parse_single_outlet(pl)
            total_products += count

        print(f"\n{'='*60}")
        print(f"ИТОГО: {total_products} товаров из {len(self.outlets_parsed)} торговых точек")
        print(f"Ошибок: {len(self.errors)}")
        print(f"{'='*60}")

    def save_to_json(self, filename: str = None):
        """Сохраняет товары в JSON"""
        filename = filename or PRODUCTS_JSON

        data = {
            "source": "siriust.ru (Profi RAW)",
            "date": datetime.now().isoformat(),
            "total": len(self.products),
            "stats": self.stats,
            "outlets": self.outlets_parsed,
            "products": self.products,
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"Сохранено в {filename}")


def save_to_db_all(products: List[Dict], outlets: List[Dict]):
    """
    Сохранение СЫРЫХ данных в profi_nomenclature_all
    БЕЗ нормализации - как есть из Excel
    """
    if not products:
        print("Нет товаров для сохранения")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        # Создаём outlets
        for outlet in outlets:
            cur.execute("""
                INSERT INTO zip_outlets (code, city, name, is_active)
                VALUES (%s, %s, %s, true)
                ON CONFLICT (code) DO UPDATE SET
                    city = EXCLUDED.city,
                    name = EXCLUDED.name
            """, (outlet["outlet_code"], outlet["city"], outlet["shop"]))
        conn.commit()
        print(f"\nСоздано/обновлено {len(outlets)} outlets")

        inserted = 0
        updated = 0

        for p in products:
            article = p.get("article", "")
            if not article:
                continue

            # UPSERT в profi_nomenclature_all (СЫРЫЕ данные)
            cur.execute("""
                INSERT INTO profi_nomenclature_all (
                    article, barcode, name,
                    brand, model, part_type, category,
                    city, shop, outlet_code,
                    price, in_stock,
                    updated_at, processed
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), false)
                ON CONFLICT (article) DO UPDATE SET
                    name = EXCLUDED.name,
                    barcode = EXCLUDED.barcode,
                    brand = EXCLUDED.brand,
                    model = EXCLUDED.model,
                    part_type = EXCLUDED.part_type,
                    category = EXCLUDED.category,
                    city = EXCLUDED.city,
                    shop = EXCLUDED.shop,
                    outlet_code = EXCLUDED.outlet_code,
                    price = EXCLUDED.price,
                    in_stock = EXCLUDED.in_stock,
                    updated_at = NOW(),
                    processed = false
                RETURNING (xmax = 0) as is_insert
            """, (
                article,
                p.get("barcode", ""),
                p.get("name", ""),
                p.get("brand", ""),
                p.get("model", ""),
                p.get("part_type", ""),
                p.get("category", ""),
                p.get("city", ""),
                p.get("shop", ""),
                p.get("outlet_code", ""),
                p.get("price", 0),
                p.get("in_stock", False)
            ))

            row = cur.fetchone()
            if row and row[0]:
                inserted += 1
            else:
                updated += 1

        conn.commit()

        print(f"\n{'='*60}")
        print(f"СОХРАНЕНО В profi_nomenclature_all:")
        print(f"  +{inserted} новых товаров")
        print(f"  ~{updated} обновлено")

        # Итоговая статистика
        cur.execute("SELECT COUNT(*) FROM profi_nomenclature_all")
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM profi_nomenclature_all WHERE processed = false")
        unprocessed = cur.fetchone()[0]
        print(f"\nИтого в БД: {total} товаров ({unprocessed} не обработаны)")
        print(f"{'='*60}")

    finally:
        cur.close()
        conn.close()


def main():
    parser = argparse.ArgumentParser(description='Парсер Profi - ЭТАП 1 (СЫРЫЕ данные)')
    parser.add_argument('--all', action='store_true',
                       help='Полный парсинг всех городов + сохранение в БД')
    parser.add_argument('--dynamic', action='store_true',
                       help='Динамически получить список прайс-листов с сайта')
    parser.add_argument('--no-db', action='store_true',
                       help='Не сохранять в БД (только JSON)')
    parser.add_argument('--city', '-c', type=str, default=None,
                       help='Парсить только указанный город')
    parser.add_argument('--test', action='store_true',
                       help='Тестовый режим: 1 город')
    args = parser.parse_args()

    # Парсинг
    profi_parser = ProfiParserRaw()

    if args.test:
        # Тестовый режим: парсим только Москву
        price_lists = [pl for pl in PRICE_LISTS if 'москва' in pl["city"].lower()][:1]
        if price_lists:
            print(f"Тестовый режим: парсим {price_lists[0]['city']}")
            profi_parser.parse_single_outlet(price_lists[0])
    elif args.city:
        # Парсим только один город
        city_lower = args.city.lower()
        price_lists = [pl for pl in PRICE_LISTS if pl["city"].lower() == city_lower]
        if not price_lists:
            print(f"Город '{args.city}' не найден. Доступные города:")
            for city in sorted(set(pl["city"] for pl in PRICE_LISTS)):
                print(f"  - {city}")
            return

        print(f"Парсинг города: {args.city}")
        for pl in price_lists:
            profi_parser.parse_single_outlet(pl)
    else:
        # Парсим все города
        profi_parser.parse_all_outlets(use_dynamic=args.dynamic)

    profi_parser.save_to_json()

    # Сохранение в БД
    if not args.no_db and profi_parser.products:
        save_to_db_all(profi_parser.products, profi_parser.outlets_parsed)

    print("\nПарсинг завершён!")
    print(f"\nСледующий шаг: запустить n8n Upload.json для нормализации")


if __name__ == "__main__":
    main()
