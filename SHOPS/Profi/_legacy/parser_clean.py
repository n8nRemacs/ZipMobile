"""
Единый парсер прайс-листов Profi с встроенной нормализацией

Объединяет:
1. Парсинг Excel (определение категорий по font-size)
2. Нормализацию (логика из Normalize_v2.json)

На выходе: только очищенная номенклатура готовая для AI
"""

import json
import re
import os
import tempfile
import subprocess
import argparse
import httpx
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import load_workbook

from config import DATA_DIR, PRODUCTS_CSV, PRODUCTS_JSON
from price_lists_config import PRICE_LISTS
from fetch_price_lists import fetch_price_lists

# === Конфигурация БД (Supabase) ===
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db_wrapper import get_db

# === Размеры шрифта для определения категорий ===
FONT_SIZE_BRAND = 11      # Бренд
FONT_SIZE_MODEL = 10      # Модель
FONT_SIZE_PART_TYPE = 9   # Тип запчасти

# === Словарь нормализации part_type ===
PART_TYPE_MAPPING = {
    'ДИСПЛЕИ': 'ДИСПЛЕЙ',
    'ДИСПЛЕИ INCELL': 'ДИСПЛЕЙ',
    'ДИСПЛЕИ OLED': 'ДИСПЛЕЙ',
    'ДИСПЛЕИ ORIGINAL': 'ДИСПЛЕЙ',
    'ДИСПЛЕИ PREMIUM': 'ДИСПЛЕЙ',
    'ДИСПЛЕИ TIANMA': 'ДИСПЛЕЙ',
    'КАМЕРЫ': 'КАМЕРА',
    'ДЕРЖАТЕЛИ SIM': 'SIM-ЛОТОК',
    'ДЕРЖАТЕЛИ, КОННЕКТОРЫ SIM/КАРТЫ ПАМЯТИ': 'SIM-ЛОТОК',
    'КОННЕКТОРЫ SIM/КАРТЫ ПАМЯТИ': 'SIM-ЛОТОК',
    'ДЖОЙСТИКИ': 'ДЖОЙСТИК',
    'ДИНАМИК': 'ДИНАМИК',
    'ЗВОНКИ, ВИБРОЗВОНКИ, ДИНАМИКИ': 'ДИНАМИК',
    'ЗВОНКИ, ДИНАМИКИ': 'ДИНАМИК',
    'ЗВОНКИ, ДИНАМИКИ, ВИБРОМОТОРЫ': 'ДИНАМИК',
    'ЗАГЛУШКИ': 'ЗАГЛУШКА',
    'ЗАДНИЕ КРЫШКИ': 'КРЫШКА ЗАДНЯЯ',
    'ЗАДНЯЯ КРЫШКИ': 'КРЫШКА ЗАДНЯЯ',
    'КНОПКИ': 'КНОПКА',
    'КНОПКИ ВКЛЮЧЕНИЯ': 'КНОПКА',
    'КНОПКИ ВКЛЮЧЕНИЯ, ТОЛКАТЕЛИ': 'КНОПКА',
    'КОРПУСА': 'КОРПУС',
    'МИКРОСХЕМЫ': 'МИКРОСХЕМА',
    'МИКРОФОНЫ': 'МИКРОФОН',
    'ПЛАТЫ КЛАВИАТУРЫ': 'ПЛАТА',
    'ПРОКЛЕЙКИ ДЛЯ ДИСПЛЕЙНЫХ МОДУЛЕЙ И ЗАДНИХ КРЫШЕК': 'СКОТЧ',
    'СКОТЧ ДЛЯ ФИКСАЦИИ АКБ': 'СКОТЧ',
    'РАЗЪЕМЫ': 'РАЗЪЕМ',
    'РАЗЪЕМЫ ДЛЯ ЗАРЯДКИ': 'РАЗЪЕМ ЗАРЯДКИ',
    'РАЗЪЕМЫ ДЛЯ ЗАРЯДКИ, АУДИО РАЗЪЕМЫ': 'РАЗЪЕМ ЗАРЯДКИ',
    'РАЗЪЕМЫ ЗАРЯДКИ': 'РАЗЪЕМ ЗАРЯДКИ',
    'РАМКИ': 'РАМКА',
    'СТЕКЛА КАМЕРЫ': 'СТЕКЛО КАМЕРЫ',
    'СТЕКЛО КАМЕРЫ': 'СТЕКЛО КАМЕРЫ',
    'ТАЧСКРИНЫ': 'ТАЧСКРИН',
    'ТАЧСКРИНЫ ДЛЯ IPAD': 'ТАЧСКРИН',
    'ШЛЕЙФА': 'ШЛЕЙФ',
    'ШЛЕЙФА ДЛЯ IPAD': 'ШЛЕЙФ',
    'ШЛЕЙФЫ': 'ШЛЕЙФ',
}


class ProfiNormalizer:
    """Логика нормализации из Normalize_v2.json"""

    @staticmethod
    def is_spare_part(brand_raw: str, model_raw: str) -> bool:
        """2.1 Фильтрация: только запчасти"""
        if not brand_raw:
            return False

        # Оставляем только запчасти
        if brand_raw == '3. АКСЕССУАРЫ' and model_raw == 'АКБ ДЛЯ МОБИЛЬНОЙ ТЕХНИКИ':
            return True

        if brand_raw in ['2. ЗАПЧАСТИ ДЛЯ СОТОВЫХ', '1. ЗАПЧАСТИ ДЛЯ APPLE']:
            return True

        return False

    @staticmethod
    def normalize_brand_model(brand_raw: str, model_raw: str) -> tuple[str, str]:
        """2.2 Нормализация brand и model"""
        # Сначала определяем верхний уровень
        if brand_raw == '1. ЗАПЧАСТИ ДЛЯ APPLE':
            brand = 'Apple'
        elif brand_raw == '2. ЗАПЧАСТИ ДЛЯ СОТОВЫХ':
            brand = 'Android'
        elif brand_raw == '3. АКСЕССУАРЫ':
            brand = 'АКБ'
        else:
            brand = brand_raw

        # Нормализуем model
        if model_raw == 'ЗАПЧАСТИ ДЛЯ APPLE IPAD':
            model = 'iPad'
        elif model_raw == 'ЗАПЧАСТИ ДЛЯ APPLE IPHONE':
            model = 'iPhone'
        elif model_raw == 'ЗАПЧАСТИ ДЛЯ APPLE WATCH':
            model = 'Apple Watch'
        elif model_raw == 'ЗАПЧАСТИ ДЛЯ APPLE MACBOOK':
            model = 'MacBook'
        elif model_raw == 'ЗАПЧАСТИ ДЛЯ APPLE AIRPODS':
            model = 'AirPods'
        else:
            # Убираем "ЗАПЧАСТИ ДЛЯ " из начала
            model = re.sub(r'^ЗАПЧАСТИ ДЛЯ\s+', '', model_raw, flags=re.IGNORECASE)

        return brand, model

    @staticmethod
    def normalize_part_type_iphone(part_type: str, model: str) -> str:
        """2.3 Обработка part_type для iPhone - обрезаем до слова 'ДЛЯ'"""
        if not model or 'iphone' not in model.lower():
            return part_type

        if not part_type or 'для' not in part_type.upper():
            return part_type

        # Находим первое вхождение "ДЛЯ"
        pos = part_type.upper().find('ДЛЯ')
        if pos > 0:
            return part_type[:pos].strip()

        return part_type

    @staticmethod
    def transfer_model_to_brand(brand: str, model: str) -> tuple[str, Optional[str]]:
        """2.4 Перенос model в brand (исправление структуры)"""
        # Для Apple: brand='Apple', model='iPhone' -> brand='iPhone', model=NULL
        if brand == 'Apple' and model and model.strip():
            return model.strip(), None

        # Для Android: brand='Android', model='SAMSUNG' -> brand='SAMSUNG', model=NULL
        if brand == 'Android' and model and model.strip():
            return model.strip(), None

        return brand, model

    @staticmethod
    def process_akb_android(brand_raw: str, model_raw: str, part_type_raw: str) -> tuple[str, str]:
        """2.6 АКБ для мобильной техники"""
        if model_raw != 'АКБ ДЛЯ МОБИЛЬНОЙ ТЕХНИКИ':
            return None, None

        # Убираем ведущие "АКБ " и "ДЛЯ "
        brand = part_type_raw
        brand = re.sub(r'^АКБ\s+', '', brand, flags=re.IGNORECASE)
        brand = re.sub(r'^ДЛЯ\s+', '', brand, flags=re.IGNORECASE)
        brand = brand.strip()

        return brand, 'АКБ'

    @staticmethod
    def normalize_universal(brand: str) -> str:
        """2.7 Нормализация УНИВЕРСАЛЬНЫЙ"""
        if not brand:
            return brand

        if brand == 'УНИВЕРСАЛЬНЫЕ ЗАПЧАСТИ':
            return 'УНИВЕРСАЛЬНЫЙ'

        if 'универсал' in brand.lower() and brand != 'УНИВЕРСАЛЬНЫЙ':
            return 'УНИВЕРСАЛЬНЫЙ'

        return brand

    @staticmethod
    def normalize_part_type(part_type: str) -> str:
        """2.9 Приведение part_type к единому виду"""
        if not part_type:
            return part_type

        key = part_type.strip().upper()
        return PART_TYPE_MAPPING.get(key, part_type.strip())

    @classmethod
    def normalize_product(cls, brand_raw: str, model_raw: str, part_type_raw: str) -> Dict:
        """Полная нормализация продукта - применяет всю логику из Normalize_v2.json"""

        # 2.1 Фильтрация
        if not cls.is_spare_part(brand_raw, model_raw):
            return None  # Пропускаем не-запчасти

        # 2.2 Нормализация brand/model
        brand, model = cls.normalize_brand_model(brand_raw, model_raw)
        part_type = part_type_raw

        # 2.3 Обработка part_type для iPhone
        part_type = cls.normalize_part_type_iphone(part_type, model)

        # 2.4 Перенос model в brand
        brand, model = cls.transfer_model_to_brand(brand, model)

        # 2.6 АКБ для Android
        akb_brand, akb_part_type = cls.process_akb_android(brand_raw, model_raw, part_type_raw)
        if akb_brand:
            brand = akb_brand
            part_type = akb_part_type

        # 2.7 Универсальные запчасти
        brand = cls.normalize_universal(brand)

        # 2.9 Нормализация part_type
        part_type = cls.normalize_part_type(part_type)

        return {
            'brand': brand,
            'model': model,
            'part_type': part_type
        }


class ProfiParserClean:
    """Парсер с встроенной нормализацией"""

    def __init__(self):
        self.products: List[Dict] = []
        self.outlets_parsed: List[Dict] = []
        self.errors: List[Dict] = []
        self.stats = {
            'total_raw': 0,
            'filtered_out': 0,
            'normalized': 0
        }
        os.makedirs(DATA_DIR, exist_ok=True)

    def download_price_list(self, url: str) -> Optional[str]:
        """Скачивает прайс-лист и возвращает путь к временному файлу"""
        try:
            with httpx.Client(timeout=60, follow_redirects=True) as client:
                response = client.get(url)
                response.raise_for_status()

                suffix = ".xls" if ".xls" in url.lower() else ".xlsx"
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
                    f.write(response.content)
                    return f.name
        except Exception as e:
            self.errors.append({
                "url": url,
                "error": str(e),
                "time": datetime.now().isoformat()
            })
            return None

    def parse_price(self, price_text) -> float:
        """Парсит цену из текста"""
        if price_text is None:
            return 0.0
        price_match = re.search(r'([\d\s]+[.,]?\d*)', str(price_text))
        if price_match:
            price_str = price_match.group(1).replace(' ', '').replace(',', '.')
            try:
                return float(price_str)
            except:
                return 0.0
        return 0.0

    def _get_font_size(self, cell) -> Optional[int]:
        """Получает размер шрифта ячейки"""
        try:
            if cell and cell.font and cell.font.sz:
                return int(round(float(cell.font.sz)))
        except:
            pass
        return None

    def _canon(self, s: str) -> str:
        """Канонизация строки для поиска заголовков"""
        s = s.strip().lower().replace("ё", "е")
        return re.sub(r"[^a-z0-9а-я]+", "", s)

    def _find_header_row(self, ws) -> Optional[int]:
        """Ищет строку с заголовками (содержит 'наимен')"""
        for r in range(1, min(ws.max_row + 1, 20)):
            for c in range(1, min(ws.max_column, 50) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "наимен" in v.lower():
                    return r
        return None

    def _resolve_columns(self, ws, header_row: int) -> Dict[str, Optional[int]]:
        """Определяет индексы колонок по заголовкам"""
        headers_map = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip():
                headers_map[self._canon(v)] = c

        targets = {
            "name": ["наимен", "наименование", "name"],
            "article": ["артикул", "код", "code", "sku"],
            "barcode": ["штрихкод", "штрихкод", "barcode", "ean"],
            "price": ["цена", "розница", "retail", "price"],
            "unit": ["ед", "единица", "unit"],
            "quantity": ["количество", "наличие", "остаток", "stock", "qty"],
            "warehouse": ["склад", "warehouse"],
            "note": ["примечание", "комментарий", "comment", "note"],
        }

        resolved = {}
        for target, keys in targets.items():
            found = None
            for key in keys:
                for hdr_canon, col in headers_map.items():
                    if key in hdr_canon:
                        found = col
                        break
                if found:
                    break
            resolved[target] = found

        return resolved

    def _row_is_empty(self, ws, row: int, cols: List[Optional[int]]) -> bool:
        """Проверяет пустая ли строка по ключевым колонкам"""
        for col in cols:
            if col:
                v = ws.cell(row, col).value
                if v not in (None, "", " "):
                    return False
        return True

    def _convert_xls_to_xlsx(self, xls_path: str) -> Optional[str]:
        """Конвертирует .xls в .xlsx через LibreOffice"""
        try:
            out_dir = tempfile.mkdtemp()

            result = subprocess.run([
                'soffice', '--headless', '--convert-to', 'xlsx',
                '--outdir', out_dir, xls_path
            ], capture_output=True, text=True, timeout=60)

            if result.returncode != 0:
                return None

            base_name = os.path.splitext(os.path.basename(xls_path))[0]
            xlsx_path = os.path.join(out_dir, f"{base_name}.xlsx")

            if os.path.exists(xlsx_path):
                return xlsx_path
            return None
        except Exception as e:
            return None

    def parse_excel_file(self, file_path: str, city: str, shop: str, outlet_code: str) -> List[Dict]:
        """Парсит Excel файл с встроенной нормализацией"""
        products = []
        converted_path = None

        # Конвертация .xls -> .xlsx
        if file_path.lower().endswith('.xls') and not file_path.lower().endswith('.xlsx'):
            converted_path = self._convert_xls_to_xlsx(file_path)
            if converted_path:
                file_path = converted_path
            else:
                self.errors.append({
                    "file": file_path,
                    "error": "Failed to convert .xls to .xlsx",
                    "time": datetime.now().isoformat()
                })
                return []

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            self.errors.append({
                "file": file_path,
                "error": f"Failed to read Excel: {e}",
                "time": datetime.now().isoformat()
            })
            if converted_path:
                try:
                    os.unlink(converted_path)
                    os.rmdir(os.path.dirname(converted_path))
                except:
                    pass
            return []

        # Находим заголовки
        header_row = self._find_header_row(ws)
        if not header_row:
            self.errors.append({
                "file": file_path,
                "error": "Header row not found",
                "time": datetime.now().isoformat()
            })
            return []

        # Определяем колонки
        cols = self._resolve_columns(ws, header_row)
        name_col = cols.get("name")
        if not name_col:
            self.errors.append({
                "file": file_path,
                "error": "Name column not found",
                "time": datetime.now().isoformat()
            })
            return []

        article_col = cols.get("article")
        price_col = cols.get("price")
        quantity_col = cols.get("quantity")
        check_cols = [name_col, article_col, price_col, quantity_col]

        # Текущие категории (определяются по font-size)
        current_brand_raw = None
        current_model_raw = None
        current_part_type_raw = None

        for r in range(header_row + 1, ws.max_row + 1):
            if self._row_is_empty(ws, r, check_cols):
                continue

            name_cell = ws.cell(r, name_col)
            name_val = name_cell.value
            if not name_val:
                continue
            name_val = str(name_val).strip()
            if not name_val:
                continue

            # Определяем тип строки по размеру шрифта
            font_size = self._get_font_size(name_cell)

            if font_size == FONT_SIZE_BRAND:
                current_brand_raw = name_val
                current_model_raw = None
                current_part_type_raw = None
                continue
            elif font_size == FONT_SIZE_MODEL:
                current_model_raw = name_val
                current_part_type_raw = None
                continue
            elif font_size == FONT_SIZE_PART_TYPE:
                current_part_type_raw = name_val
                continue

            # Это товар - применяем нормализацию
            self.stats['total_raw'] += 1

            # Применяем логику нормализации
            normalized = ProfiNormalizer.normalize_product(
                current_brand_raw or "",
                current_model_raw or "",
                current_part_type_raw or ""
            )

            if not normalized:
                # Товар отфильтрован (не запчасть)
                self.stats['filtered_out'] += 1
                continue

            self.stats['normalized'] += 1

            # Извлекаем остальные данные
            article = ""
            if article_col:
                article_val = ws.cell(r, article_col).value
                if article_val:
                    article = str(article_val).strip()

            price = 0.0
            if price_col:
                price = self.parse_price(ws.cell(r, price_col).value)

            in_stock = False
            if quantity_col:
                qty_val = ws.cell(r, quantity_col).value
                if qty_val:
                    qty_str = str(qty_val).strip().lower()
                    in_stock = qty_str == '*' or 'наличи' in qty_str or qty_str not in ('', '0', 'нет')

            # Формируем финальный продукт
            products.append({
                'article': article,
                'name': name_val,
                'price': price,
                'in_stock': in_stock,

                # НОРМАЛИЗОВАННЫЕ поля
                'brand': normalized['brand'],
                'model': normalized['model'],
                'part_type': normalized['part_type'],

                # Метаданные
                'city': city,
                'shop': shop,
                'outlet_code': outlet_code,
                'category': f"{current_brand_raw or ''} / {current_model_raw or ''} / {current_part_type_raw or ''}".strip(' /'),
            })

        try:
            wb.close()
        except:
            pass

        # Удаляем временный файл
        if converted_path:
            try:
                os.unlink(converted_path)
                os.rmdir(os.path.dirname(converted_path))
            except:
                pass

        return products

    def parse_single_outlet(self, price_list: Dict) -> int:
        """Парсит один прайс-лист"""
        url = price_list.get("url", "")
        city = price_list.get("city", "")
        shop = price_list.get("shop", "")
        outlet_code = price_list.get("outlet_code", "")

        if not outlet_code:
            filename = url.split("/")[-1].replace(".xls", "").replace(".xlsx", "")
            outlet_code = f"profi-{filename.lower().replace(' ', '-')}"

        print(f"  [{city}] {shop}...")

        # Скачиваем
        file_path = self.download_price_list(url)
        if not file_path:
            print(f"    [SKIP] Не удалось скачать")
            return 0

        try:
            # Парсим с нормализацией
            products = self.parse_excel_file(file_path, city, shop, outlet_code)
            self.products.extend(products)

            self.outlets_parsed.append({
                "city": city,
                "shop": shop,
                "outlet_code": outlet_code,
                "products_count": len(products)
            })

            print(f"    +{len(products)} товаров")
            return len(products)
        finally:
            try:
                os.unlink(file_path)
            except:
                pass

    def parse_all_outlets(self, use_dynamic: bool = False):
        """Парсит все прайс-листы"""
        print(f"\n{'='*60}")
        print(f"Парсер Profi с встроенной нормализацией")
        print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}\n")

        if use_dynamic:
            print("[Режим] Динамическая загрузка списка с сайта\n")
            price_lists = fetch_price_lists()
            for pl in price_lists:
                filename = pl["url"].split("/")[-1].replace(".xls", "").replace(".xlsx", "")
                pl["outlet_code"] = f"profi-{filename.lower().replace(' ', '-').replace('%20', '-')}"
        else:
            print(f"[Режим] Статический список ({len(PRICE_LISTS)} прайс-листов)\n")
            price_lists = PRICE_LISTS

        total_products = 0
        for pl in price_lists:
            count = self.parse_single_outlet(pl)
            total_products += count

        print(f"\n{'='*60}")
        print(f"СТАТИСТИКА НОРМАЛИЗАЦИИ:")
        print(f"  Всего строк обработано: {self.stats['total_raw']}")
        print(f"  Отфильтровано (не запчасти): {self.stats['filtered_out']}")
        print(f"  Нормализовано и сохранено: {self.stats['normalized']}")
        print(f"\nИТОГО: {total_products} товаров из {len(self.outlets_parsed)} торговых точек")
        print(f"Ошибок: {len(self.errors)}")
        print(f"{'='*60}")

    def print_stats(self):
        """Выводит детальную статистику"""
        print(f"\n{'='*60}")
        print("ДЕТАЛЬНАЯ СТАТИСТИКА")
        print(f"{'='*60}")

        # По брендам
        brand_counts = {}
        for p in self.products:
            brand = p.get("brand", "Неизвестно")
            brand_counts[brand] = brand_counts.get(brand, 0) + 1

        print(f"\nБрендов: {len(brand_counts)}")
        print("\nТоп-15 брендов:")
        for brand, count in sorted(brand_counts.items(), key=lambda x: -x[1])[:15]:
            print(f"  {count:5d} | {brand}")

        # По типам запчастей
        part_type_counts = {}
        for p in self.products:
            pt = p.get("part_type", "Неизвестно")
            part_type_counts[pt] = part_type_counts.get(pt, 0) + 1

        print(f"\nТипов запчастей: {len(part_type_counts)}")
        print("\nТоп-10 типов запчастей:")
        for pt, count in sorted(part_type_counts.items(), key=lambda x: -x[1])[:10]:
            print(f"  {count:5d} | {pt}")

        # По городам
        city_counts = {}
        for p in self.products:
            city = p.get("city", "Неизвестно")
            city_counts[city] = city_counts.get(city, 0) + 1

        print(f"\nГородов: {len(city_counts)}")

        # По наличию
        in_stock_count = sum(1 for p in self.products if p.get("in_stock", False))
        print(f"\nВ наличии: {in_stock_count} из {len(self.products)}")

        # Ценовой диапазон
        prices = [p.get("price", 0) for p in self.products if p.get("price", 0) > 0]
        if prices:
            print(f"\nЦены: от {min(prices):.0f} до {max(prices):.0f} руб")
            print(f"Средняя: {sum(prices)/len(prices):.0f} руб")

    def save_to_json(self, filename: str = None):
        """Сохраняет товары в JSON"""
        filename = filename or PRODUCTS_JSON

        data = {
            "source": "siriust.ru (Profi Clean)",
            "date": datetime.now().isoformat(),
            "total": len(self.products),
            "stats": self.stats,
            "outlets": self.outlets_parsed,
            "products": self.products,
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"Сохранено в {filename}")


def save_to_db(products: List[Dict], outlets: List[Dict]):
    """Сохранение в БД: profi_nomenclature + profi_current_prices"""
    if not products:
        print("Нет товаров для сохранения")
        return

    conn = get_db()
    cur = conn.cursor()
    try:
        # Создаём outlets
        for outlet in outlets:
            cur.execute("""
                INSERT INTO zip_outlets (code, city, name, is_active)
                VALUES (%s, %s, %s, true)
                ON CONFLICT (code) DO UPDATE SET
                    city = EXCLUDED.city,
                    name = EXCLUDED.name
            """, (outlet["outlet_code"], outlet["city"], outlet["shop"]))
        conn.commit()
        print(f"\nСоздано/обновлено {len(outlets)} outlets")

        # Кэш outlet_id
        cur.execute("SELECT code, id FROM zip_outlets WHERE code LIKE 'profi-%'")
        outlet_cache = {row[0]: row[1] for row in cur.fetchall()}

        nom_inserted = 0
        nom_updated = 0
        prices_upserted = 0

        for p in products:
            outlet_code = p.get("outlet_code", "")
            outlet_id = outlet_cache.get(outlet_code)

            if not outlet_id:
                continue

            article = p.get("article", "")
            if not article:
                continue

            # UPSERT в profi_nomenclature (НОРМАЛИЗОВАННЫЕ данные)
            cur.execute("""
                INSERT INTO profi_nomenclature (
                    article, name, category,
                    brand, model, part_type,
                    first_seen_at, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                ON CONFLICT (article) DO UPDATE SET
                    name = EXCLUDED.name,
                    category = EXCLUDED.category,
                    brand = EXCLUDED.brand,
                    model = EXCLUDED.model,
                    part_type = EXCLUDED.part_type,
                    updated_at = NOW()
                RETURNING id, (xmax = 0) as inserted
            """, (
                article,
                p.get("name", ""),
                p.get("category", ""),
                p.get("brand", ""),
                p.get("model"),
                p.get("part_type", "")
            ))

            row = cur.fetchone()
            nomenclature_id = row[0]
            if row[1]:
                nom_inserted += 1
            else:
                nom_updated += 1

            # UPSERT в profi_current_prices
            price = p.get("price", 0)
            in_stock = p.get("in_stock", False)
            product_url = f"https://siriust.ru/product/{article}"

            cur.execute("""
                INSERT INTO profi_current_prices (
                    nomenclature_id, outlet_id, price, in_stock,
                    product_url, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, NOW())
                ON CONFLICT (nomenclature_id, outlet_id) DO UPDATE SET
                    price = EXCLUDED.price,
                    in_stock = EXCLUDED.in_stock,
                    product_url = EXCLUDED.product_url,
                    updated_at = NOW()
            """, (nomenclature_id, outlet_id, price, in_stock, product_url))
            prices_upserted += 1

        conn.commit()

        print(f"\n{'='*60}")
        print(f"СОХРАНЕНО В БД:")
        print(f"  profi_nomenclature: +{nom_inserted} новых, ~{nom_updated} обновлено")
        print(f"  profi_current_prices: {prices_upserted} записей")

        # Итоговая статистика
        cur.execute("SELECT COUNT(*) FROM profi_nomenclature")
        total_nom = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM profi_current_prices WHERE in_stock = true")
        total_in_stock = cur.fetchone()[0]
        print(f"\nИтого в БД: {total_nom} товаров, {total_in_stock} в наличии")
        print(f"{'='*60}")

    finally:
        cur.close()
        conn.close()


def main():
    parser = argparse.ArgumentParser(description='Парсер Profi с встроенной нормализацией')
    parser.add_argument('--all', action='store_true',
                       help='Полный парсинг всех городов + сохранение в БД')
    parser.add_argument('--dynamic', action='store_true',
                       help='Динамически получить список прайс-листов с сайта')
    parser.add_argument('--no-db', action='store_true',
                       help='Не сохранять в БД (только JSON)')
    parser.add_argument('--city', '-c', type=str, default=None,
                       help='Парсить только указанный город')
    parser.add_argument('--test', action='store_true',
                       help='Тестовый режим: 1 город')
    args = parser.parse_args()

    # Парсинг
    profi_parser = ProfiParserClean()

    if args.test:
        # Тестовый режим: парсим только Москву
        price_lists = [pl for pl in PRICE_LISTS if 'москва' in pl["city"].lower()][:1]
        if price_lists:
            print(f"Тестовый режим: парсим {price_lists[0]['city']}")
            profi_parser.parse_single_outlet(price_lists[0])
    elif args.city:
        # Парсим только один город
        city_lower = args.city.lower()
        price_lists = [pl for pl in PRICE_LISTS if pl["city"].lower() == city_lower]
        if not price_lists:
            print(f"Город '{args.city}' не найден. Доступные города:")
            for city in sorted(set(pl["city"] for pl in PRICE_LISTS)):
                print(f"  - {city}")
            return

        print(f"Парсинг города: {args.city}")
        for pl in price_lists:
            profi_parser.parse_single_outlet(pl)
    else:
        # Парсим все города
        profi_parser.parse_all_outlets(use_dynamic=args.dynamic)

    profi_parser.print_stats()
    profi_parser.save_to_json()

    # Сохранение в БД
    if not args.no_db and profi_parser.products:
        save_to_db(profi_parser.products, profi_parser.outlets_parsed)

    print("\nПарсинг завершён!")


if __name__ == "__main__":
    main()
