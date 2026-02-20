#!/usr/bin/env python3
"""
Парсер прайс-листа Profi (Astraxan.xlsx)
Извлекает иерархию категорий по размеру шрифта:
  11px = Бренд
  10px = Модель
  9px  = Тип запчасти
"""
import sys
import time
import re
import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Пути по умолчанию
DEFAULT_SRC = "/data/in/Astraxan.xlsx"
DEFAULT_OUT = "/data/out/Astraxan_profi.csv"


def canon(s: str) -> str:
    """Канонизация строки для сравнения"""
    s = s.strip().lower().replace("ё", "е")
    return re.sub(r"[^a-z0-9а-я]+", "", s)


def cell_font_size(cell):
    """Получить размер шрифта ячейки"""
    try:
        return int(round(float(cell.font.sz))) if cell and cell.font and cell.font.sz else None
    except:
        return None


def find_header_row(ws):
    """Найти строку заголовка (содержит 'наимен')"""
    for r in range(1, min(ws.max_row + 1, 50)):
        for c in range(1, min(ws.max_column, 50) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "наимен" in v.lower():
                return r
    return None


def row_empty_fast(ws, r, cols):
    """Быстрая проверка пустой строки по ключевым колонкам"""
    for idx in cols:
        if idx:
            v = ws.cell(r, idx).value
            if v not in (None, "", " "):
                return False
    return True


def parse_price(src_path: str, out_path: str):
    """Основная функция парсинга"""
    start_all = time.time()

    SRC = Path(src_path)
    OUT = Path(out_path)

    if not SRC.exists():
        print(f"ERROR: File not found: {SRC}", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] Start. File: {SRC}")

    # Открываем с поддержкой стилей
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    print(f"[INFO] Workbook opened in {time.time()-start_all:.3f}s")

    # Ищем заголовок
    header_row = find_header_row(ws)
    if not header_row:
        print("ERROR: Header row with 'Наименование' not found", file=sys.stderr)
        sys.exit(1)
    print(f"[INFO] Header row: {header_row}")

    # Карта заголовков
    headers_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            headers_map[v.strip()] = c
    canon_map = {canon(k): v for k, v in headers_map.items()}

    # Синонимы колонок
    targets = {
        "Наименование": ["наимен", "наименование", "name"],
        "Артикул": ["артикул", "код", "code", "sku"],
        "Штрихкод": ["штрихкод", "штрих-код", "barcode", "ean"],
        "Цена": ["цена", "розница", "retail", "price"],
        "Ед.": ["ед", "единица", "unit"],
        "Количество": ["количество", "наличие", "остаток", "stock", "qty"],
        "Склад": ["склад", "warehouse"],
        "Примечание": ["примечание", "комментарий", "comment", "note"],
    }

    # Резолвим колонки
    resolved = {}
    for target, keys in targets.items():
        found = None
        for key in keys:
            for hdr_canon, col in canon_map.items():
                if key in hdr_canon:
                    found = col
                    break
            if found:
                break
        resolved[target] = found

    name_col = resolved["Наименование"]
    if not name_col:
        print("ERROR: Column 'Наименование' not found", file=sys.stderr)
        sys.exit(1)

    print(f"[INFO] Resolved columns: {resolved}")

    product_order = ["Артикул", "Штрихкод", "Цена", "Ед.", "Количество", "Склад", "Примечание"]
    product_cols = [resolved.get(t) for t in product_order]
    empty_check_cols = [name_col] + [c for c in product_cols if c]

    out_headers = ["Наименование"] + product_order + ["Признак бренда", "Признак модели", "Тип запчасти"]

    # Парсим данные
    current_brand = current_model = current_type = None
    rows_out = []

    t0 = time.time()
    for r in range(header_row + 1, ws.max_row + 1):
        if row_empty_fast(ws, r, empty_check_cols):
            print(f"[INFO] Stop at row {r} (empty)")
            break

        name_cell = ws.cell(r, name_col)
        name_val = str(name_cell.value).strip() if name_cell.value else None
        if not name_val:
            continue

        fsize = cell_font_size(name_cell)

        # Иерархия по размеру шрифта
        if fsize == 11:
            current_brand = name_val
            continue
        elif fsize == 10:
            current_model = name_val
            continue
        elif fsize == 9:
            current_type = name_val
            continue
        else:
            # Это товар
            prod_values = [(ws.cell(r, idx).value if idx else None) for idx in product_cols]
            rows_out.append([name_val, *prod_values, current_brand, current_model, current_type])

        if len(rows_out) % 1000 == 0:
            print(f"[PROGRESS] {len(rows_out)} rows...")

    print(f"[INFO] Parsed {len(rows_out)} products in {time.time()-t0:.3f}s")

    # Записываем CSV
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(out_headers)
        w.writerows(rows_out)

    print(f"[OK] Saved: {OUT} ({len(rows_out)} rows)")
    print(f"[INFO] Total time: {time.time()-start_all:.3f}s")

    return len(rows_out)


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT
    parse_price(src, out)
